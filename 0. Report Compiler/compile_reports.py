"""Compile 3 source files per property into a single Financial Report workbook.

Usage:
    python compile_reports.py --report-month 2025-11

Inputs (from ./Input/):
    - 12_Month_Statement_<code>_Accrual.xlsx
    - Cash_Flow_<code>_Accrual.xlsx
    - GeneralLedger_<code>_Accrual.xlsx

Output (to ./Output/):
    - SMS-<PropertyName> <MM.DD.YY> Financial Report_Final.xlsx
"""
from __future__ import annotations

import argparse
import calendar
import re
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

DEBUG = False


def debug_print(msg: str) -> None:
    if DEBUG:
        print(f"[DEBUG] {msg}")


# Property code mappings
PROPERTY_CONFIG: dict[str, str] = {
    "smsaltoo": "Altoona",
    "smscary": "Cary",
    "smscrys": "Crystal Lake",
    "smsnfss": "NFSS",
}

# Sheet type to output sheet name mapping
SHEET_MAPPINGS: dict[str, str] = {
    "GeneralLedger": "General Ledger",
    "Cash_Flow": "Cash Flow",
    "12_Month_Statement": "12 Month Statement",
}

# Order of sheets in output workbook
SHEET_ORDER = ["General Ledger", "Cash Flow", "12 Month Statement"]


@dataclass
class PropertyFiles:
    """Container for the 3 source files of a property."""
    code: str
    general_ledger: Path | None = None
    cash_flow: Path | None = None
    statement_12m: Path | None = None

    @property
    def property_name(self) -> str:
        return PROPERTY_CONFIG.get(self.code, self.code.upper())

    @property
    def is_complete(self) -> bool:
        return all([self.general_ledger, self.cash_flow, self.statement_12m])


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Compile source files into Financial Report workbooks.",
    )
    parser.add_argument(
        "--input-folder",
        type=Path,
        default=Path("./Input"),
        help="Folder containing source files (default: ./Input)",
    )
    parser.add_argument(
        "--output-folder",
        type=Path,
        default=Path("./Output"),
        help="Folder for compiled reports (default: ./Output)",
    )
    parser.add_argument(
        "--report-month",
        required=True,
        help="Reporting month in YYYY-MM format (e.g., 2025-11)",
    )
    return parser.parse_args()


def extract_property_code(filename: str) -> str | None:
    """Extract property code from filename like 'Cash_Flow_smsaltoo_Accrual.xlsx'."""
    for code in PROPERTY_CONFIG:
        if code in filename.lower():
            return code
    return None


def get_file_type(filename: str) -> str | None:
    """Determine file type from filename."""
    name_lower = filename.lower()
    if name_lower.startswith("generalledger"):
        return "GeneralLedger"
    elif name_lower.startswith("cash_flow"):
        return "Cash_Flow"
    elif name_lower.startswith("12_month_statement"):
        return "12_Month_Statement"
    return None


def scan_input_files(input_folder: Path) -> dict[str, PropertyFiles]:
    """Scan input folder and group files by property."""
    properties: dict[str, PropertyFiles] = {}

    for file in input_folder.glob("*.xlsx"):
        code = extract_property_code(file.name)
        file_type = get_file_type(file.name)

        if not code or not file_type:
            debug_print(f"Skipping unrecognized file: {file.name}")
            continue

        if code not in properties:
            properties[code] = PropertyFiles(code=code)

        prop = properties[code]
        if file_type == "GeneralLedger":
            prop.general_ledger = file
        elif file_type == "Cash_Flow":
            prop.cash_flow = file
        elif file_type == "12_Month_Statement":
            prop.statement_12m = file

        debug_print(f"Found {file_type} for {code}: {file.name}")

    return properties


def copy_cell_style(src: Cell, dst: Cell) -> None:
    """Copy cell styling from source to destination."""
    if src.has_style:
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)
        dst.alignment = copy(src.alignment)


def copy_sheet(src_sheet: Worksheet, dst_sheet: Worksheet) -> None:
    """Copy all content and formatting from source to destination sheet."""
    # Copy cell values and styles
    for row in src_sheet.iter_rows():
        for cell in row:
            dst_cell = dst_sheet.cell(row=cell.row, column=cell.column)
            dst_cell.value = cell.value
            copy_cell_style(cell, dst_cell)

    # Copy column dimensions
    for col_letter, dim in src_sheet.column_dimensions.items():
        dst_sheet.column_dimensions[col_letter].width = dim.width
        dst_sheet.column_dimensions[col_letter].hidden = dim.hidden

    # Copy row dimensions
    for row_num, dim in src_sheet.row_dimensions.items():
        dst_sheet.row_dimensions[row_num].height = dim.height
        dst_sheet.row_dimensions[row_num].hidden = dim.hidden

    # Copy merged cells
    for merged_range in src_sheet.merged_cells.ranges:
        dst_sheet.merge_cells(str(merged_range))


def compile_property(
    prop: PropertyFiles,
    output_path: Path,
) -> None:
    """Compile a property's source files into a single workbook."""
    debug_print(f"Compiling {prop.property_name}...")

    # Create new workbook (remove default sheet later)
    output_wb = Workbook()
    default_sheet = output_wb.active

    # Source file mapping to output sheet names
    sources = [
        (prop.general_ledger, "General Ledger"),
        (prop.cash_flow, "Cash Flow"),
        (prop.statement_12m, "12 Month Statement"),
    ]

    for src_path, sheet_name in sources:
        if not src_path:
            print(f"  [WARN] Missing {sheet_name} for {prop.property_name}")
            continue

        debug_print(f"  Copying {src_path.name} -> {sheet_name}")
        src_wb = load_workbook(src_path)
        src_sheet = src_wb.active  # All inputs have 'Report1' as single sheet

        # Create destination sheet
        dst_sheet = output_wb.create_sheet(title=sheet_name)
        copy_sheet(src_sheet, dst_sheet)

        src_wb.close()

    # Remove the default empty sheet
    if default_sheet.title not in SHEET_ORDER:
        output_wb.remove(default_sheet)

    # Save output
    output_wb.save(output_path)
    output_wb.close()
    print(f"[OK] {prop.property_name} -> {output_path.name}")


def generate_output_filename(property_name: str, year: int, month: int) -> str:
    """Generate output filename with last day of month."""
    last_day = calendar.monthrange(year, month)[1]
    date_str = f"{month:02d}.{last_day:02d}.{year % 100:02d}"
    return f"SMS-{property_name} {date_str} Financial Report_Final.xlsx"


def main() -> None:
    args = parse_args()

    input_folder = args.input_folder.resolve()
    output_folder = args.output_folder.resolve()

    if not input_folder.exists():
        raise FileNotFoundError(f"Input folder not found: {input_folder}")

    output_folder.mkdir(parents=True, exist_ok=True)

    # Parse report month
    year, month = map(int, args.report_month.split("-"))

    # Scan and group input files
    properties = scan_input_files(input_folder)

    if not properties:
        print(f"No source files found in {input_folder}")
        return

    print(f"Found {len(properties)} properties to compile\n")

    # Compile each property
    successes = 0
    for code, prop in sorted(properties.items()):
        if not prop.is_complete:
            missing = []
            if not prop.general_ledger:
                missing.append("GeneralLedger")
            if not prop.cash_flow:
                missing.append("Cash_Flow")
            if not prop.statement_12m:
                missing.append("12_Month_Statement")
            print(f"[SKIP] {prop.property_name}: Missing {', '.join(missing)}")
            continue

        output_name = generate_output_filename(prop.property_name, year, month)
        output_path = output_folder / output_name

        try:
            compile_property(prop, output_path)
            successes += 1
        except Exception as exc:
            print(f"[ERROR] {prop.property_name}: {exc}")

    print(f"\nCompiled {successes}/{len(properties)} properties")


if __name__ == "__main__":
    main()
