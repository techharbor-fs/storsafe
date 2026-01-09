"""Prepare a monthly Property Cash Flow workbook.

Implements the workflow:
1) Prompt month folder (e.g. "11. Nov")
2) Prompt target spreadsheet link/ID
3) Copy template tabs from the Templates workbook into the target:
   - PORTFOLIO CASH FLOW (tab color: blue)
   - PROPERTY CODES (tab color: black, hidden)
   - PROPERTY STATUS (tab color: black, hidden)
4) Copy the Month End list tab from the Month End workbook into the target:
   - Bank Rec priorities -> MONTH END LIST (tab color: black, hidden)
5) Import Excel exports with formatting (like File > Import in Google Sheets):
   - HA-CF-<MON> (tab color: red)
   - HA-CF-3MOS (tab color: red)
   - HA-CF-YTD (tab color: red)
   - HA-BS-<MON> (tab color: red)
6) Apply minor tweaks: freeze panes

Notes:
- Excel formatting is preserved by uploading via Drive API as Google Sheets.
- No dependency on past months - each run is self-contained.
- Requires the target spreadsheet to be shared with the service account.

Usage:
  python prepare_monthly_cashflow_workbook.py
  python prepare_monthly_cashflow_workbook.py --month-folder "11. Nov" --target-sheet <link_or_id> --confirm

"""

from __future__ import annotations

import argparse
import json
import os
import re
import tempfile
from dataclasses import dataclass
from pathlib import Path

try:
    from dotenv import load_dotenv

    load_dotenv()
except Exception:
    pass

import gspread
import openpyxl
import openpyxl.utils
import time
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from gspread.exceptions import APIError


def _retry_on_rate_limit(func, *args, max_retries: int = 3, base_delay: float = 30.0, **kwargs):
    """Retry a function call on rate limit (429) errors with exponential backoff."""
    for attempt in range(max_retries + 1):
        try:
            return func(*args, **kwargs)
        except APIError as e:
            if "429" in str(e) and attempt < max_retries:
                delay = base_delay * (2 ** attempt)
                print(f"    [Rate limited] Waiting {delay:.0f}s before retry {attempt + 1}/{max_retries}...")
                time.sleep(delay)
            else:
                raise


# Source workbooks
TEMPLATE_BOOK_ID = "1_27gMJWp2xjql24akf4CUa-7Qm1Ng_I0hb0kT9P_PbE"
TEMPLATE_TABS = ["PORTFOLIO CASH FLOW", "PROPERTY CODES", "PROPERTY STATUS"]

MONTH_END_BOOK_ID = "1_Z_gMcpFYgOdsRHWvnTk-LuLo0dR_TAWGSPWYTJd5FA"
MONTH_END_TAB = "Bank Rec priorities"
MONTH_END_TARGET_TAB = "MONTH END LIST"

# Defaults
# This script lives under internal/. Month folders (e.g. "11. Nov") live one level up.
DEFAULT_WORKDIR = Path(__file__).resolve().parent.parent
DEFAULT_TARGET_SHEET_ID = "160onO2dxp7fewgcibKoQ9biFH0qW69wg215J9phMVSs"

# Project root is the repo root (one level above the Cash Flow Report 2 folder)
PROJECT_ROOT = DEFAULT_WORKDIR.parent
STATE_PATH = PROJECT_ROOT / "data" / "output" / "cashflow_prep_state.json"

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]


# Tab colors (Google API uses 0..1 floats)
COLOR_BLUE = {"red": 0.0, "green": 0.0, "blue": 1.0}
COLOR_BLACK = {"red": 0.0, "green": 0.0, "blue": 0.0}
COLOR_RED = {"red": 1.0, "green": 0.0, "blue": 0.0}


MONTH_NAMES = {
    "JAN": "January",
    "FEB": "February",
    "MAR": "March",
    "APR": "April",
    "MAY": "May",
    "JUN": "June",
    "JUL": "July",
    "AUG": "August",
    "SEP": "September",
    "OCT": "October",
    "NOV": "November",
    "DEC": "December",
}


def _add_months(year: int, month: int, delta_months: int) -> tuple[int, int]:
    # month is 1..12
    new_month_index = (year * 12 + (month - 1)) + delta_months
    new_year = new_month_index // 12
    new_month = (new_month_index % 12) + 1
    return new_year, new_month


def _month_num_from_abbrev(mon3: str) -> int | None:
    mon3 = (mon3 or "").strip().upper()
    mapping = {
        "JAN": 1,
        "FEB": 2,
        "MAR": 3,
        "APR": 4,
        "MAY": 5,
        "JUN": 6,
        "JUL": 7,
        "AUG": 8,
        "SEP": 9,
        "OCT": 10,
        "NOV": 11,
        "DEC": 12,
    }
    return mapping.get(mon3)


def _parse_ha_cf_period_label(text: str) -> tuple[int, int] | None:
    # Examples seen in HA-CF tab:
    # "Period = Nov 2025" (case-insensitive)
    # "Period = NOV 2025"
    m = re.search(r"\bperiod\s*=\s*([A-Za-z]{3})\s+(\d{4})\b", text or "", flags=re.IGNORECASE)
    if not m:
        return None
    mon3 = m.group(1).upper()
    year = int(m.group(2))
    month = _month_num_from_abbrev(mon3)
    if not month:
        return None
    return year, month


def _month_name_year(year: int, month: int) -> str:
    inv = {v: k for k, v in {"JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6, "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12}.items()}
    mon3 = inv[month]
    return f"{MONTH_NAMES[mon3]} {year}"


def _range_label_last_3_months(end_year: int, end_month: int) -> str:
    start_year, start_month = _add_months(end_year, end_month, -2)
    return f"{_month_name_year(start_year, start_month)} - {_month_name_year(end_year, end_month)}"


def _range_label_trailing_12(end_year: int, end_month: int) -> str:
    # Trailing 12 months inclusive: start is end_month minus 11.
    start_year, start_month = _add_months(end_year, end_month, -11)
    return f"{_month_name_year(start_year, start_month)} - {_month_name_year(end_year, end_month)}"


def _normalize_title(title: str) -> str:
    return re.sub(r"\s+", " ", (title or "").strip()).upper()


def _worksheet_by_title_ci(sheet: gspread.Spreadsheet, title: str) -> gspread.Worksheet | None:
    want = _normalize_title(title)
    for ws in sheet.worksheets():
        if _normalize_title(ws.title) == want:
            return ws
    return None


def _batch_update_with_retry(sheet: gspread.Spreadsheet, body: dict, max_retries: int = 3) -> None:
    """Call batch_update with retry on rate limit errors."""
    for attempt in range(max_retries + 1):
        try:
            sheet.batch_update(body)
            return
        except APIError as e:
            if "429" in str(e) and attempt < max_retries:
                delay = 30 * (2 ** attempt)
                print(f"    [Rate limited] Waiting {delay}s before retry...")
                time.sleep(delay)
            else:
                raise


def _set_tab_properties(sheet: gspread.Spreadsheet, worksheet: gspread.Worksheet, *, tab_color: dict | None = None, hidden: bool | None = None, index: int | None = None) -> None:
    props: dict = {"sheetId": worksheet.id}
    fields: list[str] = []
    if tab_color is not None:
        props["tabColor"] = tab_color
        fields.append("tabColor")
    if hidden is not None:
        props["hidden"] = bool(hidden)
        fields.append("hidden")
    if index is not None:
        props["index"] = int(index)
        fields.append("index")
    if not fields:
        return
    _batch_update_with_retry(
        sheet,
        {
            "requests": [
                {
                    "updateSheetProperties": {
                        "properties": props,
                        "fields": ",".join(fields),
                    }
                }
            ]
        }
    )


def _set_freeze(sheet: gspread.Spreadsheet, worksheet: gspread.Worksheet, *, frozen_rows: int, frozen_cols: int) -> None:
    _batch_update_with_retry(
        sheet,
        {
            "requests": [
                {
                    "updateSheetProperties": {
                        "properties": {
                            "sheetId": worksheet.id,
                            "gridProperties": {
                                "frozenRowCount": int(frozen_rows),
                                "frozenColumnCount": int(frozen_cols),
                            },
                        },
                        "fields": "gridProperties.frozenRowCount,gridProperties.frozenColumnCount",
                    }
                }
            ]
        }
    )


def _update_portfolio_headers_from_period(
    sheet: gspread.Spreadsheet,
    *,
    portfolio_ws_title: str,
    period_year: int,
    period_month: int,
    scan_rows: int = 10,
) -> None:
    ws = _worksheet_by_title_ci(sheet, portfolio_ws_title)
    if not ws:
        return
    values = ws.get_values(f"A1:Z{scan_rows}")
    targets: list[tuple[int, int, str]] = []
    for r_idx, row in enumerate(values, start=1):
        for c_idx, cell in enumerate(row, start=1):
            text = (cell or "").strip()
            if not text:
                continue
            up = text.upper()
            if "| LAST MONTH" in up:
                targets.append((r_idx, c_idx, "LAST_MONTH"))
            elif "| LAST 3 MONTHS" in up:
                targets.append((r_idx, c_idx, "LAST_3"))
            elif "| TRAILING 12 MONTHS" in up:
                targets.append((r_idx, c_idx, "T12"))

    if not targets:
        return

    month_label = _month_name_year(period_year, period_month)
    last_3_label = _range_label_last_3_months(period_year, period_month)
    t12_label = _range_label_trailing_12(period_year, period_month)

    updates: list[dict] = []
    for r_idx, c_idx, kind in targets:
        a1 = gspread.utils.rowcol_to_a1(r_idx, c_idx)
        old = (values[r_idx - 1][c_idx - 1] or "").strip()
        # Preserve the suffix exactly as-is after the first '|'
        suffix = "|" + old.split("|", 1)[1] if "|" in old else ""
        if kind == "LAST_MONTH":
            new_value = f"{month_label} {suffix}".strip()
        elif kind == "LAST_3":
            new_value = f"{last_3_label} {suffix}".strip()
        else:
            new_value = f"{t12_label} {suffix}".strip()
        updates.append({"range": f"'{portfolio_ws_title}'!{a1}", "values": [[new_value]]})

    sheet.values_batch_update(
        {
            "valueInputOption": "RAW",
            "data": updates,
        }
    )


def _sheet_url(sheet_id: str) -> str:
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}"


def _extract_sheet_id(value: str) -> str:
    value = (value or "").strip()
    if not value:
        return value
    marker = "/spreadsheets/d/"
    if marker in value:
        after = value.split(marker, 1)[1]
        return after.split("/", 1)[0]
    return value


def _load_state() -> dict:
    try:
        return json.loads(STATE_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _save_state(state: dict) -> None:
    STATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    STATE_PATH.write_text(json.dumps(state, indent=2, sort_keys=True), encoding="utf-8")


def _resolve_service_account_file() -> str:
    service_account_json = os.environ.get("SERVICE_ACCOUNT_JSON")
    if service_account_json:
        temp_json_path = Path(tempfile.gettempdir()) / "service_account.json"
        temp_json_path.write_text(json.dumps(json.loads(service_account_json)), encoding="utf-8")
        return str(temp_json_path)

    env_path = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS") or os.environ.get("SERVICE_ACCOUNT_FILE")
    if not env_path:
        raise RuntimeError(
            "Missing Google service account credentials. Set SERVICE_ACCOUNT_JSON or "
            "GOOGLE_APPLICATION_CREDENTIALS/SERVICE_ACCOUNT_FILE."
        )
    return env_path


def _service_account_email(creds_path: str) -> str | None:
    try:
        data = json.loads(Path(creds_path).read_text(encoding="utf-8"))
    except Exception:
        return None
    email = data.get("client_email")
    return str(email).strip() if email else None


def _open_sheet_or_exit(gc: gspread.Client, *, sheet_id: str, label: str, service_account_email: str | None) -> gspread.Spreadsheet:
    max_retries = 3
    base_delay = 30.0

    for attempt in range(max_retries + 1):
        try:
            return gc.open_by_key(sheet_id)
        except PermissionError:
            email_msg = f" Share it with: {service_account_email}" if service_account_email else ""
            raise SystemExit(
                "\n".join(
                    [
                        f"Permission denied opening {label} spreadsheet.",
                        f"sheet_id: {sheet_id}",
                        f"url: {_sheet_url(sheet_id)}",
                        f"Fix: Share this spreadsheet with the service account.{email_msg}",
                    ]
                )
            )
        except APIError as e:
            if "429" in str(e) and attempt < max_retries:
                delay = base_delay * (2 ** attempt)
                print(f"  [Rate limited opening {label}] Waiting {delay:.0f}s before retry {attempt + 1}/{max_retries}...")
                time.sleep(delay)
            else:
                raise SystemExit(
                    "\n".join(
                        [
                            f"Failed opening {label} spreadsheet.",
                            f"sheet_id: {sheet_id}",
                            f"url: {_sheet_url(sheet_id)}",
                            f"error: {type(e).__name__}: {e}",
                        ]
                    )
                )
        except Exception as exc:
            raise SystemExit(
                "\n".join(
                    [
                        f"Failed opening {label} spreadsheet.",
                        f"sheet_id: {sheet_id}",
                        f"url: {_sheet_url(sheet_id)}",
                        f"error: {type(exc).__name__}: {exc}",
                    ]
                )
            )

    # Shouldn't reach here but just in case
    raise SystemExit(f"Failed to open {label} spreadsheet after {max_retries} retries")


@dataclass(frozen=True)
class TabSpec:
    source_book_id: str
    source_tab: str
    target_tab: str
    tab_color: dict
    hidden: bool


def _list_month_folders(workdir: Path) -> list[str]:
    # Expect folders like "11. Nov", "10. Oct", etc.
    out: list[str] = []
    for child in workdir.iterdir():
        if child.is_dir() and re.match(r"^\d{1,2}\.\s+", child.name):
            out.append(child.name)
    out.sort()
    return out


def _month_abbrev_from_folder(folder_name: str) -> str:
    # "11. Nov" -> "NOV"; "10. Oct" -> "OCT"
    parts = folder_name.split(".", 1)
    if len(parts) == 2:
        rest = parts[1].strip()
        if rest:
            return rest[:3].upper()
    return "MON"


def _load_excel_values(xlsx_path: Path, sheet_name: str | None = None) -> list[list[object]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _table_shape(table: list[list[object]]) -> tuple[int, int]:
    rows = len(table)
    cols = max((len(r) for r in table), default=0)
    return rows, cols


def _is_balance_sheet_export(table: list[list[object]]) -> bool:
    """Check if a table looks like a Balance Sheet export (vs Cash Flow).

    Heuristic: Balance Sheet exports typically contain:
    - "Balance Sheet" or "HAB" in early rows
    - Asset/Liability account codes (1xxx, 2xxx) without income/expense (4xxx-8xxx)
    """
    # Check first 20 rows for Balance Sheet indicators
    for row in table[:20]:
        for cell in row:
            text = str(cell or "").lower()
            if "balance sheet" in text or "hab s" in text or "ha-bs" in text:
                return True

    # Check if it has Asset/Liability accounts but no Income/Expense accounts
    # by scanning the first column for GL code patterns
    has_assets = False
    has_income_expense = False
    for row in table[:200]:
        if not row:
            continue
        first_cell = str(row[0] or "").strip()
        # GL code pattern: ####-####
        if re.match(r"^\d{4}-\d{4}$", first_cell):
            first_digit = first_cell[0]
            if first_digit in ("1", "2", "3"):
                has_assets = True
            elif first_digit in ("4", "5", "6", "7", "8"):
                has_income_expense = True

    # Balance Sheet has assets/liabilities but no income/expense accounts
    if has_assets and not has_income_expense:
        return True

    return False


def _detect_exports(month_dir: Path) -> dict[str, Path]:
    """Detect NOV/3MOS/YTD/BS exports in the month folder.

    Opens any .xlsx file and classifies by content:
    - Balance Sheet vs Cash Flow (by GL code patterns)
    - For Cash Flow: NOV/3MOS/YTD by row count (smallest=NOV, middle=3MOS, largest=YTD)

    No filename pattern assumptions - purely content-based.
    """

    # Find all xlsx files in the folder
    all_files = sorted(
        [p for p in month_dir.glob("*.xlsx") if p.is_file()],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )

    if len(all_files) < 3:
        raise SystemExit(
            f"Could not find at least 3 .xlsx files in {month_dir} (found {len(all_files)})."
        )

    # Load tables and classify
    tables: list[tuple[Path, list[list[object]], bool]] = []  # (path, table, is_bs)
    for p in all_files:
        table = _load_excel_values(p)
        is_bs = _is_balance_sheet_export(table)
        tables.append((p, table, is_bs))

    # Separate Balance Sheet from Cash Flow
    bs_files = [(p, t) for p, t, is_bs in tables if is_bs]
    cf_files = [(p, t) for p, t, is_bs in tables if not is_bs]

    # Sort Cash Flow files by row count
    cf_sorted = sorted(cf_files, key=lambda item: _table_shape(item[1])[0])

    if len(cf_sorted) < 3:
        raise SystemExit(
            f"Could not find 3 Cash Flow exports in {month_dir}. Found {len(cf_sorted)} CF files and {len(bs_files)} BS files."
        )

    result = {
        "NOV": cf_sorted[0][0],
        "3MOS": cf_sorted[1][0],
        "YTD": cf_sorted[2][0],
    }

    if bs_files:
        result["BS"] = bs_files[0][0]

    return result


def _make_unique_local_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    parent = path.parent
    i = 2
    while True:
        candidate = parent / f"{stem} ({i}){suffix}"
        if not candidate.exists():
            return candidate
        i += 1


def _infer_year_from_filename(path: Path) -> int | None:
    years = re.findall(r"\b(20\d{2})\b", path.name)
    if not years:
        return None
    try:
        return int(years[-1])
    except Exception:
        return None


def _rename_exports_in_place(
    exports: dict[str, Path],
    *,
    month_abbrev: str,
    dry_run: bool,
) -> dict[str, Path]:
    """Rename detected exports to deterministic filenames (in the same folder).

    Goal: stabilize downstream processing by ensuring consistent filenames in each month folder.
    Safety: only renames when not in dry-run mode.
    """

    mon = (month_abbrev or "MON").strip().lower()[:3] or "mon"
    any_year = (
        _infer_year_from_filename(exports.get("NOV", Path("")))
        or _infer_year_from_filename(exports.get("3MOS", Path("")))
        or _infer_year_from_filename(exports.get("YTD", Path("")))
        or _infer_year_from_filename(exports.get("BS", Path("")))
    )
    year_part = f"_{any_year}" if any_year else ""

    desired_names = {
        "NOV": f"yardi_property_comparison_allharda_accrual_{mon}{year_part}.xlsx",
        "3MOS": f"yardi_property_comparison_allharda_accrual_{mon}{year_part}_3mos.xlsx",
        "YTD": f"yardi_property_comparison_allharda_accrual_{mon}{year_part}_ytd.xlsx",
        "BS": f"yardi_balance_sheet_allharda_{mon}{year_part}.xlsx",
    }

    updated = dict(exports)
    print("\nExport rename plan (local files):")
    for key in ("NOV", "3MOS", "YTD", "BS"):
        src = exports.get(key)
        if not src:
            continue
        desired = src.parent / desired_names[key]
        if src.resolve() == desired.resolve():
            print(f"  {key}: (already named) {src.name}")
            continue
        if desired.exists() and desired.resolve() != src.resolve():
            desired = _make_unique_local_path(desired)
        if dry_run:
            print(f"  [DRY RUN] {key}: would rename {src.name} -> {desired.name}")
        else:
            print(f"  {key}: renaming {src.name} -> {desired.name}")
            src.rename(desired)
        updated[key] = desired

    return updated


def _try_get_period_from_ha_cf(target_ss: gspread.Spreadsheet, month_tab: str) -> tuple[int, int] | None:
    ws = _worksheet_by_title_ci(target_ss, month_tab)
    if not ws:
        return None
    scan = ws.get_values("A1:B8")
    for row in scan:
        for cell in row:
            parsed = _parse_ha_cf_period_label(str(cell or ""))
            if parsed:
                return parsed
    return None


def _finalize_target_workbook(
    target_ss: gspread.Spreadsheet,
    *,
    month_tab: str,
    bs_tab: str,
    dry_run: bool,
) -> None:
    if dry_run:
        return

    # Allow API quota to recover before finalization
    print("  Finalizing workbook (waiting for API quota)...")
    time.sleep(5)

    # Ensure expected naming for Month End tab.
    month_end_ws = _worksheet_by_title_ci(target_ss, MONTH_END_TAB)
    if month_end_ws and not _worksheet_by_title_ci(target_ss, MONTH_END_TARGET_TAB):
        month_end_ws.update_title(MONTH_END_TARGET_TAB)

    # Re-apply expected colors/hidden.
    portfolio_ws = _worksheet_by_title_ci(target_ss, "PORTFOLIO CASH FLOW")
    if portfolio_ws:
        _set_tab_properties(target_ss, portfolio_ws, tab_color=COLOR_BLUE, hidden=False)

    codes_ws = _worksheet_by_title_ci(target_ss, "PROPERTY CODES")
    if codes_ws:
        _set_tab_properties(target_ss, codes_ws, tab_color=COLOR_BLACK, hidden=True)

    month_end_ws2 = _worksheet_by_title_ci(target_ss, MONTH_END_TARGET_TAB) or _worksheet_by_title_ci(target_ss, MONTH_END_TAB)
    if month_end_ws2:
        _set_tab_properties(target_ss, month_end_ws2, tab_color=COLOR_BLACK, hidden=True)

    for title in (month_tab, "HA-CF-3MOS", "HA-CF-YTD", bs_tab):
        ws = _worksheet_by_title_ci(target_ss, title)
        if ws:
            _set_tab_properties(target_ss, ws, tab_color=COLOR_RED, hidden=False)

    # Freeze panes: match your example (HA-CF freeze at row 6, col B).
    for title in (month_tab, "HA-CF-3MOS", "HA-CF-YTD", bs_tab):
        ws = _worksheet_by_title_ci(target_ss, title)
        if ws:
            _set_freeze(target_ss, ws, frozen_rows=6, frozen_cols=2)

    # Portfolio freeze: keep headers and the label columns visible.
    if portfolio_ws:
        _set_freeze(target_ss, portfolio_ws, frozen_rows=4, frozen_cols=3)

    # Add column groupings in PORTFOLIO CASH FLOW (matches your screenshot brackets).
    if portfolio_ws:
        meta = target_ss.fetch_sheet_metadata()
        sheet_entry = None
        for s in meta.get("sheets", []):
            props = s.get("properties", {})
            if int(props.get("sheetId", -1)) == int(portfolio_ws.id):
                sheet_entry = s
                break

        existing_groups: list[dict] = []
        if sheet_entry:
            existing_groups = sheet_entry.get("dimensionGroups", []) or []

        def has_group(start_idx: int, end_idx: int) -> bool:
            for g in existing_groups:
                r = (g.get("range") or {})
                if r.get("dimension") != "COLUMNS":
                    continue
                if int(r.get("sheetId", -1)) != int(portfolio_ws.id):
                    continue
                if int(r.get("startIndex", -999)) == start_idx and int(r.get("endIndex", -999)) == end_idx:
                    return True
            return False

        group_ranges = [
            # D:I (Last Month metrics)
            (3, 9),
            # K:P (Last 3 Months metrics)
            (10, 16),
            # R:W (Trailing 12 Months metrics)
            (17, 23),
        ]

        requests: list[dict] = []
        for start_idx, end_idx in group_ranges:
            if has_group(start_idx, end_idx):
                continue
            requests.append(
                {
                    "addDimensionGroup": {
                        "range": {
                            "sheetId": int(portfolio_ws.id),
                            "dimension": "COLUMNS",
                            "startIndex": int(start_idx),
                            "endIndex": int(end_idx),
                        }
                    }
                }
            )

        if requests:
            target_ss.batch_update({"requests": requests})

    # Update PORTFOLIO header date ranges based on the HA-CF period label.
    parsed = _try_get_period_from_ha_cf(target_ss, month_tab)
    if parsed:
        year, month = parsed
        _update_portfolio_headers_from_period(
            target_ss,
            portfolio_ws_title="PORTFOLIO CASH FLOW",
            period_year=year,
            period_month=month,
        )

    # Enforce tab order.
    desired_titles = [
        "PORTFOLIO CASH FLOW",
        month_tab,
        "HA-CF-3MOS",
        "HA-CF-YTD",
        bs_tab,
        "PROPERTY CODES",
        MONTH_END_TARGET_TAB,
        "PROPERTY STATUS",
    ]
    title_to_ws: dict[str, gspread.Worksheet] = {}
    for ws in target_ss.worksheets():
        norm = _normalize_title(ws.title)
        for desired in desired_titles:
            if norm == _normalize_title(desired):
                title_to_ws[desired] = ws
                break

    for idx, title in enumerate(desired_titles):
        ws = title_to_ws.get(title)
        if ws:
            _set_tab_properties(target_ss, ws, index=idx)


def _find_sheet_id_by_title(ss: gspread.Spreadsheet, title: str) -> int | None:
    def key(v: object) -> str:
        return re.sub(r"\s+", " ", str(v)).strip().upper()

    want = key(title)
    meta = ss.fetch_sheet_metadata()
    for s in meta.get("sheets", []):
        props = s.get("properties", {})
        if key(props.get("title", "")) == want:
            return int(props.get("sheetId"))
    return None


def _existing_titles(ss: gspread.Spreadsheet) -> set[str]:
    meta = ss.fetch_sheet_metadata()
    titles: set[str] = set()
    for s in meta.get("sheets", []):
        props = s.get("properties", {})
        t = props.get("title")
        if t is not None:
            titles.add(str(t))
    return titles


def _norm_title(v: object) -> str:
    return re.sub(r"\s+", " ", str(v)).strip().upper()


def _cleanup_draft_tabs(dst_ss: gspread.Spreadsheet, *, dry_run: bool) -> None:
    """Remove tabs created as artifacts by prior failed runs.

    Only deletes very specific names that this automation may create.
    """

    meta = dst_ss.fetch_sheet_metadata()
    title_to_id: dict[str, int] = {}
    for s in meta.get("sheets", []):
        props = s.get("properties", {})
        title = str(props.get("title", ""))
        sheet_id = props.get("sheetId")
        if sheet_id is None:
            continue
        title_to_id[title] = int(sheet_id)

    # Delete "Copy of PROPERTY CODES*" (these came from earlier rename collisions)
    for title, sheet_id in list(title_to_id.items()):
        nt = _norm_title(title)
        if nt.startswith("COPY OF PROPERTY CODES"):
            if dry_run:
                print(f"  [DRY RUN] Would delete draft tab: {title}")
            else:
                print(f"  Deleting draft tab: {title}")
                _delete_sheet_id(dst_ss, sheet_id=sheet_id, dry_run=dry_run)

    # If we have the new month-end tab, remove the legacy one if present.
    has_bank_rec = any(_norm_title(t) == _norm_title(MONTH_END_TAB) for t in title_to_id.keys())
    if has_bank_rec:
        for title, sheet_id in list(title_to_id.items()):
            if _norm_title(title) == _norm_title("Month End Reports"):
                if dry_run:
                    print(f"  [DRY RUN] Would delete legacy tab: {title}")
                else:
                    print(f"  Deleting legacy tab: {title}")
                    _delete_sheet_id(dst_ss, sheet_id=sheet_id, dry_run=dry_run)


def _make_unique_title(ss: gspread.Spreadsheet, desired: str) -> str:
    existing = _existing_titles(ss)
    if desired not in existing:
        return desired
    i = 2
    while True:
        candidate = f"{desired} ({i})"
        if candidate not in existing:
            return candidate
        i += 1


def _rename_sheet_id(ss: gspread.Spreadsheet, *, sheet_id: int, new_title: str, dry_run: bool) -> None:
    body = {
        "requests": [
            {
                "updateSheetProperties": {
                    "properties": {"sheetId": int(sheet_id), "title": new_title},
                    "fields": "title",
                }
            }
        ]
    }
    if dry_run:
        return
    ss.batch_update(body)


def _delete_sheet_id(ss: gspread.Spreadsheet, *, sheet_id: int, dry_run: bool) -> None:
    body = {"requests": [{"deleteSheet": {"sheetId": int(sheet_id)}}]}
    if dry_run:
        return
    ss.batch_update(body)


def _rename_if_exists(ss: gspread.Spreadsheet, title: str, new_title: str, *, dry_run: bool) -> str | None:
    sheet_id = _find_sheet_id_by_title(ss, title)
    if sheet_id is None:
        return None
    unique = _make_unique_title(ss, new_title)
    if dry_run:
        print(f"  [DRY RUN] Would rename tab: {title} -> {unique}")
        return unique
    print(f"  Renaming existing tab: {title} -> {unique}")
    _rename_sheet_id(ss, sheet_id=sheet_id, new_title=unique, dry_run=dry_run)
    return unique


def _delete_if_exists(ss: gspread.Spreadsheet, title: str, *, dry_run: bool) -> None:
    sheet_id = _find_sheet_id_by_title(ss, title)
    if sheet_id is None:
        return

    if dry_run:
        print(f"  [DRY RUN] Would delete tab: {title}")
        return

    print(f"  Deleting tab: {title}")
    _delete_sheet_id(ss, sheet_id=sheet_id, dry_run=dry_run)


def _delete_other_tabs(
    ss: gspread.Spreadsheet,
    *,
    keep_titles: set[str],
    dry_run: bool,
) -> None:
    """Delete all tabs not in keep_titles (case-insensitive match).

    Safety:
    - Always keeps the tabs listed in keep_titles.
    - If no keep tabs exist in the workbook, does nothing (avoids deleting everything by accident).
    """

    keep_norm = {_normalize_title(t) for t in keep_titles if str(t).strip()}
    meta = ss.fetch_sheet_metadata()
    sheets = meta.get("sheets", [])

    existing_keep = 0
    delete_candidates: list[tuple[str, int]] = []

    for s in sheets:
        props = s.get("properties", {})
        title = str(props.get("title", ""))
        sheet_id = int(props.get("sheetId"))
        if _normalize_title(title) in keep_norm:
            existing_keep += 1
            continue
        delete_candidates.append((title, sheet_id))

    if existing_keep == 0:
        print("\n⚠️  Skip deleting other tabs: none of the keep tabs exist yet.")
        return

    if not delete_candidates:
        print("\nNo extra tabs to delete.")
        return

    print("\nOther tabs to delete:")
    for title, sheet_id in delete_candidates:
        print(f"  - {title} (sheetId={sheet_id})")

    if dry_run:
        print("\n[DRY RUN] No tabs deleted. Re-run with --confirm to delete.")
        return

    print("\nDeleting other tabs...")
    for title, sheet_id in delete_candidates:
        print(f"  Deleting: {title}")
        _delete_sheet_id(ss, sheet_id=sheet_id, dry_run=dry_run)


def _copy_tab(src: gspread.Spreadsheet, dst: gspread.Spreadsheet, src_title: str, dst_title: str, *, dry_run: bool) -> None:
    print(f"  Copying: {src_title} -> {dst_title}")
    if dry_run:
        return

    src_ws = src.worksheet(src_title)
    copy_result = src_ws.copy_to(dst.id)
    if isinstance(copy_result, dict):
        new_sheet_id = copy_result.get("sheetId")
    else:
        new_sheet_id = copy_result

    if new_sheet_id is None:
        raise RuntimeError(f"Unexpected copy result for {src_title}: {copy_result!r}")

    # Rename using sheetId to avoid stale worksheet caches.
    _rename_sheet_id(dst, sheet_id=int(new_sheet_id), new_title=dst_title, dry_run=dry_run)


def _batch_update_tab_properties(
    ss: gspread.Spreadsheet,
    *,
    title: str,
    tab_color: dict | None,
    hidden: bool | None,
    dry_run: bool,
) -> None:
    sheet_id = _find_sheet_id_by_title(ss, title)
    if sheet_id is None:
        print(f"  ⚠️ Could not find sheetId for tab: {title}")
        return

    fields: list[str] = []
    props: dict = {"sheetId": sheet_id}

    if tab_color is not None:
        props["tabColor"] = tab_color
        fields.append("tabColor")

    if hidden is not None:
        props["hidden"] = hidden
        fields.append("hidden")

    if not fields:
        return

    body = {
        "requests": [
            {
                "updateSheetProperties": {
                    "properties": props,
                    "fields": ",".join(fields),
                }
            }
        ]
    }

    if dry_run:
        print(f"  [DRY RUN] Would update tab properties: {title} fields={fields}")
        return

    ss.batch_update(body)


def _ensure_tab_from_source(
    src_ss: gspread.Spreadsheet,
    dst_ss: gspread.Spreadsheet,
    *,
    src_title: str,
    dst_title: str,
    tab_color: dict,
    hidden: bool,
    dry_run: bool,
) -> None:
    renamed_old = _rename_if_exists(dst_ss, dst_title, f"OLD__{dst_title}", dry_run=dry_run)
    _copy_tab(src_ss, dst_ss, src_title, dst_title, dry_run=dry_run)
    if renamed_old:
        _delete_if_exists(dst_ss, renamed_old, dry_run=dry_run)

    _batch_update_tab_properties(dst_ss, title=dst_title, tab_color=tab_color, hidden=hidden, dry_run=dry_run)


def _upload_table(ws: gspread.Worksheet, table: list[list[object]]) -> None:
    # Preserve formatting: do not resize (can drop formatted rows/cols).
    # Clearing values does not clear formatting.
    ws.clear()
    values = [["" if v is None else v for v in row] for row in table]
    ws.update(values=values, range_name="A1", value_input_option="USER_ENTERED")


def _import_excel_with_formatting(
    creds: Credentials,
    xlsx_path: Path,
    target_ss: gspread.Spreadsheet,
    target_tab_title: str,
    *,
    dry_run: bool = False,
) -> None:
    """Import an Excel file into a target Google Sheet as a new tab, preserving key formatting.

    Reads Excel file with openpyxl, creates a new sheet, uploads values,
    and applies column widths from the Excel file.
    """
    if dry_run:
        print(f"  [DRY RUN] Would import {xlsx_path.name} -> {target_tab_title}")
        return

    print(f"  Importing {xlsx_path.name}...")

    # Load Excel file
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = wb.active

    # Get values
    table = _load_excel_values(xlsx_path)
    rows, cols = _table_shape(table)

    # Get column widths from Excel
    col_widths: dict[int, float] = {}
    for col_letter, dim in sheet.column_dimensions.items():
        if dim.width:
            col_idx = openpyxl.utils.column_index_from_string(col_letter) - 1  # 0-indexed
            col_widths[col_idx] = dim.width

    wb.close()

    # Rename old tab if it exists
    renamed_old = _rename_if_exists(target_ss, target_tab_title, f"OLD__{target_tab_title}", dry_run=False)

    # Create new sheet or get existing
    ws = _worksheet_by_title_ci(target_ss, target_tab_title)
    if not ws:
        ws = target_ss.add_worksheet(title=target_tab_title, rows=max(rows, 100), cols=max(cols, 50))
        print(f"    Created new sheet: {target_tab_title}")
    else:
        # Resize if needed
        if ws.row_count < rows:
            ws.add_rows(rows - ws.row_count)
        if ws.col_count < cols:
            ws.add_cols(cols - ws.col_count)

    # Upload values
    values = [["" if v is None else v for v in row] for row in table]
    ws.update(values=values, range_name="A1", value_input_option="USER_ENTERED")
    print(f"    Uploaded {rows}x{cols} values")

    # Apply column widths
    if col_widths:
        requests = []
        for col_idx, width in col_widths.items():
            if col_idx < cols:
                # Convert Excel width to pixels (approximate: Excel width * 7)
                pixel_width = int(width * 7)
                requests.append({
                    "updateDimensionProperties": {
                        "range": {
                            "sheetId": ws.id,
                            "dimension": "COLUMNS",
                            "startIndex": col_idx,
                            "endIndex": col_idx + 1,
                        },
                        "properties": {"pixelSize": pixel_width},
                        "fields": "pixelSize",
                    }
                })
        if requests:
            target_ss.batch_update({"requests": requests})
            print(f"    Applied column widths")

    # Delete the old tab if we renamed it
    if renamed_old:
        _delete_if_exists(target_ss, renamed_old, dry_run=False)

    print(f"    Done: {target_tab_title}")

    # Small delay to avoid rate limiting
    time.sleep(2)


def _apply_ha_tab_tweaks(
    target_ss: gspread.Spreadsheet,
    tab_title: str,
    *,
    tab_color: dict | None = None,
    freeze_row: int = 6,
    freeze_col: int = 2,
) -> None:
    """Apply standard tweaks after importing an Excel file:
    - Set tab color
    - Freeze panes
    - Left-align and unmerge header row (row 6)
    """
    ws = _worksheet_by_title_ci(target_ss, tab_title)
    if not ws:
        print(f"    [WARN] Tab not found for tweaks: {tab_title}")
        return

    # Set tab color
    if tab_color:
        _set_tab_properties(target_ss, ws, tab_color=tab_color, hidden=False)

    # Freeze panes
    requests = [
        {
            "updateSheetProperties": {
                "properties": {
                    "sheetId": ws.id,
                    "gridProperties": {
                        "frozenRowCount": freeze_row,
                        "frozenColumnCount": freeze_col,
                    },
                },
                "fields": "gridProperties.frozenRowCount,gridProperties.frozenColumnCount",
            }
        }
    ]
    target_ss.batch_update({"requests": requests})
    print(f"    Applied tweaks to {tab_title}: color, freeze at row {freeze_row}/col {freeze_col}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--month-folder",
        default="",
        help="Month folder under this directory (e.g., '11. Nov')",
    )
    parser.add_argument(
        "--target-sheet",
        default="",
        help="Target Google Sheet link or ID",
    )
    parser.add_argument(
        "--confirm",
        action="store_true",
        help="Actually modify the target spreadsheet (default is dry-run)",
    )
    parser.add_argument(
        "--skip-copy",
        action="store_true",
        help="Skip copying tabs from source spreadsheets; only upload values into existing HA-CF tabs in the target.",
    )
    parser.add_argument(
        "--cleanup-drafts",
        action="store_true",
        help="Delete known draft/duplicate tabs created by earlier failed runs (safe, specific patterns only).",
    )
    parser.add_argument(
        "--delete-other-tabs",
        action="store_true",
        help="Delete ALL other tabs in the target workbook besides the required Cash Flow tabs (destructive).",
    )
    parser.add_argument(
        "--skip-upload",
        action="store_true",
        help="Copy tabs and set formatting, but do not upload HA-CF values",
    )
    args = parser.parse_args()

    dry_run = not args.confirm

    state = _load_state()

    # Select month folder
    month_folder = args.month_folder.strip()
    if not month_folder:
        options = _list_month_folders(DEFAULT_WORKDIR)
        if not options:
            raise SystemExit(f"No month folders found under: {DEFAULT_WORKDIR}")

        last_month_folder = str(state.get("last_month_folder") or "").strip()
        print("\nSelect month folder:")
        for i, name in enumerate(options, start=1):
            print(f"  {i}. {name}")

        default_idx = 1
        if last_month_folder and last_month_folder in options:
            default_idx = options.index(last_month_folder) + 1

        choice = input(f"Enter number [1-{len(options)}] (default {default_idx}): ").strip()
        try:
            idx = int(choice)
        except Exception:
            idx = default_idx
        idx = max(1, min(len(options), idx))
        month_folder = options[idx - 1]

    month_dir = (DEFAULT_WORKDIR / month_folder).resolve()
    if not month_dir.exists():
        raise SystemExit(f"Month folder not found: {month_dir}")

    month_abbrev = _month_abbrev_from_folder(month_folder)
    month_tab = f"HA-CF-{month_abbrev}"

    # Target sheet id
    target_id = _extract_sheet_id(args.target_sheet)
    if not target_id:
        last_target_id = str(state.get("last_target_sheet_id") or "").strip()
        default_target = last_target_id or DEFAULT_TARGET_SHEET_ID
        entered = input(f"Paste TARGET sheet link/ID [default: {default_target}]: ").strip()
        target_id = _extract_sheet_id(entered) or default_target

    creds_path = _resolve_service_account_file()
    email = _service_account_email(creds_path)
    if email:
        print(f"Service account: {email}")
        print("Ensure the target spreadsheet is shared with this email.")

    creds = Credentials.from_service_account_file(creds_path, scopes=SCOPES)
    gc = gspread.authorize(creds)

    target_ss = _open_sheet_or_exit(
        gc,
        sheet_id=target_id,
        label="TARGET",
        service_account_email=email,
    )

    if args.cleanup_drafts:
        print("\nCleaning up draft/duplicate tabs...")
        _cleanup_draft_tabs(target_ss, dry_run=dry_run)

    # Persist last-used inputs once we know TARGET is accessible.
    state["last_target_sheet_id"] = target_id
    state["last_month_folder"] = month_folder
    _save_state(state)

    template_ss = None
    month_end_ss = None

    if not args.skip_copy:
        template_ss = _open_sheet_or_exit(
            gc,
            sheet_id=TEMPLATE_BOOK_ID,
            label="TEMPLATE (PORTFOLIO CASH FLOW / PROPERTY CODES)",
            service_account_email=email,
        )
        month_end_ss = _open_sheet_or_exit(
            gc,
            sheet_id=MONTH_END_BOOK_ID,
            label="MONTH END (Bank Rec priorities)",
            service_account_email=email,
        )

    print("\nPlanned operations:")
    bs_tab = f"HA-BS-{month_abbrev}"
    print(f"  - Month folder: {month_folder} (HA-CF month tab: {month_tab})")
    if args.skip_copy:
        print("  - Skip copying tabs from source spreadsheets")
    else:
        print("  - Copy templates:")
        print(f"      - {TEMPLATE_TABS[0]} (blue)")
        print(f"      - {TEMPLATE_TABS[1]} (black, hidden)")
        print(f"      - {TEMPLATE_TABS[2]} (black, hidden)")
        print("  - Copy month end list:")
        print(f"      - {MONTH_END_TAB} -> {MONTH_END_TARGET_TAB} (black, hidden)")
    if args.skip_upload:
        print("  - Skip importing HA-CF/HA-BS tabs")
    else:
        print("  - Import Excel exports with formatting (like File > Import):")
        print(f"      - {month_tab} (red)")
        print("      - HA-CF-3MOS (red)")
        print("      - HA-CF-YTD (red)")
        print(f"      - HA-BS-{month_abbrev} (red, if BS export found)")

    if args.delete_other_tabs:
        print("  - Delete all other tabs in target workbook (keep Cash Flow tabs only)")

    if dry_run:
        print("\n[DRY RUN] No changes will be applied. Re-run with --confirm to apply.")

    if not args.skip_copy:
        assert template_ss is not None
        assert month_end_ss is not None

        # Copy template tabs
        _ensure_tab_from_source(
            template_ss,
            target_ss,
            src_title="PORTFOLIO CASH FLOW",
            dst_title="PORTFOLIO CASH FLOW",
            tab_color=COLOR_BLUE,
            hidden=False,
            dry_run=dry_run,
        )
        _ensure_tab_from_source(
            template_ss,
            target_ss,
            src_title="PROPERTY CODES",
            dst_title="PROPERTY CODES",
            tab_color=COLOR_BLACK,
            hidden=True,
            dry_run=dry_run,
        )
        _ensure_tab_from_source(
            template_ss,
            target_ss,
            src_title="PROPERTY STATUS",
            dst_title="PROPERTY STATUS",
            tab_color=COLOR_BLACK,
            hidden=True,
            dry_run=dry_run,
        )

        # Copy month end tab
        _ensure_tab_from_source(
            month_end_ss,
            target_ss,
            src_title=MONTH_END_TAB,
            dst_title=MONTH_END_TARGET_TAB,
            tab_color=COLOR_BLACK,
            hidden=True,
            dry_run=dry_run,
        )

        # Note: HA-CF and HA-BS tabs are NOT copied from a reference workbook.
        # They are imported directly from the Excel exports with their formatting preserved.

    if args.delete_other_tabs:
        keep_titles = {
            "PORTFOLIO CASH FLOW",
            "PROPERTY CODES",
            "PROPERTY STATUS",
            MONTH_END_TARGET_TAB,
            month_tab,
            "HA-CF-3MOS",
            "HA-CF-YTD",
            bs_tab,
        }
        _delete_other_tabs(target_ss, keep_titles=keep_titles, dry_run=dry_run)

    if args.skip_upload:
        _finalize_target_workbook(target_ss, month_tab=month_tab, bs_tab=bs_tab, dry_run=dry_run)
        print("\n✅ Done (tabs copied / styled).")
        return

    exports = _detect_exports(month_dir)
    exports = _rename_exports_in_place(exports, month_abbrev=month_abbrev, dry_run=dry_run)
    bs_xlsx = exports.get("BS")
    print("\nDetected exports:")
    print(f"  NOV : {exports['NOV'].name}")
    print(f"  3MOS: {exports['3MOS'].name}")
    print(f"  YTD : {exports['YTD'].name}")
    print(f"  BS  : {bs_xlsx.name if bs_xlsx else '(not found)'}")

    if dry_run:
        print("\n[DRY RUN] Would import exports into HA-CF/HA-BS tabs (with Excel formatting):")
        for key, tab in (("NOV", month_tab), ("3MOS", "HA-CF-3MOS"), ("YTD", "HA-CF-YTD")):
            table = _load_excel_values(exports[key])
            rows, cols = _table_shape(table)
            print(f"  - {tab}: {rows}x{cols} from {exports[key].name}")
        if bs_xlsx:
            table = _load_excel_values(bs_xlsx)
            rows, cols = _table_shape(table)
            print(f"  - {bs_tab}: {rows}x{cols} from {bs_xlsx.name}")
        else:
            print(f"  - {bs_tab}: (no Balance Sheet file found)")
        return

    # Import each Excel file with formatting (mimics File > Import in Google Sheets)
    import_plan: list[tuple[Path, str]] = [
        (exports["NOV"], month_tab),
        (exports["3MOS"], "HA-CF-3MOS"),
        (exports["YTD"], "HA-CF-YTD"),
    ]

    for xlsx_path, tab in import_plan:
        print("\n" + "=" * 100)
        print(f"Importing with formatting: {xlsx_path.name} -> {tab}")
        _import_excel_with_formatting(creds, xlsx_path, target_ss, tab, dry_run=False)
        # Refresh target_ss after each import to see new tabs
        target_ss = gc.open_by_key(target_id)
        # Apply tweaks: red tab color, freeze panes
        _apply_ha_tab_tweaks(target_ss, tab, tab_color=COLOR_RED, freeze_row=6, freeze_col=2)

    # Import Balance Sheet if found
    if bs_xlsx:
        print("\n" + "=" * 100)
        print(f"Importing Balance Sheet with formatting: {bs_xlsx.name} -> {bs_tab}")
        _import_excel_with_formatting(creds, bs_xlsx, target_ss, bs_tab, dry_run=False)
        target_ss = gc.open_by_key(target_id)
        _apply_ha_tab_tweaks(target_ss, bs_tab, tab_color=COLOR_RED, freeze_row=6, freeze_col=2)
    else:
        print(f"\n[WARN] Balance Sheet not found in {month_dir}. Skipping HA-BS-{month_abbrev} import.")

    # Refresh again before finalization
    target_ss = gc.open_by_key(target_id)
    _finalize_target_workbook(target_ss, month_tab=month_tab, bs_tab=bs_tab, dry_run=dry_run)

    print("\n✅ Monthly workbook prepared.")


if __name__ == "__main__":
    main()
