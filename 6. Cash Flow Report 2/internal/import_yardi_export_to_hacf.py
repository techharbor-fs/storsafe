"""Import a Yardi Excel export into a HA-CF tab.

This is the beginning of a repeatable pipeline:
- Reads an .xlsx export
- Detects whether it contains HA-CF-style account codes (####-####)
- Reports its inferred layout

Write support is optional and off by default.

Usage:
  python import_yardi_export_to_hacf.py --xlsx "path\\to\\export.xlsx" --tab "HA-CF-OCT" --dry-run
  python import_yardi_export_to_hacf.py --xlsx "path\\to\\export.xlsx" --tab "HA-CF-OCT" --write

Credentials are required only when using --write.

"""

from __future__ import annotations

import argparse
import json
import os
import tempfile
from pathlib import Path

try:
    from dotenv import load_dotenv

    load_dotenv()
except Exception:
    pass

import openpyxl

from hacf_utils import ACCOUNT_CODE_PATTERN, infer_layout, required_codes


DEFAULT_SHEET_ID = "1iZNLklMpAPeVo57nJVFfGBqUmQ3PD_bow4IlYOvtQj0"
SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]


def _extract_sheet_id(value: str) -> str:
    value = (value or "").strip()
    if not value:
        return value
    marker = "/spreadsheets/d/"
    if marker in value:
        after = value.split(marker, 1)[1]
        return after.split("/", 1)[0]
    return value


def _resolve_service_account_file() -> str:
    service_account_json = os.environ.get("SERVICE_ACCOUNT_JSON")
    if service_account_json:
        temp_json_path = Path(tempfile.gettempdir()) / "service_account.json"
        temp_json_path.write_text(json.dumps(json.loads(service_account_json)), encoding="utf-8")
        return str(temp_json_path)

    env_path = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS") or os.environ.get("SERVICE_ACCOUNT_FILE")
    if not env_path:
        raise RuntimeError(
            "Missing Google service account credentials. Set SERVICE_ACCOUNT_JSON or "
            "GOOGLE_APPLICATION_CREDENTIALS/SERVICE_ACCOUNT_FILE."
        )
    return env_path


def load_sheet_values(xlsx_path: Path, sheet_name: str | None) -> list[list[object]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=False, read_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
        values: list[list[object]] = []
        for row in ws.iter_rows(values_only=True):
            values.append(list(row))
        return values
    finally:
        wb.close()


def any_account_codes(table: list[list[object]]) -> bool:
    for row in table:
        for cell in row:
            if isinstance(cell, str) and ACCOUNT_CODE_PATTERN.match(cell.strip()):
                return True
    return False


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True, help="Path to the Yardi Excel export")
    parser.add_argument("--sheet", default=None, help="Optional worksheet name inside the workbook")
    parser.add_argument("--tab", required=True, help="Target Google Sheet tab name (e.g., HA-CF-OCT)")
    parser.add_argument(
        "--sheet-id",
        default=os.environ.get("CASHFLOW_SHEET_ID") or DEFAULT_SHEET_ID,
        help="Google Sheet ID or full link (used for --write later)",
    )
    parser.add_argument("--dry-run", action="store_true", help="Print findings only")
    parser.add_argument("--write", action="store_true", help="(Future) upload data to the tab")

    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        raise SystemExit(f"XLSX not found: {xlsx_path}")

    table = load_sheet_values(xlsx_path, args.sheet)
    layout = infer_layout(table)

    print("=" * 100)
    print("EXPORT INSPECTION")
    print("=" * 100)
    print("file:", xlsx_path)
    print("rows:", len(table))
    print("cols (max row width):", max((len(r) for r in table), default=0))
    print("layout:")
    print(f"  account_code_col: {layout.account_code_col}")
    print(f"  description_col:  {layout.description_col}")
    print(f"  property_code_row:{layout.property_code_row}")
    print(f"  period_row:       {layout.period_row}")
    print(f"  cfads_row:        {layout.cfads_row}")

    has_codes = any_account_codes(table)
    print("contains ####-#### account codes:", has_codes)

    found_required = []
    for code in required_codes():
        found = False
        for row in table:
            for cell in row:
                if isinstance(cell, str) and cell.strip() == code:
                    found = True
                    break
            if found:
                break
        found_required.append((code, found))

    print("required HA-CF codes present in export:")
    for code, ok in found_required:
        print(f"  {code}: {'YES' if ok else 'NO'}")

    if args.write:
        sheet_id = _extract_sheet_id(args.sheet_id)
        if not sheet_id:
            raise SystemExit("Missing --sheet-id (or set CASHFLOW_SHEET_ID)")
        # We intentionally fail loudly for now; write support should only be enabled after
        # we confirm the export format and mapping rules.
        raise SystemExit(
            "--write is not implemented yet. Next step is to confirm the exact Yardi export format "
            "that includes account codes, then we will map it into the HA-CF tab structure."
        )

    if args.dry_run or not args.write:
        print("\nDone (dry run).")


if __name__ == "__main__":
    main()
