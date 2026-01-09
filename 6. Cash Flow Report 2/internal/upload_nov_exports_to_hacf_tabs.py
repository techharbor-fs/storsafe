"""Upload November Yardi exports into a Google Spreadsheet.

Goal (per workflow):
1) User provides a Google Sheet link/ID (prompted); for unattended testing, pass --sheet-id.
2) Spreadsheet must be shared with the service account email.
3) Upload the exports into HA-CF tabs verbatim (all cells/ranges as-is).

Note:
- These Yardi Property Comparison exports do not include HA account codes.
- The cashflow generator is responsible for locating the correct rows by labels.

Usage:
  python upload_nov_exports_to_hacf_tabs.py --sheet-id <link_or_id>

"""

from __future__ import annotations

import argparse
import json
import os
import tempfile
from pathlib import Path

try:
    from dotenv import load_dotenv

    load_dotenv()
except Exception:
    pass

import openpyxl
import gspread
from google.oauth2.service_account import Credentials
import re


NOV_SHEET_DEFAULT = "160onO2dxp7fewgcibKoQ9biFH0qW69wg215J9phMVSs"
SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]


_ACCT_CODE_RE = re.compile(r"^\d{4}-\d{4}$")


def _extract_sheet_id(value: str) -> str:
    value = (value or "").strip()
    if not value:
        return value
    marker = "/spreadsheets/d/"
    if marker in value:
        after = value.split(marker, 1)[1]
        return after.split("/", 1)[0]
    return value


def _resolve_service_account_file() -> str:
    service_account_json = os.environ.get("SERVICE_ACCOUNT_JSON")
    if service_account_json:
        temp_json_path = Path(tempfile.gettempdir()) / "service_account.json"
        temp_json_path.write_text(json.dumps(json.loads(service_account_json)), encoding="utf-8")
        return str(temp_json_path)

    env_path = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS") or os.environ.get("SERVICE_ACCOUNT_FILE")
    if not env_path:
        raise RuntimeError(
            "Missing Google service account credentials. Set SERVICE_ACCOUNT_JSON or "
            "GOOGLE_APPLICATION_CREDENTIALS/SERVICE_ACCOUNT_FILE."
        )
    return env_path


def _service_account_email(creds_path: str) -> str | None:
    try:
        data = json.loads(Path(creds_path).read_text(encoding="utf-8"))
    except Exception:
        return None
    email = data.get("client_email")
    return str(email).strip() if email else None


def _load_excel_values(xlsx_path: Path, sheet_name: str | None = None) -> list[list[object]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _table_shape(table: list[list[object]]) -> tuple[int, int]:
    rows = len(table)
    cols = max((len(r) for r in table), default=0)
    return rows, cols


def _count_account_codes_in_first_col(table: list[list[object]]) -> int:
    count = 0
    for row in table:
        if not row:
            continue
        cell = row[0]
        if cell is None:
            continue
        text = str(cell).strip()
        if _ACCT_CODE_RE.match(text):
            count += 1
    return count


def _auto_detect_exports(xlsx_dir: Path) -> dict[str, Path]:
    """Return export paths for NOV/3MOS/YTD.

    Preference order:
    1) Use the previously-renamed, explicit filenames if present.
    2) Otherwise, auto-detect from matching Property Comparison exports in the folder.

    Auto-detect heuristic:
    - Load each candidate once to get row-count.
    - YTD tends to be the largest; NOV the smallest; 3MOS the middle.
    """

    explicit = {
        "NOV": xlsx_dir / "yardi_property_comparison_allharda_accrual_nov_2025.xlsx",
        "3MOS": xlsx_dir / "yardi_property_comparison_allharda_accrual_sep_2025_nov_2025_3mos.xlsx",
        "YTD": xlsx_dir / "yardi_property_comparison_allharda_accrual_dec_2024_nov_2025_ytd.xlsx",
    }
    if all(p.exists() for p in explicit.values()):
        return explicit

    candidates = sorted(
        [p for p in xlsx_dir.glob("Property_Comparison_allharda_Accrual*.xlsx") if p.is_file()],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if len(candidates) < 3:
        raise SystemExit(
            "Could not auto-detect exports. Expected at least 3 matching files in: "
            f"{xlsx_dir} (found {len(candidates)})."
        )

    # Only consider the three most recent matches; that's the normal workflow.
    candidates = candidates[:3]
    tables: list[tuple[Path, list[list[object]]]] = [(p, _load_excel_values(p)) for p in candidates]
    tables_sorted = sorted(tables, key=lambda item: _table_shape(item[1])[0])

    nov_path = tables_sorted[0][0]
    mos3_path = tables_sorted[1][0]
    ytd_path = tables_sorted[2][0]
    return {"NOV": nov_path, "3MOS": mos3_path, "YTD": ytd_path}


def _ensure_worksheet(ss: gspread.Spreadsheet, title: str, rows: int, cols: int) -> gspread.Worksheet:
    try:
        ws = ss.worksheet(title)
        # ensure big enough
        if ws.row_count < rows or ws.col_count < cols:
            ws.resize(rows=max(rows, ws.row_count), cols=max(cols, ws.col_count))
        return ws
    except Exception:
        return ss.add_worksheet(title=title, rows=rows, cols=cols)


def _upload_table(ws: gspread.Worksheet, table: list[list[object]]) -> None:
    # Convert None to "" for gspread
    values = [["" if v is None else v for v in row] for row in table]
    rows = max(1, len(values))
    cols = max(1, max((len(r) for r in values), default=1))
    ws.resize(rows=rows, cols=cols)
    ws.update(values=values, range_name="A1", value_input_option="USER_ENTERED")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--sheet-id", default="", help="Google Sheet link or ID")
    parser.add_argument("--dry-run", action="store_true", help="Do not write to Google Sheets")
    args = parser.parse_args()

    sheet_id = _extract_sheet_id(args.sheet_id)
    if not sheet_id:
        entered = input(f"Paste Google Sheet link/ID [default: {NOV_SHEET_DEFAULT}]: ").strip()
        sheet_id = _extract_sheet_id(entered) or NOV_SHEET_DEFAULT

    creds_path = _resolve_service_account_file()
    email = _service_account_email(creds_path)
    if email:
        print(f"Service account: {email}")
        print("Ensure the spreadsheet is shared with this email.")

    xlsx_dir = Path(r"C:\Users\jayry\python projects\automation files\03. storsafe\6. Cash Flow Report 2\11. Nov")
    exports = _auto_detect_exports(xlsx_dir)
    print("\nDetected exports:")
    for key in ("NOV", "3MOS", "YTD"):
        path = exports[key]
        print(f"  {key}: {path.name}")

    if args.dry_run:
        print("DRY RUN: would upload to sheet", sheet_id)
        for key, path in exports.items():
            table = _load_excel_values(path)
            rows, cols = _table_shape(table)
            acct_codes = _count_account_codes_in_first_col(table)
            print(f"{key}: export rows={rows} cols~={cols} acct_codes_in_colA={acct_codes}")
        return

    creds = Credentials.from_service_account_file(creds_path, scopes=SCOPES)
    gc = gspread.authorize(creds)
    ss = gc.open_by_key(sheet_id)

    # Upload verbatim Excel grids into the HA-CF tabs
    plan = [
        (exports["NOV"], "HA-CF-NOV"),
        (exports["3MOS"], "HA-CF-3MOS"),
        (exports["YTD"], "HA-CF-YTD"),
    ]

    for xlsx_path, hacf_tab in plan:
        print("\n" + "=" * 100)
        print(f"Uploading: {xlsx_path.name}")
        raw_table = _load_excel_values(xlsx_path)
        rows, cols = _table_shape(raw_table)
        acct_codes = _count_account_codes_in_first_col(raw_table)
        print(f"HA-CF (verbatim) -> {hacf_tab} ({rows}x{cols}) acct_codes_in_colA={acct_codes}")
        hacf_ws = _ensure_worksheet(
            ss,
            hacf_tab,
            rows=max(200, len(raw_table) + 10),
            cols=max(60, max((len(r) for r in raw_table), default=0) + 5),
        )
        _upload_table(hacf_ws, raw_table)

    print("\n✅ Upload complete.")


if __name__ == "__main__":
    main()
