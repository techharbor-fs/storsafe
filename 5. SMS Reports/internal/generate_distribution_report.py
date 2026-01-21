"""
Generate SMS Distribution Recommendation Workbook

Creates a distribution recommendation workbook for a single property,
with formulas linking to the financial report's Cash Flow sheet.

This is the core helper used by generate_monthly_distribution_reports.py
"""

import calendar
import re
from datetime import date
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# Styling constants
FONT_NAME = "Aptos Narrow"
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
CURRENCY_FORMAT = '_($* #,##0.00_)'
CURRENCY_FORMAT_MANUAL = '"$"#,##0.00'


def find_label_row(ws, label: str) -> tuple[int, int] | None:
    """
    Find a row containing the given label text.
    Dynamically scans the entire used range of the sheet.
    Returns (row, col) or None if not found.
    """
    max_row = ws.max_row or 200
    max_col = ws.max_column or 15
    label_lower = label.lower().strip()
    
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            val = ws.cell(row, col).value
            if val and isinstance(val, str) and label_lower in val.lower():
                return (row, col)
    return None


def find_total_cash_ending_balance(ws) -> tuple[float, int, int] | None:
    """
    Find the 'Total Cash' row in the 'Period to Date' section and return the Ending Balance value.
    Dynamically scans the entire sheet - no hardcoded row/column positions.
    
    Strategy: Find a row that has BOTH "Period to Date" AND "Ending Balance" labels,
    then find "Total Cash" below it and get the value in the Ending Balance column.
    
    Returns (value, row, col) or None.
    """
    max_row = ws.max_row or 200
    max_col = ws.max_column or 15
    
    # Find a row that contains BOTH "Period to Date" AND "Ending Balance"
    # This distinguishes the cash balance section from other "Period to Date" headers
    ptd_row = None
    ending_col = None
    
    for row in range(1, max_row + 1):
        row_has_ptd = False
        row_ending_col = None
        
        for col in range(1, max_col + 1):
            val = ws.cell(row, col).value
            if val and isinstance(val, str):
                val_lower = val.lower()
                if "period to date" in val_lower:
                    row_has_ptd = True
                if "ending balance" in val_lower:
                    row_ending_col = col
        
        # If this row has both, we found the cash balance section
        if row_has_ptd and row_ending_col:
            ptd_row = row
            ending_col = row_ending_col
            break
    
    if not ptd_row or not ending_col:
        return None
    
    # Find "Total Cash" row below Period to Date
    for row in range(ptd_row + 1, min(ptd_row + 20, max_row + 1)):
        for col in range(1, max_col + 1):
            val = ws.cell(row, col).value
            if val and isinstance(val, str) and "total cash" in val.lower():
                # Get the value in the Ending Balance column
                ending_val = ws.cell(row, ending_col).value
                if ending_val is not None:
                    try:
                        return (float(ending_val), row, ending_col)
                    except (ValueError, TypeError):
                        pass
    
    return None


def find_net_income_value(ws) -> tuple[float, int, int] | None:
    """
    Find NET INCOME row and return the monthly value (first numeric value after label).
    Dynamically searches the row for the first numeric value.
    Returns (value, row, col) or None.
    """
    result = find_label_row(ws, "NET INCOME")
    if not result:
        return None
    
    row, label_col = result
    max_col = ws.max_column or 15
    
    # Search for the first numeric value in the row after the label
    for col in range(label_col, max_col + 1):
        val = ws.cell(row, col).value
        if val is not None and not isinstance(val, str):
            try:
                return (float(val), row, col)
            except (ValueError, TypeError):
                pass
    return None


def find_note_principal_value(ws) -> tuple[float, int, int] | None:
    """
    Find Note 1 Principal row and return the monthly value (first numeric value after label).
    Dynamically searches the row for the first numeric value.
    Returns (value, row, col) or None.
    """
    result = find_label_row(ws, "Note 1 Principal")
    if not result:
        return None
    
    row, label_col = result
    max_col = ws.max_column or 15
    
    # Search for the first numeric value in the row after the label
    for col in range(label_col, max_col + 1):
        val = ws.cell(row, col).value
        if val is not None and not isinstance(val, str):
            try:
                return (float(val), row, col)
            except (ValueError, TypeError):
                pass
    return None


def find_crown_castle_value(ws_gl) -> float | None:
    """
    For NFSS: Find Crown Castle DEBIT payments in General Ledger sheet.
    Dynamically scans the entire sheet for "crown castle" text.
    Only counts rows where there is a debit entry (not credit/expense entries).
    
    Pattern in GL: ..., description, ref, check#, debit, credit, balance, ...
    A debit entry has: debit > 0 AND credit == 0
    A credit entry has: debit == 0 AND credit > 0
    
    Returns the total of all Crown Castle debit payments, or None if not found.
    """
    max_row = ws_gl.max_row or 500
    max_col = ws_gl.max_column or 15
    
    total = 0.0
    found_any = False
    
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            val = ws_gl.cell(row, col).value
            if val and isinstance(val, str) and "crown castle" in val.lower():
                # Find pairs of numeric values (debit, credit) after the description
                # The pattern is: description, ref, check#, DEBIT, CREDIT, balance
                numeric_values = []
                for c in range(col + 1, min(col + 6, max_col + 1)):
                    cell_val = ws_gl.cell(row, c).value
                    try:
                        num_val = float(cell_val) if cell_val is not None else None
                        # Skip strings (reference numbers stored as strings like '3116878')
                        if isinstance(ws_gl.cell(row, c).value, str):
                            continue
                        if num_val is not None:
                            numeric_values.append((c, num_val))
                    except (ValueError, TypeError):
                        pass
                
                # Look for debit/credit pair: debit > 0 and next value == 0
                for i, (c, num_val) in enumerate(numeric_values):
                    if 0 < num_val < 100000:  # Reasonable debit amount
                        # Check if next numeric value is 0 (credit column)
                        if i + 1 < len(numeric_values) and numeric_values[i + 1][1] == 0:
                            total += num_val
                            found_any = True
                            break
                break  # Found crown castle in this row, move to next row
    
    return total if found_any else None


def generate_distribution_report(
    financial_path: Path,
    property_name: str,
    report_month: str,  # YYYY-MM format
    output_path: Path,
    operating_hold: float = -10000.0,
    b15_value: Optional[float] = None,
    current_balance_date: Optional[str] = None,  # YYYY-MM-DD format
) -> None:
    """
    Generate a distribution recommendation workbook for a single property.
    
    Args:
        financial_path: Path to the property's financial report workbook
        property_name: Display name (e.g., "Cary", "NFSS")
        report_month: Reporting month in YYYY-MM format
        output_path: Where to save the generated workbook
        operating_hold: Value for A4 (default -10000)
        b15_value: Optional cash balance for B15
        current_balance_date: Override for the "Current Balance @" date
    """
    # Parse report month
    year, month = map(int, report_month.split("-"))
    last_day = calendar.monthrange(year, month)[1]
    
    # Default balance date to last day of report month
    if current_balance_date:
        bal_year, bal_month, bal_day = map(int, current_balance_date.split("-"))
    else:
        bal_year, bal_month, bal_day = year, month, last_day
    
    # Check if this is NFSS
    is_nfss = "nfss" in property_name.lower() or "northfield" in property_name.lower()
    
    # Load the financial workbook
    fin_wb = openpyxl.load_workbook(financial_path, data_only=True)
    
    if "Cash Flow" not in fin_wb.sheetnames:
        raise ValueError(f"No 'Cash Flow' sheet found in {financial_path}")
    
    cf_ws = fin_wb["Cash Flow"]
    
    # Extract values from Cash Flow sheet
    ending_balance_result = find_total_cash_ending_balance(cf_ws)
    if not ending_balance_result:
        raise ValueError(f"Could not find Total Cash Ending Balance in {financial_path}")
    book_balance, _, _ = ending_balance_result
    
    net_income_result = find_net_income_value(cf_ws)
    if not net_income_result:
        raise ValueError(f"Could not find NET INCOME in {financial_path}")
    net_income, _, _ = net_income_result
    
    note_principal_result = find_note_principal_value(cf_ws)
    if not note_principal_result:
        # Note principal might not exist for all properties
        note_principal = 0.0
    else:
        note_principal, _, _ = note_principal_result
    
    # For NFSS, also get Crown Castle payment
    crown_castle_value = None
    if is_nfss and "General Ledger" in fin_wb.sheetnames:
        crown_castle_value = find_crown_castle_value(fin_wb["General Ledger"])
    
    fin_wb.close()
    
    # Create the distribution workbook
    out_wb = openpyxl.Workbook()
    ws = out_wb.active
    ws.title = "Sheet1"
    
    # Set column widths (B, C, D, E are fixed; A will auto-fit later)
    ws.column_dimensions["B"].width = 34.14
    ws.column_dimensions["C"].width = 13.0
    ws.column_dimensions["D"].width = 9.14
    ws.column_dimensions["E"].width = 9.29
    
    # Set row heights
    ws.row_dimensions[6].height = 6.75
    ws.row_dimensions[9].height = 7.5
    
    # Helper to set cell with style
    def set_cell(row, col, value, bold=False, currency=False, manual_currency=False, yellow=False):
        cell = ws.cell(row, col)
        cell.value = value
        cell.font = Font(name=FONT_NAME, bold=bold)
        if currency:
            cell.number_format = CURRENCY_FORMAT
        elif manual_currency:
            cell.number_format = CURRENCY_FORMAT_MANUAL
        if yellow:
            cell.fill = YELLOW_FILL
    
    # Row 1: Property name
    set_cell(1, 1, property_name, bold=True)
    
    # Row 2: "Book Balance"
    set_cell(2, 1, "Book Balance")
    
    # Row 3: Book balance value and "as of" date
    set_cell(3, 1, book_balance, currency=True)
    set_cell(3, 2, f'="as of "&TEXT(DATE({year},{month},{last_day}),"mm.dd.yyyy")')
    
    # Row 4: Operating hold
    set_cell(4, 1, operating_hold, currency=True)
    set_cell(4, 2, "Operating hold")
    
    # Row 5: Sum formula and label
    set_cell(5, 1, "=SUM(A3:A4)")
    ws.cell(5, 1).number_format = CURRENCY_FORMAT
    set_cell(5, 2, "Cash available for distribution")
    
    # Row 6: blank
    
    # Row 7: Net income calculation
    # Note principal is typically negative, we need to make it positive for subtraction display
    note_principal_abs = abs(note_principal)
    set_cell(7, 1, f"={net_income}-{note_principal_abs}")
    ws.cell(7, 1).number_format = CURRENCY_FORMAT
    set_cell(7, 2, f'=TEXT(DATE({year},{month},1),"mmmm")&" Net Income after Note 1 Principal"')
    
    if is_nfss and crown_castle_value:
        # NFSS special layout with Crown Castle row
        # Row 8: Crown Castle payment
        set_cell(8, 1, crown_castle_value, currency=True)
        set_cell(8, 2, "Crown Castle payment")
        
        # Row 9: Distribution recommendation (yellow highlight)
        set_cell(9, 1, "=IF(A8>0,MAX(0,MIN(A7,A5))+A8,MAX(0,MIN(A7,A5)))", yellow=True)
        ws.cell(9, 1).number_format = CURRENCY_FORMAT
        set_cell(9, 2, "Total Net Income recommended for distribution", yellow=True)
    else:
        # Standard layout
        # Row 8: Distribution recommendation (yellow highlight)
        set_cell(8, 1, "=MAX(0,MIN(A7,A5))", yellow=True)
        ws.cell(8, 1).number_format = CURRENCY_FORMAT
        set_cell(8, 2, "Total Net Income recommended for distribution", yellow=True)
    
    # Row 15: Current balance label and value
    set_cell(15, 1, f'="Current Balance @ "&TEXT(DATE({bal_year},{bal_month},{bal_day}),"mm.dd.yyyy")')
    if b15_value is not None:
        set_cell(15, 2, b15_value, manual_currency=True)
    
    # Row 16: "includes" label
    set_cell(16, 1, f'="includes "&TEXT(DATE({year},{month},1),"mmm")&" income"')
    
    # Auto-fit column A width based on content
    # This mimics double-clicking the column border in Excel
    max_length = 0
    for row in range(1, 20):
        cell = ws.cell(row, 1)
        if cell.value:
            # For formulas, estimate the displayed length
            cell_value = str(cell.value)
            if cell_value.startswith("="):
                # Estimate rendered length for common formulas
                if "Current Balance" in cell_value:
                    estimated_length = 30  # "Current Balance @ 12.31.2025"
                elif "includes" in cell_value:
                    estimated_length = 20  # "includes Dec income"
                elif "SUM" in cell_value or "MAX" in cell_value:
                    estimated_length = 15  # Currency value
                else:
                    estimated_length = 25
            else:
                estimated_length = len(cell_value)
            
            # Account for bold font (slightly wider)
            if cell.font and cell.font.bold:
                estimated_length *= 1.1
            
            max_length = max(max_length, estimated_length)
    
    # Add padding and set width (Excel width units ≈ characters * 1.2)
    ws.column_dimensions["A"].width = max_length * 1.2 + 2
    
    # Save the workbook
    output_path.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(output_path)
    out_wb.close()


# CLI interface for standalone usage
if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Generate SMS distribution recommendation workbook")
    parser.add_argument("--financial", required=True, type=Path, help="Path to financial workbook")
    parser.add_argument("--property-name", required=True, help="Property display name")
    parser.add_argument("--report-month", required=True, help="Report month (YYYY-MM)")
    parser.add_argument("--output", required=True, type=Path, help="Output path")
    parser.add_argument("--operating-hold", type=float, default=-10000.0)
    parser.add_argument("--b15-value", type=float, default=None)
    parser.add_argument("--current-balance-date", default=None)
    
    args = parser.parse_args()
    
    generate_distribution_report(
        financial_path=args.financial,
        property_name=args.property_name,
        report_month=args.report_month,
        output_path=args.output,
        operating_hold=args.operating_hold,
        b15_value=args.b15_value,
        current_balance_date=args.current_balance_date,
    )
    print(f"Generated: {args.output}")
