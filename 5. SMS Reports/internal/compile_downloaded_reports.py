"""
Compile Downloaded SMS Reports into Standard Format

When reports are downloaded from the accounting system, they come as separate files:
  - Cash_Flow_sms[property]_Accrual.xlsx
  - 12_Month_Statement_sms[property]_Accrual.xlsx
  - GeneralLedger_sms[property]_Accrual.xlsx

This script compiles them into the standard Financial Report format:
  - SMS-[Property] MM.DD.YY Financial Report_Final.xlsx
  - Sheets: General Ledger, Cash Flow, 12 Month Statement

Usage:
    py -3 compile_downloaded_reports.py --reports-folder ".Reports/12. Dec" --report-date "12.31.25"
"""

import argparse
import openpyxl
from pathlib import Path
from copy import copy
import re


# Property mapping: code in filename -> display name
PROPERTY_MAP = {
    "smscary": "Cary",
    "smsaltoo": "Altoona",
    "smscrys": "Crystal Lake",
    "smsnfss": "NFSS",
}

# Expected file prefixes and their target sheet names
# Also includes validation: expected text that should appear in cell A1 or nearby
SHEET_CONFIG = [
    {
        "prefix": "GeneralLedger",
        "sheet_name": "General Ledger",
        "validate_text": ["General Ledger", "GL"],  # Text to look for in first rows
    },
    {
        "prefix": "Cash_Flow",
        "sheet_name": "Cash Flow",
        "validate_text": ["Cash Flow", "Statement of Cash"],
    },
    {
        "prefix": "12_Month_Statement",
        "sheet_name": "12 Month Statement",
        "validate_text": ["Income Statement", "12 Month", "Revenue"],
    },
]


def copy_sheet_content(source_ws, target_ws):
    """Copy all content, formatting, and dimensions from source to target worksheet."""
    
    # Copy column dimensions
    for col_letter, col_dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = col_dim.width
        target_ws.column_dimensions[col_letter].hidden = col_dim.hidden
    
    # Copy row dimensions
    for row_num, row_dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row_num].height = row_dim.height
        target_ws.row_dimensions[row_num].hidden = row_dim.hidden
    
    # Copy cell values and formatting
    for row in source_ws.iter_rows():
        for cell in row:
            target_cell = target_ws.cell(row=cell.row, column=cell.column)
            target_cell.value = cell.value
            
            if cell.has_style:
                target_cell.font = copy(cell.font)
                target_cell.border = copy(cell.border)
                target_cell.fill = copy(cell.fill)
                target_cell.number_format = cell.number_format
                target_cell.protection = copy(cell.protection)
                target_cell.alignment = copy(cell.alignment)
    
    # Copy merged cells
    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))
    
    # Copy print settings
    target_ws.print_options = source_ws.print_options
    target_ws.page_margins = source_ws.page_margins
    target_ws.page_setup = source_ws.page_setup


def validate_report_content(file_path: Path, expected_texts: list) -> tuple[bool, str]:
    """
    Validate that a downloaded file contains expected content.
    Checks first 10 rows for any of the expected text patterns.
    Returns (is_valid, message).
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # Collect text from first 10 rows, columns A-E
        found_text = []
        for row in range(1, 11):
            for col in range(1, 6):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str):
                    found_text.append(cell_value)
        
        wb.close()
        
        # Check if any expected text is found
        all_text = " ".join(found_text).upper()
        for expected in expected_texts:
            if expected.upper() in all_text:
                return True, f"Found '{expected}'"
        
        # Show what was found for debugging
        preview = found_text[:5] if found_text else ["(empty)"]
        return False, f"Expected one of {expected_texts}, found: {preview}"
        
    except Exception as e:
        return False, f"Error reading file: {e}"


def scan_downloaded_files(folder: Path) -> dict:
    """
    Scan folder for downloaded report files.
    Returns dict with structure:
    {
        property_code: {
            "property_name": str,
            "files": {prefix: Path or None},
            "complete": bool,
            "missing": [list of missing prefixes]
        }
    }
    """
    results = {}
    
    for property_code, property_name in PROPERTY_MAP.items():
        files = {}
        missing = []
        
        for config in SHEET_CONFIG:
            prefix = config["prefix"]
            file_path = folder / f"{prefix}_{property_code}_Accrual.xlsx"
            
            if file_path.exists():
                files[prefix] = file_path
            else:
                files[prefix] = None
                missing.append(prefix)
        
        # Only include properties that have at least one file
        if any(f is not None for f in files.values()):
            results[property_code] = {
                "property_name": property_name,
                "files": files,
                "complete": len(missing) == 0,
                "missing": missing
            }
    
    return results


def compile_property_report(folder: Path, property_code: str, property_name: str,
                           files: dict, report_date: str) -> tuple[Path, list]:
    """
    Compile the 3 downloaded files into one Financial Report workbook.
    Returns (output_path, list_of_source_files_to_delete).
    """
    print(f"\nCompiling {property_name}...")
    
    output_wb = openpyxl.Workbook()
    default_sheet = output_wb.active
    source_files_to_delete = []
    
    for config in SHEET_CONFIG:
        prefix = config["prefix"]
        target_sheet_name = config["sheet_name"]
        validate_texts = config["validate_text"]
        source_file = files.get(prefix)
        
        if not source_file or not source_file.exists():
            raise ValueError(f"Missing required file: {prefix}_{property_code}_Accrual.xlsx")
        
        # Validate content before adding
        is_valid, msg = validate_report_content(source_file, validate_texts)
        if not is_valid:
            raise ValueError(f"{source_file.name} validation failed: {msg}")
        
        print(f"  + {target_sheet_name} from {source_file.name} (validated: {msg})")
        
        # Load and copy
        source_wb = openpyxl.load_workbook(source_file)
        source_ws = source_wb.active
        target_ws = output_wb.create_sheet(title=target_sheet_name)
        copy_sheet_content(source_ws, target_ws)
        source_wb.close()
        
        source_files_to_delete.append(source_file)
    
    # Remove default empty sheet
    if default_sheet.title == "Sheet":
        output_wb.remove(default_sheet)
    
    # Save output
    output_filename = f"SMS-{property_name} {report_date} Financial Report_Final.xlsx"
    output_path = folder / output_filename
    output_wb.save(output_path)
    output_wb.close()
    
    # Verify output has correct sheets
    verify_wb = openpyxl.load_workbook(output_path)
    expected_sheets = [c["sheet_name"] for c in SHEET_CONFIG]
    if verify_wb.sheetnames != expected_sheets:
        verify_wb.close()
        raise ValueError(f"Output verification failed. Expected sheets {expected_sheets}, got {verify_wb.sheetnames}")
    verify_wb.close()
    
    print(f"  Saved: {output_filename}")
    
    return output_path, source_files_to_delete


def main():
    parser = argparse.ArgumentParser(
        description="Compile downloaded SMS reports into standard Financial Report format"
    )
    parser.add_argument(
        "--reports-folder",
        required=True,
        help="Path to the month folder containing downloaded files"
    )
    parser.add_argument(
        "--report-date",
        required=True,
        help="Date string for output filename (e.g., '12.31.25')"
    )
    
    args = parser.parse_args()
    
    # Resolve folder path
    folder = Path(args.reports_folder)
    if not folder.is_absolute():
        folder = Path(__file__).parent / folder
    
    if not folder.exists():
        print(f"ERROR: Folder not found: {folder}")
        return 1
    
    print("=" * 60)
    print("SMS Reports Compilation")
    print("=" * 60)
    print(f"Folder: {folder}")
    print(f"Report date: {args.report_date}")
    
    # Scan for downloaded files
    scan_results = scan_downloaded_files(folder)
    
    if not scan_results:
        print("\nNo downloaded report files found.")
        print("Expected files like: Cash_Flow_smscary_Accrual.xlsx")
        return 1
    
    # Report findings
    complete_properties = {k: v for k, v in scan_results.items() if v["complete"]}
    incomplete_properties = {k: v for k, v in scan_results.items() if not v["complete"]}
    
    if incomplete_properties:
        print("\n" + "!" * 60)
        print("WARNING: Incomplete file sets found (will NOT be compiled):")
        print("!" * 60)
        for code, info in incomplete_properties.items():
            print(f"  {info['property_name']}: Missing {info['missing']}")
        print()
    
    if not complete_properties:
        print("\nNo complete file sets to compile.")
        return 1
    
    print(f"\nReady to compile: {', '.join(v['property_name'] for v in complete_properties.values())}")
    
    # Compile each complete property
    compiled_files = []
    all_files_to_delete = []
    
    for property_code, info in complete_properties.items():
        try:
            output_path, files_to_delete = compile_property_report(
                folder,
                property_code,
                info["property_name"],
                info["files"],
                args.report_date
            )
            compiled_files.append(output_path)
            all_files_to_delete.extend(files_to_delete)
        except Exception as e:
            print(f"  ERROR compiling {info['property_name']}: {e}")
    
    # Delete source files only after all compilations succeed
    if compiled_files:
        print("\nCleaning up source files...")
        for source_file in all_files_to_delete:
            if source_file.exists():
                source_file.unlink()
                print(f"  Deleted: {source_file.name}")
    
    # Summary
    print("\n" + "=" * 60)
    print(f"Compilation complete. {len(compiled_files)} files created:")
    for f in compiled_files:
        print(f"  - {f.name}")
    
    if incomplete_properties:
        print(f"\n{len(incomplete_properties)} property(ies) skipped due to missing files.")
    
    print("=" * 60)
    
    return 0


if __name__ == "__main__":
    exit(main())
