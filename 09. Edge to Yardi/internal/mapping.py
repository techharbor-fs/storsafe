"""Mapping utilities for Edge → Yardi.

This phase intentionally keeps mapping sources simple and auditable.

Rules:
- Property mapping is required: property_name -> yardi_entity.
- The PDF "Code" column is *usually* the Yardi account code, but some rows have
    blank codes. Those blank-code rows are reported as "unmapped accounts".
- An optional account mapping table can be used for overrides:
        edge_code -> yardi_account
    If no override exists, we default to using the code as-is.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import cast

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class MappingResult:
    mapped_rows: list[dict]
    unmapped_properties: list[str]
    unmapped_accounts: list[dict]


def _read_mapping_csv(path: Path, *, key_field: str, value_field: str) -> dict[str, str]:
    if not path.exists():
        return {}

    out: dict[str, str] = {}
    with path.open("r", newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for raw in reader:
            if not raw:
                continue
            k = str(raw.get(key_field, "") or "").strip()
            v = str(raw.get(value_field, "") or "").strip()
            if not k or k.startswith("#"):
                continue
            if not v or v.startswith("#"):
                # Treat empty target as unmapped.
                continue
            out[k] = v
    return out


def ensure_mapping_templates(project_dir: Path, *, properties: list[str], edge_codes: list[str]) -> dict:
    """Create mapping CSVs without overwriting canonical mapping files."""

    mappings_dir = project_dir / "internal" / "mappings"
    mappings_dir.mkdir(parents=True, exist_ok=True)

    property_map_path = mappings_dir / "property_mapping.csv"
    account_map_path = mappings_dir / "account_mapping.csv"
    property_code_map_path = mappings_dir / "property_code_mapping.csv"
    missing_code_account_map_path = mappings_dir / "missing_code_account_mapping.csv"
    property_assignee_map_path = mappings_dir / "property_assignee.csv"

    # Only create templates if missing. We don't auto-overwrite canonical mappings.
    if not property_map_path.exists():
        property_map_path.write_text("property_name,yardi_entity\n", encoding="utf-8")

    # Optional overrides. We intentionally do not auto-populate with all seen codes.
    if not account_map_path.exists():
        account_map_path.write_text("edge_code,yardi_account\n", encoding="utf-8")

    # Used for property-code reference (authoritative source is typically an Excel export).
    # Includes optional pdf_property column for months where the PDF naming differs.
    if not property_code_map_path.exists():
        property_code_map_path.write_text("property_code,property_name,pdf_property\n", encoding="utf-8")

    # Used to fill *blank* PDF code rows by (Category, Subcategory, Description) -> yardi_account.
    if not missing_code_account_map_path.exists():
        missing_code_account_map_path.write_text(
            "category,subcategory,description,yardi_account\n",
            encoding="utf-8",
        )

    # Optional: property_code -> assignee (used for splitting exports, e.g. Jay-only import files).
    if not property_assignee_map_path.exists():
        property_assignee_map_path.write_text("property_code,assignee\n", encoding="utf-8")

    return {
        "property_mapping_csv": str(property_map_path),
        "account_mapping_csv": str(account_map_path),
        "property_code_mapping_csv": str(property_code_map_path),
        "missing_code_account_mapping_csv": str(missing_code_account_map_path),
        "property_assignee_csv": str(property_assignee_map_path),
    }


@dataclass(frozen=True)
class MissingCodeKey:
    category: str
    subcategory: str
    description: str

    @staticmethod
    def from_row(category: str, subcategory: str, description: str) -> "MissingCodeKey":
        return MissingCodeKey(
            (category or "").strip(),
            (subcategory or "").strip(),
            (description or "").strip(),
        )


def _read_missing_code_account_mapping(path: Path) -> dict[MissingCodeKey, str]:
    if not path.exists():
        return {}

    out: dict[MissingCodeKey, str] = {}
    with path.open("r", newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for raw in reader:
            if not raw:
                continue
            key = MissingCodeKey.from_row(
                str(raw.get("category", "") or ""),
                str(raw.get("subcategory", "") or ""),
                str(raw.get("description", "") or ""),
            )
            yardi_account = str(raw.get("yardi_account", "") or "").strip()
            if not key.category or key.category.startswith("#"):
                continue
            if not yardi_account or yardi_account.startswith("#"):
                continue
            if not key.subcategory or not key.description:
                continue
            out[key] = yardi_account
    return out


def apply_mappings(
    *,
    rows: list[dict],
    property_mapping_csv: Path,
    account_mapping_csv: Path,
    missing_code_account_mapping_csv: Path | None = None,
    property_field: str = "property_name",
    edge_code_field: str = "edge_code",
) -> MappingResult:
    prop_map = _read_mapping_csv(property_mapping_csv, key_field="property_name", value_field="yardi_entity")
    acct_map = _read_mapping_csv(account_mapping_csv, key_field="edge_code", value_field="yardi_account")
    missing_code_map = (
        _read_missing_code_account_mapping(missing_code_account_mapping_csv)
        if missing_code_account_mapping_csv
        else {}
    )

    unmapped_properties_set: set[str] = set()
    unmapped_accounts_set: set[tuple[str, str, str]] = set()

    mapped_rows: list[dict] = []
    for r in rows:
        prop = str(r.get(property_field, "") or "").strip()
        code = str(r.get(edge_code_field, "") or "").strip()

        yardi_entity = prop_map.get(prop)
        if not yardi_entity:
            unmapped_properties_set.add(prop)

        # Account logic:
        # - If code is present, assume it's the Yardi account code unless an override exists.
        # - If code is blank, we can optionally fill it using the missing-code mapping table.
        if code:
            yardi_account = acct_map.get(code) or code
        else:
            category = str(r.get("category", "") or "").strip()
            subcategory = str(r.get("subcategory", "") or "").strip()
            desc = str(r.get("description", "") or "").strip()

            key = MissingCodeKey.from_row(category, subcategory, desc)
            yardi_account = missing_code_map.get(key, "")
            if not yardi_account:
                unmapped_accounts_set.add((category, subcategory, desc))

        rr = dict(r)
        rr["yardi_entity"] = yardi_entity or ""
        rr["yardi_account"] = yardi_account or ""
        mapped_rows.append(rr)

    return MappingResult(
        mapped_rows=mapped_rows,
        unmapped_properties=sorted(p for p in unmapped_properties_set if p),
        unmapped_accounts=[
            {"category": c, "subcategory": s, "description": d}
            for (c, s, d) in sorted(unmapped_accounts_set, key=lambda t: (t[0].lower(), t[1].lower(), t[2].lower()))
            if any([c, s, d])
        ],
    )


def write_mapping_report_xlsx(
    *,
    output_path: Path,
    unmapped_properties: list[str],
    unmapped_accounts: list[dict],
    mapping_paths: dict,
) -> str:
    wb = Workbook()

    ws = cast(Worksheet, wb.active)
    ws.title = "README"
    ws.append(["How to fix mappings"])
    ws.append(["1) Fill out the mapping CSVs in internal/mappings/"])
    ws.append(["2) Re-run the workflow to generate the final import CSV"])
    ws.append([])
    ws.append(["property_mapping.csv", mapping_paths.get("property_mapping_csv")])
    ws.append(["account_mapping.csv (optional overrides)", mapping_paths.get("account_mapping_csv")])
    ws.append(["missing_code_account_mapping.csv (blank-code rows)", mapping_paths.get("missing_code_account_mapping_csv")])
    ws.append(["property_code_mapping.csv (property codes reference)", mapping_paths.get("property_code_mapping_csv")])
    ws.append(["property_assignee.csv (optional split exports)", mapping_paths.get("property_assignee_csv")])

    ws1 = wb.create_sheet("Unmapped Properties")
    ws1.append(["property_name"])
    for p in unmapped_properties:
        ws1.append([p])

    ws2 = wb.create_sheet("Unmapped Accounts")
    ws2.append(["Category", "Subcategory", "Description"])
    for a in unmapped_accounts:
        ws2.append([
            str(a.get("category", "") or ""),
            str(a.get("subcategory", "") or ""),
            str(a.get("description", "") or ""),
        ])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(output_path)
        return str(output_path)
    except PermissionError:
        for i in range(1, 1000):
            alt = output_path.with_name(f"{output_path.stem} ({i}){output_path.suffix}")
            if not alt.exists():
                wb.save(alt)
                return str(alt)
        raise
