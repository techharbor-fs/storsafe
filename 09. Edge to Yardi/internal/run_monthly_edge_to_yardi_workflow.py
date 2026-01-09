"""Run the monthly Edge → Yardi workflow (Phase 1: rename + inspect).

This is the real workflow runner. `RUN_ME.py` is just a friendly launcher.

Current scope:
- Optional: rename/move the month’s Edge PDF into a deterministic filename
- Inspect PDF structure and output artifacts (property page ranges, basic checks)

Default behavior is dry-run; pass --confirm to apply file renames.
"""

from __future__ import annotations

import argparse
import csv
import filecmp
import json
import re
import sys
import shutil
from dataclasses import asdict, dataclass
from decimal import Decimal
from datetime import datetime, timedelta
from pathlib import Path

import fitz  # PyMuPDF

from edge_gl_pdf_parser import EdgeGlRow, extract_rows_from_pdf
from mapping import apply_mappings, ensure_mapping_templates, write_mapping_report_xlsx


PROPERTY_HEADER_RE = re.compile(r"^(StorSafe of .+?)\s+-\s+(.+)$", re.IGNORECASE)


@dataclass
class PropertySection:
    property_name: str
    header_line: str
    start_page: int  # 1-based
    end_page: int  # 1-based


def _parse_report_month(text: str) -> str:
    raw = (text or "").strip()
    if not re.fullmatch(r"\d{4}-\d{2}", raw):
        raise ValueError("report-month must be YYYY-MM")
    year, mon = raw.split("-", 1)
    if not (1 <= int(mon) <= 12):
        raise ValueError("report-month month must be 01..12")
    return raw


def _month_folder_dir(project_dir: Path, month_folder: str) -> Path:
    d = (project_dir / month_folder).resolve()
    if not d.is_dir():
        raise FileNotFoundError(f"Month folder not found: {d}")
    return d


def _find_candidate_pdfs(month_dir: Path) -> list[Path]:
    candidates: list[Path] = []
    for base in [month_dir, month_dir / "Input"]:
        if not base.exists() or not base.is_dir():
            continue
        for p in base.glob("*.pdf"):
            if p.name.lower().endswith(".pdf"):
                candidates.append(p)
    # de-dupe
    unique = {p.resolve(): p for p in candidates}
    out = list(unique.values())
    out.sort(key=lambda p: p.name.lower())
    return out


def _pick_pdf(candidates: list[Path]) -> Path:
    if not candidates:
        raise FileNotFoundError("No PDF files found in month folder (or Input/)")
    if len(candidates) == 1:
        return candidates[0]

    # Prefer filenames containing general ledger tokens.
    scored: list[tuple[int, Path]] = []
    for p in candidates:
        name = p.name.lower()
        score = 0
        if "general" in name and "ledger" in name:
            score += 10
        if "combined" in name:
            score += 3
        scored.append((score, p))

    scored.sort(key=lambda t: (-t[0], t[1].name.lower()))
    if scored[0][0] == scored[1][0]:
        raise RuntimeError(
            "Multiple PDFs found and none is clearly the general ledger PDF. "
            "Specify --pdf explicitly."
        )
    return scored[0][1]


def _target_pdf_name(*, report_month: str) -> str:
    # Keep this short and stable; this is the canonical monthly Edge GL input.
    return f"edge_gl_{report_month}.pdf"


def _legacy_target_pdf_name(*, report_month: str) -> str:
    # Backward compatibility for earlier runs.
    return f"edge_general_ledger_{report_month}.pdf"


def _rename_pdf_if_needed(
    *,
    pdf_path: Path,
    month_dir: Path,
    report_month: str,
    confirm: bool,
) -> tuple[Path, dict]:
    """Ensure the PDF lives in month_dir/Input with a deterministic filename."""

    input_dir = month_dir / "Input"
    target = input_dir / _target_pdf_name(report_month=report_month)
    legacy_target = input_dir / _legacy_target_pdf_name(report_month=report_month)

    result = {
        "source": str(pdf_path),
        "target": str(target),
        "did_rename": False,
        "skipped_reason": None,
    }

    # Already correct
    try:
        if pdf_path.resolve() == target.resolve():
            result["skipped_reason"] = "already_named_and_located"
            return pdf_path, result
        if pdf_path.resolve() == legacy_target.resolve():
            result["skipped_reason"] = "already_named_and_located_legacy"
            return pdf_path, result
    except Exception:
        pass

    if target.exists():
        result["skipped_reason"] = "target_exists_using_target"

        # Best-effort: if the picked PDF is a duplicate in the month root,
        # move/copy it into Input so the month folder stays tidy.
        if confirm:
            try:
                try:
                    already_in_input = pdf_path.parent.resolve() == input_dir.resolve()
                except Exception:
                    already_in_input = False

                if not already_in_input and pdf_path.exists():
                    # If it's identical to the canonical input, just remove the duplicate.
                    try:
                        if filecmp.cmp(pdf_path, target, shallow=False):
                            try:
                                pdf_path.unlink()
                                result["source_archived_action"] = "deleted_duplicate"
                                result["source_archived_to"] = None
                                return target, result
                            except PermissionError as e:
                                result["source_archive_warning"] = f"Duplicate source PDF is locked; could not delete: {e}"
                                return target, result
                    except Exception as e:
                        result["source_archive_warning"] = f"Could not compare source PDF to canonical input: {e}"

                    base = input_dir / f"{_target_pdf_name(report_month=report_month).removesuffix('.pdf')}__source.pdf"
                    extra = base
                    if extra.exists():
                        for i in range(1, 1000):
                            candidate = input_dir / f"{base.stem}__alt{i}{base.suffix}"
                            if not candidate.exists():
                                extra = candidate
                                break
                    try:
                        pdf_path.replace(extra)
                        result["source_archived_action"] = "moved"
                        result["source_archived_to"] = str(extra)
                    except PermissionError as e:
                        shutil.copy2(pdf_path, extra)
                        result["source_archived_action"] = "copied"
                        result["source_archived_to"] = str(extra)
                        result["source_archive_warning"] = f"Move failed (file likely open): {e}"
            except Exception as e:
                result["source_archive_warning"] = f"Could not archive duplicate source PDF: {e}"

        return target, result
    if legacy_target.exists():
        result["skipped_reason"] = "target_exists_using_legacy_target"
        return legacy_target, result

    if not confirm:
        result["skipped_reason"] = "dry_run"
        return pdf_path, result

    input_dir.mkdir(parents=True, exist_ok=True)

    # Rename/move into Input.
    try:
        moved = pdf_path.replace(target)
        result["did_rename"] = True
        return moved, result
    except PermissionError as e:
        # Common on Windows when the PDF is open in another process.
        # Fall back to copying so the workflow can continue.
        shutil.copy2(pdf_path, target)
        result["did_rename"] = False
        result["skipped_reason"] = "file_locked_copied_instead"
        result["warning"] = f"Move failed (file likely open): {e}"
        return target, result


def _extract_property_sections(doc: fitz.Document) -> tuple[list[PropertySection], list[str]]:
    """Detect property header pages and return page ranges."""

    warnings: list[str] = []

    starts: list[tuple[int, str, str]] = []  # (page_idx0, property_name, header_line)
    for i in range(doc.page_count):
        text = str(doc.load_page(i).get_text("text") or "")
        header_line = None
        property_name = None
        for line in text.splitlines():
            line = (line or "").strip()
            if not line:
                continue
            m = PROPERTY_HEADER_RE.match(line)
            if m:
                property_name = m.group(1).strip()
                header_line = line
                break
        if property_name and header_line:
            starts.append((i, property_name, header_line))

    if not starts:
        warnings.append("No property header lines found (expected 'StorSafe of <X> - <address>')")
        return [], warnings

    # Build page ranges
    starts.sort(key=lambda t: t[0])
    sections: list[PropertySection] = []
    for idx, (page0, prop, header) in enumerate(starts):
        next_start_page0 = starts[idx + 1][0] if idx + 1 < len(starts) else doc.page_count
        start_1 = page0 + 1
        end_1 = next_start_page0
        if end_1 < start_1:
            warnings.append(f"Invalid page range for {prop}: {start_1}-{end_1}")
            continue
        sections.append(PropertySection(prop, header, start_1, end_1))

    return sections, warnings


def _check_expected_headers(doc: fitz.Document, *, max_pages_to_scan: int = 5) -> dict:
    expected = ["Category", "Subcategory", "Description", "Code", "Debit", "Credit"]
    found_any = {k: False for k in expected}
    scanned = min(doc.page_count, max_pages_to_scan)
    for i in range(scanned):
        text = str(doc.load_page(i).get_text("text") or "")
        for k in expected:
            if k in text:
                found_any[k] = True
    return {
        "scanned_pages": scanned,
        "expected_headers": expected,
        "found": found_any,
        "all_found": all(found_any.values()) if expected else True,
    }


def _month_helper_artifacts_dir(project_dir: Path, *, month_dir: Path) -> Path:
    """Project-level helper artifacts folder for this month.

    Keeps month folder clean: only Input/ and Output/ deliverables live there.
    """

    root = project_dir.parent / ".helper_artifacts" / "09. Edge to Yardi" / month_dir.name
    root.mkdir(parents=True, exist_ok=True)
    return root


def _write_artifacts(
    *,
    artifacts_dir: Path,
    summary: dict,
    preview_text: str,
) -> dict:
    artifacts_dir.mkdir(parents=True, exist_ok=True)

    summary_path = artifacts_dir / "inspection_summary.json"
    preview_path = artifacts_dir / "inspection_preview.txt"

    summary_path.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    preview_path.write_text(preview_text, encoding="utf-8")

    return {
        "summary_json": str(summary_path),
        "preview_txt": str(preview_path),
    }


def _write_normalized_csv(artifacts_dir: Path, *, report_month: str, rows: list[EdgeGlRow]) -> str:
    artifacts_dir.mkdir(parents=True, exist_ok=True)
    path = artifacts_dir / "edge_normalized.csv"

    def _open_with_fallback(p: Path):
        try:
            return p.open("w", newline="", encoding="utf-8"), p
        except PermissionError:
            # File is likely open/locked (Excel). Fall back to Windows-style duplicates.
            for i in range(1, 1000):
                alt = p.with_name(f"{p.stem} ({i}){p.suffix}")
                if not alt.exists():
                    return alt.open("w", newline="", encoding="utf-8"), alt
            raise

    f, actual_path = _open_with_fallback(path)
    with f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "property_name",
                "category",
                "subcategory",
                "description",
                "edge_code",
                "debit",
                "credit",
                "source_page",
            ],
        )
        w.writeheader()
        for r in rows:
            w.writerow(
                {
                    "property_name": r.property_name,
                    "category": r.category,
                    "subcategory": r.subcategory,
                    "description": r.description,
                    "edge_code": r.edge_code,
                    "debit": f"{r.debit:.2f}",
                    "credit": f"{r.credit:.2f}",
                    "source_page": r.source_page,
                }
            )

    return str(actual_path)


def _slug(text: str) -> str:
    import re

    t = (text or "").strip().lower()
    t = re.sub(r"[^a-z0-9]+", "_", t).strip("_")
    return t or "unknown"


def _write_per_property_tables(
    artifacts_dir: Path,
    *,
    report_month: str,
    rows: list[EdgeGlRow],
    yardi_account_by_edge_code: dict[str, str] | None = None,
) -> dict:
    """Write per-property CSVs matching the PDF table headers.

    Headers must match:
      Category | Subcategory | Description | Yardi Code | Debit | Credit
    """

    out_dir = artifacts_dir / "parsed_tables"
    out_dir.mkdir(parents=True, exist_ok=True)

    grouped: dict[str, list[EdgeGlRow]] = {}
    for r in rows:
        grouped.setdefault(r.property_name, []).append(r)

    paths: dict[str, str] = {}
    for property_name, items in sorted(grouped.items(), key=lambda t: t[0].lower()):
        filename = f"edge_gl_table_{report_month}__{_slug(property_name)}.csv"
        path = out_dir / filename

        with path.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(
                f,
                fieldnames=["Category", "Subcategory", "Description", "Yardi Code", "Debit", "Credit"],
            )
            w.writeheader()
            for r in items:
                yardi_code = r.edge_code
                if yardi_account_by_edge_code is not None and r.edge_code:
                    mapped = yardi_account_by_edge_code.get(r.edge_code)
                    if mapped:
                        yardi_code = mapped
                w.writerow(
                    {
                        "Category": r.category,
                        "Subcategory": r.subcategory,
                        "Description": r.description,
                        "Yardi Code": yardi_code,
                        "Debit": f"{r.debit:.2f}",
                        "Credit": f"{r.credit:.2f}",
                    }
                )

        paths[property_name] = str(path)

    return {"parsed_tables_dir": str(out_dir), "per_property_csvs": paths}


def _draft_yardi_import_filename(report_month: str) -> str:
    # Deterministic, explicit that it's draft.
    return f"yardi_gl_import_edge_{report_month}_draft.csv"


def _final_yardi_import_filename(report_month: str) -> str:
    return f"yardi_gl_import_edge_{report_month}.csv"


def _write_draft_yardi_import(
    month_dir: Path,
    *,
    artifacts_dir: Path,
    report_month: str,
    rows: list[EdgeGlRow],
) -> tuple[str, dict]:
    """Generate a draft GL import CSV.

    This is intentionally mapping-free: account/entity fields are placeholders
    until Phase 3 mapping is implemented.
    """

    # Draft is a debug artifact; keep month Output clean.
    artifacts_dir.mkdir(parents=True, exist_ok=True)
    path = artifacts_dir / _draft_yardi_import_filename(report_month)

    total_debit = Decimal("0")
    total_credit = Decimal("0")

    try:
        f = path.open("w", newline="", encoding="utf-8")
        actual_path = path
    except PermissionError:
        actual_path = None
        for i in range(1, 1000):
            alt = path.with_name(f"{path.stem} ({i}){path.suffix}")
            if not alt.exists():
                actual_path = alt
                break
        if actual_path is None:
            raise
        f = actual_path.open("w", newline="", encoding="utf-8")

    with f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "batch_id",
                "report_month",
                "entity",  # placeholder (property -> yardi entity)
                "account",  # placeholder (edge_code -> yardi account)
                "description",
                "debit",
                "credit",
                "source_property",
                "source_page",
            ],
        )
        w.writeheader()

        batch_id = f"EDGE_GL_{report_month.replace('-', '')}"
        for r in rows:
            total_debit += r.debit
            total_credit += r.credit
            w.writerow(
                {
                    "batch_id": batch_id,
                    "report_month": report_month,
                    "entity": r.property_name,  # TEMP
                    "account": r.edge_code,  # TEMP
                    "description": f"{r.category} | {r.subcategory} | {r.description}".strip(" |"),
                    "debit": f"{r.debit:.2f}",
                    "credit": f"{r.credit:.2f}",
                    "source_property": r.property_name,
                    "source_page": r.source_page,
                }
            )

    totals = {
        "row_count": len(rows),
        "total_debit": f"{total_debit:.2f}",
        "total_credit": f"{total_credit:.2f}",
        "balanced": total_debit == total_credit,
    }

    # Control totals are an operational output worth keeping in month Output.
    out_dir = month_dir / "Output"
    out_dir.mkdir(parents=True, exist_ok=True)
    totals_path = out_dir / "control_totals.json"
    try:
        totals_path.write_text(json.dumps(totals, indent=2), encoding="utf-8")
        actual_totals_path = totals_path
    except PermissionError:
        actual_totals_path = None
        for i in range(1, 1000):
            alt = totals_path.with_name(f"{totals_path.stem} ({i}){totals_path.suffix}")
            if not alt.exists():
                actual_totals_path = alt
                break
        if actual_totals_path is None:
            raise
        actual_totals_path.write_text(json.dumps(totals, indent=2), encoding="utf-8")

    return str(actual_path), {"control_totals_json": str(actual_totals_path), **totals}


def _write_final_yardi_import(
    month_dir: Path,
    *,
    artifacts_dir: Path,
    report_month: str,
    mapped_rows: list[dict],
) -> str:
    # Internal-format output is for analysis/debug; keep month Output clean.
    artifacts_dir.mkdir(parents=True, exist_ok=True)
    path = artifacts_dir / _final_yardi_import_filename(report_month)

    batch_id = f"EDGE_GL_{report_month.replace('-', '')}"

    try:
        f = path.open("w", newline="", encoding="utf-8")
        actual_path = path
    except PermissionError:
        actual_path = None
        for i in range(1, 1000):
            alt = path.with_name(f"{path.stem} ({i}){path.suffix}")
            if not alt.exists():
                actual_path = alt
                break
        if actual_path is None:
            raise
        f = actual_path.open("w", newline="", encoding="utf-8")

    with f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "batch_id",
                "report_month",
                "entity",
                "account",
                "description",
                "debit",
                "credit",
                "source_property",
                "source_page",
            ],
        )
        w.writeheader()
        for r in mapped_rows:
            w.writerow(
                {
                    "batch_id": batch_id,
                    "report_month": report_month,
                    "entity": r.get("yardi_entity", ""),
                    "account": r.get("yardi_account", ""),
                    "description": r.get("description", ""),
                    "debit": r.get("debit", ""),
                    "credit": r.get("credit", ""),
                    "source_property": r.get("source_property", ""),
                    "source_page": r.get("source_page", ""),
                }
            )

    return str(actual_path)


def _find_detail_template_csv(month_dir: Path) -> Path | None:
    # Template input typically has no header and lives in Input/.
    candidates: list[Path] = []
    for base in [month_dir / "Input", month_dir]:
        if not base.exists():
            continue
        for p in base.glob("*.csv"):
            name = p.name.lower()
            if "general" in name and "ledger" in name and "detail" in name:
                candidates.append(p)

    if not candidates:
        return None
    candidates.sort(key=lambda p: p.name.lower())
    return candidates[0]


def _read_description_to_account(mapping_csv: Path) -> dict[str, str]:
    """Build a lookup: description -> yardi_account.

    The canonical mapping file is category/subcategory/description -> yardi_account,
    but the detail template CSV only carries the description label (col O).

    If a description maps to multiple accounts, we skip it to avoid wrong fills.
    """

    import re

    trailing_acct_re = re.compile(r"\s+\d{4}-\d{1,4}$")

    if not mapping_csv.exists():
        return {}

    seen: dict[str, set[str]] = {}
    with mapping_csv.open("r", newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for r in reader:
            desc = str(r.get("description") or r.get("Description") or "").strip()
            acct = str(
                r.get("yardi_account")
                or r.get("Yardi Code")
                or r.get("Yardi Account")
                or r.get("account")
                or ""
            ).strip()
            if not desc or not acct or desc.startswith("#") or acct.startswith("#"):
                continue

            # Some mapping sources (and/or prior PDF parsing) may have a trailing account token
            # accidentally appended to the description. Generate a cleaned variant so we still
            # match the detail template's description-only labels.
            desc_variants = {desc}
            cleaned = trailing_acct_re.sub("", desc).strip()
            if cleaned and cleaned != desc:
                desc_variants.add(cleaned)

            for dv in desc_variants:
                seen.setdefault(dv, set()).add(acct)

    out: dict[str, str] = {}
    for desc, accts in seen.items():
        if len(accts) == 1:
            out[desc] = next(iter(accts))
    return out


def _write_yardi_import_from_detail_template(
    *,
    month_dir: Path,
    artifacts_dir: Path,
    report_month: str,
    detail_csv_path: Path,
    mapped_rows: list[dict],
    missing_code_account_mapping_csv: Path,
) -> tuple[str, dict]:
    """Create a Yardi import CSV using the detail template layout.

    We preserve the 15-column layout of the production detail CSV.
    Changes:
    - Column I (index 8) is set to the *correct* property code (yardi_entity)
      by matching the template's property-id buckets to PDF properties by account totals.
    - Account code column (index 10) is filled if blank using missing-code mapping.
    """

    import re

    valid_yardi_acct_re = re.compile(r"^\d{4}-\d{4}$")

    def is_valid_yardi_account(code: str) -> bool:
        return bool(valid_yardi_acct_re.match((code or "").strip()))

    def _build_pdf_totals(
        rows: list[dict],
    ) -> tuple[dict[tuple[str, str], Decimal], set[str], dict[str, str]]:
        pdf_totals: dict[tuple[str, str], Decimal] = {}
        pdf_properties: set[str] = set()
        pdf_property_to_yardi_entity: dict[str, str] = {}
        for r in rows:
            prop = str(r.get("property_name", "") or r.get("source_property", "") or "").strip()
            acct = str(r.get("yardi_account", "") or "").strip()
            if not prop:
                continue
            pdf_properties.add(prop)
            if prop and str(r.get("yardi_entity", "") or "").strip():
                pdf_property_to_yardi_entity[prop] = str(r.get("yardi_entity", "") or "").strip()
            if not acct:
                continue
            debit = Decimal(str(r.get("debit", "0") or "0"))
            credit = Decimal(str(r.get("credit", "0") or "0"))
            net = debit - credit
            pdf_totals[(prop, acct)] = pdf_totals.get((prop, acct), Decimal("0")) + net
        return pdf_totals, pdf_properties, pdf_property_to_yardi_entity

    pdf_totals, pdf_properties, pdf_property_to_yardi_entity = _build_pdf_totals(mapped_rows)

    def _read_detail_template(
        path: Path,
    ) -> tuple[list[list[str]], set[str], dict[tuple[str, str], Decimal]]:
        detail_totals: dict[tuple[str, str], Decimal] = {}
        detail_prop_ids: set[str] = set()
        raw_rows: list[list[str]] = []

        last_prop_id: str = ""

        with path.open("r", newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            for row in reader:
                if not row:
                    continue
                if len(row) < 15:
                    # Skip malformed lines.
                    continue
                row = list(row[:15])

                # Edge exports often omit the property bucket ID on subsequent lines.
                # Forward-fill column I so every row belongs to a property bucket.
                prop_id = (row[8] or "").strip()
                if not prop_id and last_prop_id:
                    prop_id = last_prop_id
                    row[8] = prop_id
                elif prop_id:
                    last_prop_id = prop_id

                raw_rows.append(row)
                acct = (row[10] or "").strip()
                amt_raw = (row[9] or "0").strip()
                if prop_id:
                    detail_prop_ids.add(prop_id)
                if not (prop_id and acct):
                    continue
                try:
                    amt = Decimal(amt_raw)
                except Exception:
                    continue
                detail_totals[(prop_id, acct)] = detail_totals.get((prop_id, acct), Decimal("0")) + amt

        return raw_rows, detail_prop_ids, detail_totals

    raw_rows, detail_prop_ids, detail_totals = _read_detail_template(detail_csv_path)

    def _map_detail_props_to_pdf_props(
        *,
        detail_prop_ids: set[str],
        detail_totals: dict[tuple[str, str], Decimal],
        pdf_props: set[str],
        pdf_totals: dict[tuple[str, str], Decimal],
        pdf_property_to_yardi_entity: dict[str, str],
    ) -> tuple[dict[str, str], list[dict]]:
        detail_props = sorted(detail_prop_ids)
        pdf_props_sorted = sorted(pdf_props)

        # Pre-index account sets for overlap calculations.
        acct_by_detail_prop: dict[str, set[str]] = {p: set() for p in detail_props}
        for (p, acct) in detail_totals.keys():
            acct_by_detail_prop.setdefault(p, set()).add(acct)

        acct_by_pdf_prop: dict[str, set[str]] = {p: set() for p in pdf_props_sorted}
        for (p, acct) in pdf_totals.keys():
            acct_by_pdf_prop.setdefault(p, set()).add(acct)

        pair_scores: list[tuple[Decimal, int, str, str]] = []
        for dp in detail_props:
            for pp in pdf_props_sorted:
                overlap = acct_by_detail_prop.get(dp, set()) & acct_by_pdf_prop.get(pp, set())
                if len(overlap) < 3:
                    continue
                err = Decimal("0")
                for acct in overlap:
                    a = detail_totals.get((dp, acct), Decimal("0"))
                    b = pdf_totals.get((pp, acct), Decimal("0"))
                    err += abs(a - b)
                pair_scores.append((err, len(overlap), dp, pp))

        pair_scores.sort(key=lambda t: (t[0], -t[1], t[2], t[3]))

        detail_to_pdf: dict[str, str] = {}
        diagnostics: list[dict] = []

        if len(detail_props) == len(pdf_props_sorted) and pair_scores:
            used_detail: set[str] = set()
            used_pdf: set[str] = set()
            for err, overlap_n, dp, pp in pair_scores:
                if dp in used_detail or pp in used_pdf:
                    continue
                used_detail.add(dp)
                used_pdf.add(pp)
                detail_to_pdf[dp] = pp
                diagnostics.append(
                    {
                        "detail_property_id": dp,
                        "pdf_property": pp,
                        "overlap_accounts": overlap_n,
                        "total_abs_error": f"{err}",
                        "yardi_entity": pdf_property_to_yardi_entity.get(pp, ""),
                    }
                )
        else:
            # Non-equal counts: map each detail prop independently to best score.
            best_by_detail: dict[str, tuple[Decimal, int, str]] = {}
            for err, overlap_n, dp, pp in pair_scores:
                cur = best_by_detail.get(dp)
                if cur is None or (err, -overlap_n) < (cur[0], -cur[1]):
                    best_by_detail[dp] = (err, overlap_n, pp)

            for dp in detail_props:
                if dp in best_by_detail:
                    err, overlap_n, pp = best_by_detail[dp]
                    detail_to_pdf[dp] = pp
                    diagnostics.append(
                        {
                            "detail_property_id": dp,
                            "pdf_property": pp,
                            "overlap_accounts": overlap_n,
                            "total_abs_error": f"{err}",
                            "yardi_entity": pdf_property_to_yardi_entity.get(pp, ""),
                        }
                    )

        return detail_to_pdf, diagnostics

    detail_to_pdf, diagnostics = _map_detail_props_to_pdf_props(
        detail_prop_ids=detail_prop_ids,
        detail_totals=detail_totals,
        pdf_props=pdf_properties,
        pdf_totals=pdf_totals,
        pdf_property_to_yardi_entity=pdf_property_to_yardi_entity,
    )

    # Build account-fill lookup (description -> account).
    # 1) Month-level mapping (authoritative for the detail template)
    month_account_code_mapping_csv = month_dir / "Input" / "account code mapping.csv"
    desc_to_acct = _read_description_to_account(month_account_code_mapping_csv)

    # 2) Explicit missing-code mapping (curated fallback)
    curated_desc_to_acct = _read_description_to_account(missing_code_account_mapping_csv)
    for d, a in curated_desc_to_acct.items():
        desc_to_acct.setdefault(d, a)

    # 3) Fallback: infer from the PDF-mapped rows if description is unique.
    pdf_desc_to_accounts: dict[str, set[str]] = {}
    for r in mapped_rows:
        desc = str(r.get("description", "") or "").strip()
        acct = str(r.get("yardi_account", "") or "").strip()
        if not desc or not acct:
            continue
        pdf_desc_to_accounts.setdefault(desc, set()).add(acct)
    pdf_desc_to_acct_unique: dict[str, str] = {
        d: next(iter(accts)) for d, accts in pdf_desc_to_accounts.items() if len(accts) == 1
    }

    filled_property_count = 0
    filled_account_count = 0
    corrected_account_count = 0
    unmapped_property_ids: set[str] = set()
    unmapped_account_rows = 0

    out_rows: list[list[str]] = []
    for row in raw_rows:
        dp = (row[8] or "").strip()
        acct = (row[10] or "").strip()
        desc = (row[14] or "").strip()

        # Fill property code in column I.
        pp = detail_to_pdf.get(dp)
        if pp:
            yardi_entity = pdf_property_to_yardi_entity.get(pp, "")
            if yardi_entity:
                if row[8] != yardi_entity:
                    row[8] = yardi_entity
                filled_property_count += 1
            else:
                unmapped_property_ids.add(dp)
        else:
            if dp:
                unmapped_property_ids.add(dp)

        # Fill/correct account code column K based on description.
        # - If blank: fill
        # - If present but invalid format (e.g., '5000-001'): correct
        # - If present but differs from mapping: prefer mapping (authoritative)
        mapped_acct = desc_to_acct.get(desc, "") or pdf_desc_to_acct_unique.get(desc, "")
        if mapped_acct:
            if (not acct) or (not is_valid_yardi_account(acct)) or (acct != mapped_acct):
                if not acct:
                    filled_account_count += 1
                else:
                    corrected_account_count += 1
                row[10] = mapped_acct
        else:
            if (not acct) or (not is_valid_yardi_account(acct)):
                unmapped_account_rows += 1

        out_rows.append(row)

    out_dir = month_dir / "Output"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"yardi_import_template_{report_month}.csv"

    try:
        f = out_path.open("w", newline="", encoding="utf-8")
        actual_path = out_path
    except PermissionError:
        actual_path = None
        for i in range(1, 1000):
            alt = out_path.with_name(f"{out_path.stem} ({i}){out_path.suffix}")
            if not alt.exists():
                actual_path = alt
                break
        if actual_path is None:
            raise
        f = actual_path.open("w", newline="", encoding="utf-8")

    with f:
        writer = csv.writer(f)
        writer.writerows(out_rows)

    artifacts_dir.mkdir(parents=True, exist_ok=True)
    mapping_path = artifacts_dir / f"detail_property_id_to_pdf_property_{report_month}.csv"
    try:
        mf = mapping_path.open("w", newline="", encoding="utf-8")
        mapping_actual = mapping_path
    except PermissionError:
        mapping_actual = None
        for i in range(1, 1000):
            alt = mapping_path.with_name(f"{mapping_path.stem} ({i}){mapping_path.suffix}")
            if not alt.exists():
                mapping_actual = alt
                break
        if mapping_actual is None:
            raise
        mf = mapping_actual.open("w", newline="", encoding="utf-8")

    with mf:
        w = csv.DictWriter(
            mf,
            fieldnames=[
                "detail_property_id",
                "pdf_property",
                "yardi_entity",
                "overlap_accounts",
                "total_abs_error",
            ],
        )
        w.writeheader()
        for d in diagnostics:
            w.writerow(d)

    stats = {
        "detail_template_input": str(detail_csv_path),
        "yardi_import_template_csv": str(actual_path),
        "detail_property_id_mapping_csv": str(mapping_actual),
        "detail_property_ids_count": len(detail_prop_ids),
        "mapped_detail_property_ids_count": len({d.get('detail_property_id') for d in diagnostics}),
        "unmapped_detail_property_ids": sorted(unmapped_property_ids),
        "filled_property_rows_count": filled_property_count,
        "filled_account_rows_count": filled_account_count,
        "corrected_account_rows_count": corrected_account_count,
        "unmapped_account_rows_count": unmapped_account_rows,
    }

    return str(actual_path), stats


def _write_yardi_import_template_from_pdf(
    *,
    month_dir: Path,
    artifacts_dir: Path,
    report_month: str,
    mapped_rows: list[dict],
) -> tuple[str, dict]:
    """Create the 15-column "template" CSV directly from PDF-derived mapped rows.

    The PDF is authoritative; it does not provide per-transaction dates. We therefore
    use a single posting date (month end) and a single period start (month start)
    across all rows.

    Output matches the existing 15-column layout used by the prior detail-template
    based generator:
      0: 'J' (constant)
      1: day-of-month for posting date
      2-3: blank
      4: posting date (MM/DD/YYYY)
      5: period start (MM/DD/YYYY)
      6: blank
      7: 'Edge General Ledger' (constant)
      8: yardi_entity (property code)
      9: net amount (debit - credit)
     10: yardi_account
     11-12: blank
     13: '1000' (constant)
     14: description
    """

    year_s, mon_s = report_month.split("-", 1)
    year = int(year_s)
    month = int(mon_s)
    month_start = datetime(year, month, 1)
    # Month end: first day of next month minus 1 day.
    if month == 12:
        next_month = datetime(year + 1, 1, 1)
    else:
        next_month = datetime(year, month + 1, 1)
    month_end = next_month - timedelta(days=1)

    posting_date = month_end.strftime("%m/%d/%Y")
    posting_day = str(month_end.day)
    period_start = month_start.strftime("%m/%d/%Y")

    out_rows: list[list[str]] = []
    skipped_unmapped = 0
    for r in mapped_rows:
        yardi_entity = str(r.get("yardi_entity", "") or "").strip()
        yardi_account = str(r.get("yardi_account", "") or "").strip()
        desc = str(r.get("description", "") or "").strip()
        if not (yardi_entity and yardi_account):
            skipped_unmapped += 1
            continue
        debit = Decimal(str(r.get("debit", "0") or "0"))
        credit = Decimal(str(r.get("credit", "0") or "0"))
        net = debit - credit

        out_rows.append(
            [
                "J",  # 0
                posting_day,  # 1
                "",  # 2
                "",  # 3
                posting_date,  # 4
                period_start,  # 5
                "",  # 6
                "Edge General Ledger",  # 7
                yardi_entity,  # 8
                f"{net}",  # 9
                yardi_account,  # 10
                "",  # 11
                "",  # 12
                "1000",  # 13
                desc,  # 14
            ]
        )

    out_dir = month_dir / "Output"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"yardi_import_template_{report_month}.csv"

    try:
        f = out_path.open("w", newline="", encoding="utf-8")
        actual_path = out_path
    except PermissionError:
        actual_path = None
        for i in range(1, 1000):
            alt = out_path.with_name(f"{out_path.stem} ({i}){out_path.suffix}")
            if not alt.exists():
                actual_path = alt
                break
        if actual_path is None:
            raise
        f = actual_path.open("w", newline="", encoding="utf-8")

    with f:
        w = csv.writer(f)
        w.writerows(out_rows)

    stats = {
        "source": "pdf",
        "yardi_import_template_csv": str(actual_path),
        "row_count": len(out_rows),
        "skipped_rows_missing_property_or_account": skipped_unmapped,
        "posting_date": posting_date,
        "period_start": period_start,
        "constants": {"col0": "J", "col7": "Edge General Ledger", "col13": "1000"},
    }

    # Keep artifacts dir present for consistency.
    artifacts_dir.mkdir(parents=True, exist_ok=True)
    return str(actual_path), stats


def _compare_detail_csv_to_pdf(
    *,
    month_dir: Path,
    artifacts_dir: Path,
    report_month: str,
    detail_csv_path: Path,
    mapped_rows: list[dict],
    extracted_rows: list[EdgeGlRow],
    tolerance: Decimal,
    focus_property: str | None = None,
) -> tuple[str, dict]:
    """Compare detail template CSV totals to PDF totals.

    'Equal' here means: after mapping the detail CSV's property buckets to PDF properties,
    the summed net amounts per (property, account) match within the tolerance.

    We also compare the grand total across all properties/accounts.
    """

    # Reuse the same internal helpers defined in _write_yardi_import_from_detail_template
    # by calling it in a dry-ish mode isn't possible, so we reimplement the small parts here.
    # Keep logic aligned to the template writer.
    import re

    valid_yardi_acct_re = re.compile(r"^\d{4}-\d{4}$")

    def is_valid_yardi_account(code: str) -> bool:
        return bool(valid_yardi_acct_re.match((code or "").strip()))

    # Build PDF totals by (pdf_property, account)
    pdf_totals: dict[tuple[str, str], Decimal] = {}
    pdf_properties: set[str] = set()
    pdf_property_to_yardi_entity: dict[str, str] = {}
    for r in mapped_rows:
        prop = str(r.get("property_name", "") or r.get("source_property", "") or "").strip()
        acct = str(r.get("yardi_account", "") or "").strip()
        if not prop:
            continue
        pdf_properties.add(prop)
        if prop and str(r.get("yardi_entity", "") or "").strip():
            pdf_property_to_yardi_entity[prop] = str(r.get("yardi_entity", "") or "").strip()
        if not acct:
            continue
        debit = Decimal(str(r.get("debit", "0") or "0"))
        credit = Decimal(str(r.get("credit", "0") or "0"))
        net = debit - credit
        pdf_totals[(prop, acct)] = pdf_totals.get((prop, acct), Decimal("0")) + net

    # Build lookup to correct/fill detail CSV account codes based on description.
    month_account_code_mapping_csv = month_dir / "Input" / "account code mapping.csv"
    desc_to_acct = _read_description_to_account(month_account_code_mapping_csv)
    curated_missing_code_csv = month_dir.parent / "internal" / "mappings" / "missing_code_account_mapping.csv"
    curated_desc_to_acct = _read_description_to_account(curated_missing_code_csv)
    for d, a in curated_desc_to_acct.items():
        desc_to_acct.setdefault(d, a)

    pdf_desc_to_accounts: dict[str, set[str]] = {}
    for r in mapped_rows:
        desc = str(r.get("description", "") or "").strip()
        acct = str(r.get("yardi_account", "") or "").strip()
        if not desc or not acct:
            continue
        pdf_desc_to_accounts.setdefault(desc, set()).add(acct)
    pdf_desc_to_acct_unique: dict[str, str] = {
        d: next(iter(accts)) for d, accts in pdf_desc_to_accounts.items() if len(accts) == 1
    }

    # Read template totals by (detail_prop_id, account)
    detail_totals_raw: dict[tuple[str, str], Decimal] = {}
    detail_totals_corrected: dict[tuple[str, str], Decimal] = {}
    detail_prop_ids: set[str] = set()
    invalid_account_codes_before: dict[str, int] = {}
    invalid_account_codes_after: dict[str, int] = {}
    corrected_account_rows = 0
    filled_account_rows = 0
    skipped_non_month_rows = 0

    # Filter the detail template to the requested report_month using the period-start column (index 5).
    # The production export uses US-style dates like 12/01/2025 (sometimes without leading zeros).
    year_s, mon_s = report_month.split("-", 1)
    expected_period_start_date = datetime(int(year_s), int(mon_s), 1).date()

    def _parse_us_date(text: str):
        t = (text or "").strip()
        if not t:
            return None
        for fmt in ("%m/%d/%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(t, fmt).date()
            except Exception:
                continue
        return None
    with detail_csv_path.open("r", newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        last_prop_id: str = ""
        for row in reader:
            if not row or len(row) < 15:
                continue
            row = list(row[:15])
            period_start = (row[5] or "").strip()
            if period_start:
                parsed = _parse_us_date(period_start)
                if parsed is not None and parsed != expected_period_start_date:
                    skipped_non_month_rows += 1
                    continue
            prop_id = (row[8] or "").strip()
            if not prop_id and last_prop_id:
                prop_id = last_prop_id
            elif prop_id:
                last_prop_id = prop_id
            acct = (row[10] or "").strip()
            amt_raw = (row[9] or "0").strip()
            desc = (row[14] or "").strip()
            if prop_id:
                detail_prop_ids.add(prop_id)
            if acct and not is_valid_yardi_account(acct):
                invalid_account_codes_before[acct] = invalid_account_codes_before.get(acct, 0) + 1
            if not (prop_id and acct):
                continue
            try:
                amt = Decimal(amt_raw)
            except Exception:
                continue

            detail_totals_raw[(prop_id, acct)] = detail_totals_raw.get((prop_id, acct), Decimal("0")) + amt

            # Normalize account code for corrected comparison.
            new_acct = acct
            mapped_acct = desc_to_acct.get(desc, "") or pdf_desc_to_acct_unique.get(desc, "")
            if mapped_acct:
                if (not new_acct) and mapped_acct:
                    new_acct = mapped_acct
                    filled_account_rows += 1
                elif (not is_valid_yardi_account(new_acct)) or (new_acct != mapped_acct):
                    new_acct = mapped_acct
                    corrected_account_rows += 1

            if new_acct and not is_valid_yardi_account(new_acct):
                invalid_account_codes_after[new_acct] = invalid_account_codes_after.get(new_acct, 0) + 1

            detail_totals_corrected[(prop_id, new_acct)] = (
                detail_totals_corrected.get((prop_id, new_acct), Decimal("0")) + amt
            )

    # Use the exact same mapping function logic from the template writer by calling it indirectly.
    # To avoid circular nesting, we reconstruct it here with the same heuristics.
    # NOTE: keep this aligned with _map_detail_props_to_pdf_props above.
    detail_props = sorted(detail_prop_ids)
    pdf_props_sorted = sorted(pdf_properties)

    acct_by_detail_prop: dict[str, set[str]] = {p: set() for p in detail_props}
    for (p, acct) in detail_totals_corrected.keys():
        acct_by_detail_prop.setdefault(p, set()).add(acct)

    acct_by_pdf_prop: dict[str, set[str]] = {p: set() for p in pdf_props_sorted}
    for (p, acct) in pdf_totals.keys():
        acct_by_pdf_prop.setdefault(p, set()).add(acct)

    pair_scores: list[tuple[Decimal, int, str, str]] = []
    for dp in detail_props:
        for pp in pdf_props_sorted:
            overlap = acct_by_detail_prop.get(dp, set()) & acct_by_pdf_prop.get(pp, set())
            if len(overlap) < 3:
                continue
            err = Decimal("0")
            for acct in overlap:
                a = detail_totals_corrected.get((dp, acct), Decimal("0"))
                b = pdf_totals.get((pp, acct), Decimal("0"))
                err += abs(a - b)
            pair_scores.append((err, len(overlap), dp, pp))

    pair_scores.sort(key=lambda t: (t[0], -t[1], t[2], t[3]))

    detail_to_pdf: dict[str, str] = {}
    diagnostics: list[dict] = []
    if len(detail_props) == len(pdf_props_sorted) and pair_scores:
        used_detail: set[str] = set()
        used_pdf: set[str] = set()
        for err, overlap_n, dp, pp in pair_scores:
            if dp in used_detail or pp in used_pdf:
                continue
            used_detail.add(dp)
            used_pdf.add(pp)
            detail_to_pdf[dp] = pp
            diagnostics.append(
                {
                    "detail_property_id": dp,
                    "pdf_property": pp,
                    "overlap_accounts": overlap_n,
                    "total_abs_error": f"{err}",
                    "yardi_entity": pdf_property_to_yardi_entity.get(pp, ""),
                }
            )
    else:
        best_by_detail: dict[str, tuple[Decimal, int, str]] = {}
        for err, overlap_n, dp, pp in pair_scores:
            cur = best_by_detail.get(dp)
            if cur is None or (err, -overlap_n) < (cur[0], -cur[1]):
                best_by_detail[dp] = (err, overlap_n, pp)
        for dp in detail_props:
            if dp in best_by_detail:
                err, overlap_n, pp = best_by_detail[dp]
                detail_to_pdf[dp] = pp
                diagnostics.append(
                    {
                        "detail_property_id": dp,
                        "pdf_property": pp,
                        "overlap_accounts": overlap_n,
                        "total_abs_error": f"{err}",
                        "yardi_entity": pdf_property_to_yardi_entity.get(pp, ""),
                    }
                )

    # Aggregate detail totals into PDF-property space for comparison.
    detail_mapped_totals: dict[tuple[str, str], Decimal] = {}
    unmapped_detail_props: set[str] = set()
    for (dp, acct), amt in detail_totals_corrected.items():
        pp = detail_to_pdf.get(dp)
        if not pp:
            unmapped_detail_props.add(dp)
            continue
        detail_mapped_totals[(pp, acct)] = detail_mapped_totals.get((pp, acct), Decimal("0")) + amt

    # Compare per (pdf_property, account)
    all_keys = set(pdf_totals.keys()) | set(detail_mapped_totals.keys())
    mismatches: list[dict] = []
    total_abs_diff = Decimal("0")
    for (pp, acct) in sorted(all_keys, key=lambda t: (t[0].lower(), t[1])):
        a = detail_mapped_totals.get((pp, acct), Decimal("0"))
        b = pdf_totals.get((pp, acct), Decimal("0"))
        diff = a - b
        if abs(diff) > tolerance:
            mismatches.append(
                {
                    "pdf_property": pp,
                    "yardi_entity": pdf_property_to_yardi_entity.get(pp, ""),
                    "account": acct,
                    "detail_total": f"{a}",
                    "pdf_total": f"{b}",
                    "diff": f"{diff}",
                }
            )
        total_abs_diff += abs(diff)

    detail_grand = sum(detail_mapped_totals.values(), Decimal("0"))
    pdf_grand = sum(pdf_totals.values(), Decimal("0"))
    grand_diff = detail_grand - pdf_grand

    ok = (not unmapped_detail_props) and (abs(grand_diff) <= tolerance) and (len(mismatches) == 0)

    def _slug_simple2(text: str) -> str:
        t = (text or "").strip().lower()
        t = re.sub(r"[^a-z0-9]+", "_", t).strip("_")
        return t or "unknown"

    focus_report: dict | None = None
    focus_report_path: str | None = None
    if focus_property:
        needle = (focus_property or "").strip().casefold()

        # Prefer matching by yardi entity code first.
        matched_pdf_props: list[str] = [
            p for p, ent in pdf_property_to_yardi_entity.items() if (ent or "").strip().casefold() == needle
        ]
        if not matched_pdf_props:
            # Fallback: match by pdf property name containing the needle.
            matched_pdf_props = [p for p in sorted(pdf_properties) if needle in p.casefold()]

        matched_pdf_props = sorted(set(matched_pdf_props), key=lambda s: s.casefold())

        # Which detail bucket IDs map to those PDF props?
        detail_bucket_ids_for_focus = sorted(
            {dp for dp, pp in detail_to_pdf.items() if pp in set(matched_pdf_props)},
            key=lambda s: (len(s), s),
        )

        focus_keys = {k for k in all_keys if k[0] in set(matched_pdf_props)}
        focus_mismatches = [m for m in mismatches if m.get("pdf_property") in set(matched_pdf_props)]

        # Separate PDF rows with blank codes (these are the ones that need missing-code mapping)
        blank_code_rows: list[dict] = []
        for r in extracted_rows:
            if matched_pdf_props and r.property_name not in set(matched_pdf_props):
                continue
            if not r.edge_code:
                blank_code_rows.append(
                    {
                        "property_name": r.property_name,
                        "category": r.category,
                        "subcategory": r.subcategory,
                        "description": r.description,
                        "debit": f"{r.debit:.2f}",
                        "credit": f"{r.credit:.2f}",
                        "net": f"{(r.debit - r.credit):.2f}",
                        "source_page": r.source_page,
                    }
                )

        blank_code_totals_by_desc: dict[str, Decimal] = {}
        for r in extracted_rows:
            if matched_pdf_props and r.property_name not in set(matched_pdf_props):
                continue
            if r.edge_code:
                continue
            key = f"{r.category} | {r.subcategory} | {r.description}".strip()
            blank_code_totals_by_desc[key] = blank_code_totals_by_desc.get(key, Decimal("0")) + (
                r.debit - r.credit
            )

        focus_report = {
            "focus_property_input": focus_property,
            "matched_pdf_properties": matched_pdf_props,
            "detail_bucket_ids_mapped_to_focus": detail_bucket_ids_for_focus,
            "mismatch_count": len(focus_mismatches),
            "mismatches": focus_mismatches[:500],
            "pdf_blank_code_rows_count": len(blank_code_rows),
            "pdf_blank_code_rows": blank_code_rows[:500],
            "pdf_blank_code_totals_by_description": {
                k: f"{v}" for k, v in sorted(blank_code_totals_by_desc.items(), key=lambda t: (t[0].casefold()))
            },
            "notes": [
                "detail_bucket_ids_mapped_to_focus are the L###/T### bucket codes from the detail export.",
                "pdf_blank_code_rows are PDF lines where the Code column was blank (needs missing-code mapping).",
            ],
        }

        focus_path = artifacts_dir / f"csv_vs_pdf_focus_{report_month}__{_slug_simple2(focus_property)}.json"
        focus_path.write_text(json.dumps(focus_report, indent=2), encoding="utf-8")
        focus_report_path = str(focus_path)

    # Build a full side-by-side table (one row per property+account).
    # Matched rows first, then unmatched (mismatch / detail-only / pdf-only).
    artifacts_dir.mkdir(parents=True, exist_ok=True)

    pdf_prop_to_detail_ids: dict[str, list[str]] = {}
    for dp, pp in detail_to_pdf.items():
        pdf_prop_to_detail_ids.setdefault(pp, []).append(dp)
    for pp in list(pdf_prop_to_detail_ids.keys()):
        pdf_prop_to_detail_ids[pp] = sorted(set(pdf_prop_to_detail_ids[pp]), key=lambda s: (len(s), s))

    def _status_for_key(pp: str, acct: str, a: Decimal, b: Decimal) -> str:
        in_detail = (pp, acct) in detail_mapped_totals
        in_pdf = (pp, acct) in pdf_totals
        if in_detail and in_pdf:
            return "MATCH" if abs(a - b) <= tolerance else "MISMATCH"
        if in_detail and not in_pdf:
            return "DETAIL_ONLY"
        if in_pdf and not in_detail:
            return "PDF_ONLY"
        return "UNKNOWN"

    status_rank = {
        "MATCH": 0,
        "MISMATCH": 1,
        "DETAIL_ONLY": 2,
        "PDF_ONLY": 3,
        "UNKNOWN": 9,
    }

    table_rows: list[dict] = []
    for (pp, acct) in all_keys:
        a = detail_mapped_totals.get((pp, acct), Decimal("0"))
        b = pdf_totals.get((pp, acct), Decimal("0"))
        diff = a - b
        status = _status_for_key(pp, acct, a, b)
        table_rows.append(
            {
                "status": status,
                "status_rank": status_rank.get(status, 9),
                "pdf_property": pp,
                "yardi_entity": pdf_property_to_yardi_entity.get(pp, ""),
                "detail_property_ids": ";".join(pdf_prop_to_detail_ids.get(pp, [])),
                "account": acct,
                "detail_total": f"{a}",
                "pdf_total": f"{b}",
                "diff": f"{diff}",
                "abs_diff": abs(diff),
            }
        )

    # Sort: group by property first, then MATCH rows before unmatched.
    # Within a property+status group, largest diffs first, then account.
    table_rows.sort(
        key=lambda r: (
            str(r.get("pdf_property", "")).casefold(),
            int(r.get("status_rank", 9)),
            -float(r.get("abs_diff", Decimal("0"))),
            str(r.get("account", "")),
        )
    )

    table_csv_path = artifacts_dir / f"csv_vs_pdf_account_table_{report_month}.csv"
    with table_csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "status",
                "pdf_property",
                "yardi_entity",
                "detail_property_ids",
                "account",
                "detail_total",
                "pdf_total",
                "diff",
            ],
        )
        w.writeheader()
        for r in table_rows:
            w.writerow({k: r.get(k, "") for k in w.fieldnames})

    # Optional XLSX with conditional formatting (best visibility)
    table_xlsx_path: Path | None = None
    try:
        import openpyxl
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.styles import PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            raise RuntimeError("openpyxl workbook has no active worksheet")
        ws.title = "Comparison"

        headers = [
            "status",
            "pdf_property",
            "yardi_entity",
            "detail_property_ids",
            "account",
            "detail_total",
            "pdf_total",
            "diff",
        ]
        ws.append(headers)
        for r in table_rows:
            ws.append([r.get(h, "") for h in headers])

        # Freeze header and add a filter.
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:H{ws.max_row}"

        # Simple column sizing.
        for col_letter in ("A", "B", "C", "D", "E", "F", "G", "H"):
            ws.column_dimensions[col_letter].width = 18
        ws.column_dimensions["B"].width = 32
        ws.column_dimensions["D"].width = 18

        # Conditional formatting by status.
        fill_match = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fill_mismatch = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        fill_warn = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        # Apply to full row range A:H
        data_range = f"A2:H{ws.max_row}"
        ws.conditional_formatting.add(
            data_range,
            FormulaRule(formula=["$A2=\"MATCH\""], fill=fill_match),
        )
        ws.conditional_formatting.add(
            data_range,
            FormulaRule(formula=["$A2=\"MISMATCH\""], fill=fill_mismatch),
        )
        ws.conditional_formatting.add(
            data_range,
            FormulaRule(formula=["OR($A2=\"DETAIL_ONLY\",$A2=\"PDF_ONLY\")"], fill=fill_warn),
        )

        table_xlsx_path = artifacts_dir / f"csv_vs_pdf_account_table_{report_month}.xlsx"
        wb.save(table_xlsx_path)
    except Exception:
        table_xlsx_path = None

    result = {
        "report_month": report_month,
        "detail_csv": str(detail_csv_path),
        "tolerance": f"{tolerance}",
        "property_bucket_mapping": diagnostics,
        "unmapped_detail_property_ids": sorted(unmapped_detail_props),
        "invalid_account_codes_in_detail_before": invalid_account_codes_before,
        "invalid_account_codes_in_detail_after": invalid_account_codes_after,
        "detail_grand_total_raw": f"{sum(detail_totals_raw.values(), Decimal('0'))}",
        "detail_grand_total_corrected": f"{detail_grand}",
        "pdf_grand_total": f"{pdf_grand}",
        "grand_total_diff": f"{grand_diff}",
        "skipped_detail_rows_not_in_report_month": skipped_non_month_rows,
        "expected_detail_period_start": expected_period_start_date.isoformat(),
        "filled_account_rows": filled_account_rows,
        "corrected_account_rows": corrected_account_rows,
        "mismatch_count": len(mismatches),
        "mismatches": mismatches[:500],
        "status": "EQUAL" if ok else "NOT_EQUAL",
        "tables": {
            "account_side_by_side_csv": str(table_csv_path),
            "account_side_by_side_xlsx": str(table_xlsx_path) if table_xlsx_path else None,
            "notes": [
                "Rows are sorted with MATCH first, then unmatched (MISMATCH/DETAIL_ONLY/PDF_ONLY).",
                "XLSX uses conditional formatting to highlight status when available.",
            ],
        },
        "focus_report": {
            "requested": focus_property,
            "path": focus_report_path,
        },
    }

    out_path = artifacts_dir / f"csv_vs_pdf_comparison_{report_month}.json"
    out_path.write_text(json.dumps(result, indent=2), encoding="utf-8")
    return str(out_path), result


def _read_property_assignees(path: Path) -> dict[str, str]:
    """property_code -> assignee (case-preserving value)."""
    if not path.exists():
        return {}
    out: dict[str, str] = {}
    with path.open("r", newline="", encoding="utf-8-sig") as f:
        r = csv.DictReader(f)
        for row in r:
            code = str(row.get("property_code", "") or "").strip()
            assignee = str(row.get("assignee", "") or "").strip()
            if not code or code.startswith("#"):
                continue
            if not assignee or assignee.startswith("#"):
                continue
            out[code] = assignee
    return out


def _norm_key(text: str) -> str:
    t = (text or "").strip().casefold()
    t = re.sub(r"[^a-z0-9]+", "", t)
    return t


def _load_property_code_mapping(property_code_mapping_csv: Path) -> tuple[set[str], dict[str, str]]:
    """Returns (known_codes_lower, norm_name_to_code_lower)."""
    if not property_code_mapping_csv.exists():
        return set(), {}

    known_codes: set[str] = set()
    norm_name_to_code: dict[str, str] = {}

    with property_code_mapping_csv.open("r", newline="", encoding="utf-8-sig") as f:
        r = csv.DictReader(f)
        for row in r:
            code = str(row.get("property_code", "") or "").strip()
            if not code or code.startswith("#"):
                continue
            code_l = code.casefold()
            known_codes.add(code_l)

            for k in ("property_name", "pdf_property"):
                name = str(row.get(k, "") or "").strip()
                if not name or name.startswith("#"):
                    continue
                nk = _norm_key(name)
                if nk and nk not in norm_name_to_code:
                    norm_name_to_code[nk] = code_l

    return known_codes, norm_name_to_code


def _read_property_assignees_from_assigned_property_xlsx(
    *,
    xlsx_path: Path,
    property_code_mapping_csv: Path,
) -> dict[str, str]:
    """Reads an 'Assigned Property.xlsx' style sheet and returns property_code -> assignee."""
    if not xlsx_path.exists() or not property_code_mapping_csv.exists():
        return {}

    try:
        import openpyxl  # type: ignore
    except Exception:
        return {}

    known_codes, norm_name_to_code = _load_property_code_mapping(property_code_mapping_csv)
    if not known_codes and not norm_name_to_code:
        return {}

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    except Exception:
        return {}

    try:
        sheet_name = "Upload summary" if "Upload summary" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]

        header_row_idx: int | None = None
        prop_col_idx: int | None = None
        assignee_col_idx: int | None = None

        # Find header row within the first 200 rows.
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=200, values_only=True), start=1):
            vals = [str(v).strip() if isinstance(v, str) else ("" if v is None else str(v)) for v in row[:20]]
            if not any(v for v in vals):
                continue

            # Look for a 'Property' header.
            prop_hits = [i for i, v in enumerate(vals) if v.casefold() == "property" or "property" in v.casefold()]
            if not prop_hits:
                continue
            prop_col_idx = prop_hits[0]

            # Assignee column is often immediately after Property (even if header cell is blank).
            # Prefer an explicit header if present.
            assignee_hits = [
                i
                for i, v in enumerate(vals)
                if any(k in v.casefold() for k in ("assignee", "assigned", "owner"))
            ]
            if assignee_hits:
                assignee_col_idx = assignee_hits[0]
            else:
                assignee_col_idx = prop_col_idx + 1 if (prop_col_idx + 1) < len(vals) else None

            header_row_idx = idx
            break

        if header_row_idx is None or prop_col_idx is None or assignee_col_idx is None:
            return {}

        out: dict[str, str] = {}
        empty_streak = 0
        for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=5000, values_only=True):
            # Extract property + assignee.
            raw_prop = row[prop_col_idx] if prop_col_idx < len(row) else None
            raw_assignee = row[assignee_col_idx] if assignee_col_idx < len(row) else None

            prop = str(raw_prop).strip() if raw_prop is not None else ""
            assignee = str(raw_assignee).strip() if raw_assignee is not None else ""

            if not prop and not assignee:
                empty_streak += 1
                if empty_streak >= 25:
                    break
                continue
            empty_streak = 0

            if not prop or prop.startswith("#"):
                continue
            if not assignee or assignee.startswith("#"):
                continue

            # Resolve property to canonical code.
            prop_clean = prop.strip()
            code_l: str | None = None
            if prop_clean.casefold() in known_codes:
                code_l = prop_clean.casefold()
            else:
                nk = _norm_key(prop_clean)
                code_l = norm_name_to_code.get(nk)

            if not code_l:
                continue

            out[code_l] = assignee

        return out
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _slug_simple(text: str) -> str:
    t = (text or "").strip()
    t = re.sub(r"[^A-Za-z0-9]+", "_", t).strip("_")
    return t or "unknown"


def _write_template_import_splits_by_assignee(
    *,
    month_dir: Path,
    report_month: str,
    template_csv_path: Path,
    property_assignee_csv: Path,
    assigned_property_xlsx: Path | None = None,
    property_code_mapping_csv: Path | None = None,
    include_assignees: set[str] | None = None,
) -> dict[str, str]:
    assignee_by_code = _read_property_assignees(property_assignee_csv) if property_assignee_csv else {}
    if assigned_property_xlsx and property_code_mapping_csv:
        from_xlsx = _read_property_assignees_from_assigned_property_xlsx(
            xlsx_path=assigned_property_xlsx,
            property_code_mapping_csv=property_code_mapping_csv,
        )
        # CSV wins if both specify the same code.
        for code_l, assignee in from_xlsx.items():
            assignee_by_code.setdefault(code_l, assignee)

    if not assignee_by_code:
        return {}

    include_cf = {a.casefold() for a in include_assignees} if include_assignees else None

    rows_by_assignee: dict[str, list[list[str]]] = {}
    with template_csv_path.open("r", newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for row in reader:
            if not row or len(row) < 15:
                continue
            code = str(row[8] or "").strip()  # column I (property code)
            assignee = assignee_by_code.get(code.casefold())
            if not assignee:
                continue
            if include_cf is not None and assignee.casefold() not in include_cf:
                continue
            rows_by_assignee.setdefault(assignee, []).append(list(row[:15]))

    if not rows_by_assignee:
        return {}

    out_dir = month_dir / "Output"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_paths: dict[str, str] = {}
    for assignee, rows in sorted(rows_by_assignee.items(), key=lambda t: t[0].lower()):
        out_path = out_dir / f"yardi_import_template_{report_month}__{_slug_simple(assignee)}.csv"
        try:
            f = out_path.open("w", newline="", encoding="utf-8")
            actual_path = out_path
        except PermissionError:
            actual_path = None
            for i in range(1, 1000):
                alt = out_path.with_name(f"{out_path.stem} ({i}){out_path.suffix}")
                if not alt.exists():
                    actual_path = alt
                    break
            if actual_path is None:
                raise
            f = actual_path.open("w", newline="", encoding="utf-8")

        with f:
            w = csv.writer(f)
            w.writerows(rows)
        out_paths[assignee] = str(actual_path)

    return out_paths


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--month-folder", required=True, help="Month folder under 09. Edge to Yardi (e.g. '12. Dec')")
    parser.add_argument("--report-month", required=True, help="Reporting month in YYYY-MM format")
    parser.add_argument("--pdf", default=None, help="Optional explicit PDF path override")
    parser.add_argument(
        "--confirm",
        action="store_true",
        help="Apply file renames (default is dry-run; no file changes)",
    )
    parser.add_argument(
        "--skip-rename",
        action="store_true",
        help="Skip renaming/moving input PDFs",
    )
    parser.add_argument(
        "--skip-normalize",
        action="store_true",
        help="Skip parsing PDF into edge_normalized.csv",
    )
    parser.add_argument(
        "--skip-draft-import",
        action="store_true",
        help="Skip generating the draft Yardi import CSV",
    )
    parser.add_argument(
        "--skip-mapping",
        action="store_true",
        help="Skip mapping step and do not generate final import",
    )
    parser.add_argument(
        "--compare-csv-to-pdf",
        action="store_true",
        help="Compare the month detail template CSV totals to the parsed PDF totals and write a JSON report",
    )
    parser.add_argument(
        "--compare-tolerance",
        default="0.01",
        help="Tolerance for CSV vs PDF total differences (default 0.01)",
    )
    parser.add_argument(
        "--compare-focus-property",
        default=None,
        help=(
            "Optional: focus the CSV↔PDF comparison on a single property. "
            "Matches by Yardi entity code first (e.g. 'smslawnd'), then by PDF property name substring (e.g. 'Lawndale')."
        ),
    )

    args = parser.parse_args()

    project_dir = Path(__file__).resolve().parent.parent

    # Load workspace-level .env so Google credentials (and other local settings)
    # are available when running from CLI.
    try:
        from dotenv import load_dotenv

        env_path = project_dir.parent / ".env"
        if env_path.exists():
            load_dotenv(env_path)
    except Exception:
        pass

    report_month = _parse_report_month(args.report_month)

    month_dir = _month_folder_dir(project_dir, args.month_folder)
    artifacts_dir = _month_helper_artifacts_dir(project_dir, month_dir=month_dir)

    print("=== RUN START ===")
    print(f"month_folder: {args.month_folder}")
    print(f"report_month: {report_month}")
    print(f"confirm: {bool(args.confirm)}")

    if args.compare_csv_to_pdf:
        detail_csv_path = _find_detail_template_csv(month_dir)
        if detail_csv_path is None:
            raise SystemExit("Could not find detail template CSV in month Input/")

        # Parse + map rows (same as the normal workflow path)
        if args.pdf:
            pdf_path = Path(args.pdf).expanduser().resolve()
            if not pdf_path.exists():
                raise FileNotFoundError(f"PDF not found: {pdf_path}")
        else:
            candidates = _find_candidate_pdfs(month_dir)
            pdf_path = _pick_pdf(candidates)

        doc = fitz.open(pdf_path)
        sections, warnings = _extract_property_sections(doc)
        if not sections:
            raise SystemExit("Could not detect property sections in the PDF; cannot compare")

        extracted_rows, parse_warnings = extract_rows_from_pdf(
            Path(pdf_path),
            property_sections=[asdict(s) for s in sections],
        )

        mapping_paths = ensure_mapping_templates(
            project_dir,
            properties=sorted({r.property_name for r in extracted_rows if r.property_name}),
            edge_codes=sorted({r.edge_code for r in extracted_rows if r.edge_code}),
        )

        draft_rows_for_mapping: list[dict] = []
        for r in extracted_rows:
            draft_rows_for_mapping.append(
                {
                    "property_name": r.property_name,
                    "category": r.category,
                    "subcategory": r.subcategory,
                    "description": r.description,
                    "edge_code": r.edge_code,
                    "debit": f"{r.debit:.2f}",
                    "credit": f"{r.credit:.2f}",
                    "source_property": r.property_name,
                    "source_page": r.source_page,
                }
            )

        mapped = apply_mappings(
            rows=draft_rows_for_mapping,
            property_mapping_csv=Path(mapping_paths["property_mapping_csv"]),
            account_mapping_csv=Path(mapping_paths["account_mapping_csv"]),
            missing_code_account_mapping_csv=Path(mapping_paths["missing_code_account_mapping_csv"]),
        )

        tolerance = Decimal(str(args.compare_tolerance))
        report_path, stats = _compare_detail_csv_to_pdf(
            month_dir=month_dir,
            artifacts_dir=artifacts_dir,
            report_month=report_month,
            detail_csv_path=detail_csv_path,
            mapped_rows=mapped.mapped_rows,
            extracted_rows=extracted_rows,
            tolerance=tolerance,
            focus_property=(str(args.compare_focus_property).strip() if args.compare_focus_property else None),
        )

        print("\n=== RUN END ===")
        print(f"status: {stats.get('status')}")
        print(f"csv_vs_pdf_report: {report_path}")
        if warnings or parse_warnings:
            print("warnings:")
            for w in warnings + parse_warnings:
                print(f"- {w}")
        print("FINAL_MARKER: COMPLETED")
        return

    if args.pdf:
        pdf_path = Path(args.pdf).expanduser().resolve()
        if not pdf_path.exists():
            raise FileNotFoundError(f"PDF not found: {pdf_path}")
    else:
        candidates = _find_candidate_pdfs(month_dir)
        pdf_path = _pick_pdf(candidates)

    rename_result: dict | None = None
    if not args.skip_rename:
        pdf_path, rename_result = _rename_pdf_if_needed(
            pdf_path=pdf_path,
            month_dir=month_dir,
            report_month=report_month,
            confirm=bool(args.confirm),
        )

    doc = fitz.open(pdf_path)
    header_check = _check_expected_headers(doc)
    sections, warnings = _extract_property_sections(doc)

    normalized_csv_path: str | None = None
    yardi_import_path: str | None = None
    yardi_import_final_path: str | None = None
    yardi_import_template_path: str | None = None
    yardi_import_template_stats: dict | None = None
    yardi_import_template_by_assignee: dict | None = None
    totals: dict | None = None
    mapping_report_path: str | None = None
    mapping_status: dict | None = None
    parsed_tables: dict | None = None
    parse_warnings: list[str] = []
    extracted_rows: list[EdgeGlRow] = []

    if not args.skip_normalize and sections:
        extracted_rows, parse_warnings = extract_rows_from_pdf(
            Path(pdf_path),
            property_sections=[asdict(s) for s in sections],
        )
        normalized_csv_path = _write_normalized_csv(artifacts_dir, report_month=report_month, rows=extracted_rows)

        # Always write per-property table CSVs for human validation (matches screenshot headers).
        parsed_tables = _write_per_property_tables(
            artifacts_dir,
            report_month=report_month,
            rows=extracted_rows,
            yardi_account_by_edge_code=None,
        )

    if not args.skip_draft_import and extracted_rows:
        yardi_import_path, totals = _write_draft_yardi_import(
            month_dir,
            artifacts_dir=artifacts_dir,
            report_month=report_month,
            rows=extracted_rows,
        )

    # Mapping phase (Phase 3)
    if not args.skip_mapping and extracted_rows:
        mapping_paths = ensure_mapping_templates(
            project_dir,
            properties=sorted({r.property_name for r in extracted_rows}),
            edge_codes=sorted({r.edge_code for r in extracted_rows if r.edge_code}),
        )

        # Build dict rows aligned with CSV outputs.
        draft_rows_for_mapping: list[dict] = []
        for r in extracted_rows:
            draft_rows_for_mapping.append(
                {
                    "property_name": r.property_name,
                    "category": r.category,
                    "subcategory": r.subcategory,
                    "description": r.description,
                    "edge_code": r.edge_code,
                    "debit": f"{r.debit:.2f}",
                    "credit": f"{r.credit:.2f}",
                    "source_property": r.property_name,
                    "source_page": r.source_page,
                }
            )

        mapping_result = apply_mappings(
            rows=draft_rows_for_mapping,
            property_mapping_csv=Path(mapping_paths["property_mapping_csv"]),
            account_mapping_csv=Path(mapping_paths["account_mapping_csv"]),
            missing_code_account_mapping_csv=Path(mapping_paths["missing_code_account_mapping_csv"]),
        )

        mapping_report_path = write_mapping_report_xlsx(
            output_path=(artifacts_dir / f"mapping_report_{report_month}.xlsx"),
            unmapped_properties=mapping_result.unmapped_properties,
            unmapped_accounts=mapping_result.unmapped_accounts,
            mapping_paths=mapping_paths,
        )

        mapping_status = {
            "unmapped_properties_count": len(mapping_result.unmapped_properties),
            "unmapped_accounts_count": len(mapping_result.unmapped_accounts),
            "property_mapping_csv": mapping_paths["property_mapping_csv"],
            "account_mapping_csv": mapping_paths["account_mapping_csv"],
            "missing_code_account_mapping_csv": mapping_paths.get("missing_code_account_mapping_csv"),
            "property_code_mapping_csv": mapping_paths.get("property_code_mapping_csv"),
            "mapping_report_xlsx": mapping_report_path,
        }

        fully_mapped = (
            mapping_status["unmapped_properties_count"] == 0 and mapping_status["unmapped_accounts_count"] == 0
        )

        if fully_mapped:
            yardi_import_final_path = _write_final_yardi_import(
                month_dir,
                artifacts_dir=artifacts_dir,
                report_month=report_month,
                mapped_rows=mapping_result.mapped_rows,
            )

            # Generate the 15-column template CSV directly from the PDF-derived mapped rows.
            yardi_import_template_path, yardi_import_template_stats = _write_yardi_import_template_from_pdf(
                month_dir=month_dir,
                artifacts_dir=artifacts_dir,
                report_month=report_month,
                mapped_rows=mapping_result.mapped_rows,
            )

            assignee_csv = Path(mapping_paths.get("property_assignee_csv") or "")
            if assignee_csv and assignee_csv.exists() and yardi_import_template_path:
                yardi_import_template_by_assignee = _write_template_import_splits_by_assignee(
                    month_dir=month_dir,
                    report_month=report_month,
                    template_csv_path=Path(yardi_import_template_path),
                    property_assignee_csv=assignee_csv,
                    assigned_property_xlsx=(month_dir / "Input" / "Assigned Property.xlsx"),
                    property_code_mapping_csv=Path(mapping_paths.get("property_code_mapping_csv") or ""),
                    include_assignees={"Jay"},
                )

    # Preview: first 2000 chars of first page + property section starts
    first_page_text = str(doc.load_page(0).get_text("text") or "") if doc.page_count else ""
    preview_lines = [
        f"PDF: {pdf_path}",
        f"Pages: {doc.page_count}",
        "",
        "Property section starts:",
    ]
    for s in sections:
        preview_lines.append(f"- {s.property_name}: pages {s.start_page}-{s.end_page} ({s.header_line})")
    preview_lines.append("")
    preview_lines.append("First page preview (first 2000 chars):")
    preview_lines.append(first_page_text[:2000])
    preview_text = "\n".join(preview_lines)

    summary = {
        "status": "COMPLETED" if not warnings else "COMPLETED_WITH_WARNINGS",
        "month_folder": args.month_folder,
        "report_month": report_month,
        "confirm": bool(args.confirm),
        "inputs": {
            "pdf": str(pdf_path),
            "rename": rename_result,
        },
        "pdf": {
            "page_count": doc.page_count,
            "header_check": header_check,
            "property_sections": [asdict(s) for s in sections],
        },
        "outputs": {
            "edge_normalized_csv": normalized_csv_path,
            "yardi_import_draft_csv": yardi_import_path,
            "yardi_import_final_csv": yardi_import_final_path,
            "yardi_import_template_csv": yardi_import_template_path,
            "yardi_import_template_stats": yardi_import_template_stats,
            "yardi_import_template_by_assignee": yardi_import_template_by_assignee,
            "yardi_control_totals": totals,
            "mapping_report_xlsx": mapping_report_path,
            "mapping_status": mapping_status,
            "parsed_tables": parsed_tables,
        },
        "warnings": warnings + parse_warnings,
    }

    artifacts = _write_artifacts(artifacts_dir=artifacts_dir, summary=summary, preview_text=preview_text)
    print("\n=== RUN END ===")
    print(f"status: {summary['status']}")
    print(f"artifacts: {artifacts}")
    if warnings:
        print("warnings:")
        for w in warnings:
            print(f"- {w}")
    print("FINAL_MARKER: COMPLETED")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:  # noqa: BLE001
        print("\n=== RUN END ===")
        print("status: FAILED")
        print(f"error: {exc}")
        print("FINAL_MARKER: FAILED")
        raise
