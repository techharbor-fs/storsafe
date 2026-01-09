import argparse
import shutil
from pathlib import Path

from openpyxl import load_workbook


BALANCE_SHEET_RECON_DIR = Path(__file__).resolve().parent


def replace_yardi_report(workbook_path: Path, yardi_export_path: Path, sheet_name: str = "Yardi Report") -> None:
    wb = load_workbook(workbook_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {workbook_path.name}")

    ws = wb[sheet_name]

    # Clear existing content on Yardi Report (but keep sheet itself)
    # We clear values within the used range, being careful with merged cells.
    min_row = ws.min_row
    max_row = ws.max_row
    min_col = ws.min_column
    max_col = ws.max_column

    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            # Skip merged cells' non-top-left positions to avoid openpyxl MergedCell issues
            if type(cell).__name__ == "MergedCell":
                continue
            cell.value = None

    # Load the export (assume Excel format with a single relevant sheet)
    export_wb = load_workbook(yardi_export_path, data_only=False, read_only=True)
    export_ws = export_wb[export_wb.sheetnames[0]]

    max_row = export_ws.max_row
    max_col = export_ws.max_column

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            src_cell = export_ws.cell(row=r, column=c)
            dest_cell = ws.cell(row=r, column=c)
            # Again, avoid writing into merged cell placeholders
            if type(dest_cell).__name__ == "MergedCell":
                continue
            dest_cell.value = src_cell.value

    wb.save(workbook_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate next-month BS recon by copying a template and refreshing Yardi Report.")
    parser.add_argument("template", type=str, help="Path to prior-month/template BS recon workbook (xlsx).")
    parser.add_argument("output", type=str, help="Path for new-month BS recon workbook (xlsx).")
    parser.add_argument("yardi_export", type=str, help="Path to Yardi Balance Sheet (With Period Change) export (xlsx).")
    args = parser.parse_args()

    template_path = (BALANCE_SHEET_RECON_DIR / args.template).resolve() if not Path(args.template).is_absolute() else Path(args.template)
    output_path = (BALANCE_SHEET_RECON_DIR / args.output).resolve() if not Path(args.output).is_absolute() else Path(args.output)
    yardi_export_path = Path(args.yardi_export).resolve()

    if not template_path.is_file():
        raise FileNotFoundError(f"Template workbook not found: {template_path}")
    if not yardi_export_path.is_file():
        raise FileNotFoundError(f"Yardi export file not found: {yardi_export_path}")

    if output_path.exists():
        raise FileExistsError(f"Output file already exists: {output_path}")

    # Copy template to output
    shutil.copy2(template_path, output_path)

    # Replace Yardi Report content in the new workbook
    replace_yardi_report(output_path, yardi_export_path)

    print(f"Created next-month BS recon: {output_path}")


if __name__ == "__main__":
    main()
