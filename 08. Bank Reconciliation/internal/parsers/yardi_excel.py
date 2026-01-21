"""
Yardi Excel Parser

Extracts Outstanding Checks and Other Items from Bank Reconciliation Excel files.
These files are exported from Yardi and contain the "book side" of the reconciliation.

Also provides report analysis and automatic file renaming based on detected content.
"""

import re
import shutil
from datetime import date, datetime
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple

from openpyxl import load_workbook


# Month name mappings
MONTH_NAMES = {
    1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
    7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"
}

MONTH_NAME_TO_NUM = {
    "january": 1, "jan": 1,
    "february": 2, "feb": 2,
    "march": 3, "mar": 3,
    "april": 4, "apr": 4,
    "may": 5,
    "june": 6, "jun": 6,
    "july": 7, "jul": 7,
    "august": 8, "aug": 8,
    "september": 9, "sep": 9, "sept": 9,
    "october": 10, "oct": 10,
    "november": 11, "nov": 11,
    "december": 12, "dec": 12,
}


def analyze_yardi_report(excel_path: Path) -> Dict[str, Any]:
    """
    Analyze a Yardi Bank Rec Excel report to extract metadata.
    
    Args:
        excel_path: Path to the Excel file
        
    Returns:
        Dict with keys:
        - property: str (detected property name)
        - year: int (detected year)
        - month: int (detected month, 1-12)
        - suggested_name: str (suggested standardized filename)
        - confidence: str ('high', 'medium', 'low')
    """
    excel_path = Path(excel_path)
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    assert ws is not None
    
    # Extract all header text for analysis
    header_blob = _get_header_blob(ws)
    
    # Extract property name
    property_name = _extract_property_name(ws, excel_path)
    
    # Extract period (month/year)
    year, month = _extract_period(ws, header_blob, excel_path)
    
    wb.close()
    
    # Determine confidence level
    confidence = "high"
    if not property_name:
        confidence = "low"
    elif not year or not month:
        confidence = "medium"
    
    # Generate suggested filename
    suggested_name = None
    if property_name and year and month:
        month_str = f"{month:02d}"
        suggested_name = f"{year}-{month_str}_Bank_Rec_{property_name}.xlsx"
    
    return {
        "property": property_name,
        "year": year,
        "month": month,
        "suggested_name": suggested_name,
        "confidence": confidence,
        "original_name": excel_path.name,
    }


def rename_yardi_report(excel_path: Path, dry_run: bool = False) -> Tuple[bool, str, Optional[Path]]:
    """
    Analyze and rename a Yardi Bank Rec Excel report to standardized format.
    
    Format: {Year}-{Month}_Bank_Rec_{Property}.xlsx
    Example: 2025-12_Bank_Rec_Madison.xlsx
    
    Args:
        excel_path: Path to the Excel file
        dry_run: If True, only report what would happen without renaming
        
    Returns:
        Tuple of (success: bool, message: str, new_path: Optional[Path])
    """
    excel_path = Path(excel_path)
    
    if not excel_path.exists():
        return False, f"File not found: {excel_path}", None
    
    if not excel_path.suffix.lower() in (".xlsx", ".xls"):
        return False, f"Not an Excel file: {excel_path}", None
    
    # Analyze the report
    analysis = analyze_yardi_report(excel_path)
    
    if not analysis["suggested_name"]:
        missing = []
        if not analysis["property"]:
            missing.append("property name")
        if not analysis["year"]:
            missing.append("year")
        if not analysis["month"]:
            missing.append("month")
        return False, f"Could not determine: {', '.join(missing)}", None
    
    # Check if already has correct name
    if excel_path.name == analysis["suggested_name"]:
        return True, f"Already named correctly: {excel_path.name}", excel_path
    
    # Build new path (same directory)
    new_path = excel_path.parent / analysis["suggested_name"]
    
    # Check for conflicts
    if new_path.exists() and new_path != excel_path:
        return False, f"Target file already exists: {new_path}", None
    
    if dry_run:
        return True, f"Would rename: {excel_path.name} → {analysis['suggested_name']}", new_path
    
    # Perform the rename
    try:
        excel_path.rename(new_path)
        return True, f"Renamed: {excel_path.name} → {analysis['suggested_name']}", new_path
    except Exception as e:
        return False, f"Rename failed: {e}", None


def extract_yardi_from_excel(
    excel_path: Path, 
    auto_rename: bool = False
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """
    Extract Outstanding Checks and Other Items from Bank Rec Excel file.
    
    Args:
        excel_path: Path to the Bank_Rec*.xlsx file
        auto_rename: If True, rename file to standardized format after extraction
        
    Returns:
        Tuple of:
        - List of transaction dicts with keys:
            - property: str
            - date: str (YYYY-MM-DD)
            - transaction_id: str (check number, blank for Other Items)
            - description: str
            - amount: float
            - source_type: str ('check' or 'other')
        - Analysis dict with metadata (property, year, month, etc.)
    """
    excel_path = Path(excel_path)
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    assert ws is not None
    
    transactions: List[Dict[str, Any]] = []
    
    # Get header blob for analysis
    header_blob = _get_header_blob(ws)
    
    # Extract property name from the Excel report content
    property_name = _extract_property_name(ws, excel_path)
    
    # Extract period
    year, month = _extract_period(ws, header_blob, excel_path)
    
    current_section = None
    
    for row in ws.iter_rows():
        # Get cell values for section detection
        cells = [cell.value for cell in row]
        
        # Detect section headers
        if "Outstanding Checks" in str(cells):
            current_section = "checks"
            continue
        elif "Other Items" in str(cells):
            current_section = "other"
            continue
        elif "Less:" in str(cells) or "Plus/Minus:" in str(cells):
            # End of section marker
            continue
        
        if current_section == "checks":
            # Outstanding Checks format: 
            # Column C (index 2) = Date
            # Column F (index 5) = Check#
            # Column H (index 7) = Payee
            # Column I (index 8) = Amount
            date_val = cells[2] if len(cells) > 2 else None
            check_cell = row[5] if len(row) > 5 else None
            check_num = _excel_cell_to_text(check_cell)
            payee = cells[7] if len(cells) > 7 else None
            amount = cells[8] if len(cells) > 8 else None
            
            if date_val and amount and check_num:
                # Skip header row
                if str(check_num).strip().lower() == "check number":
                    continue
                    
                check_num_str = str(check_num).strip()
                trans_id = check_num_str
                description = str(payee).strip() if payee is not None else ""
                
                # Format date
                date_str = _format_date(date_val)
                
                # Amount should be negative for checks (outflow)
                try:
                    amt = -abs(float(amount))
                except (ValueError, TypeError):
                    continue
                    
                transactions.append({
                    "property": property_name,
                    "date": date_str,
                    "transaction_id": trans_id,
                    "description": description,
                    "amount": amt,
                    "source_type": "check",
                })
        
        elif current_section == "other":
            # Other Items format:
            # Column C (index 2) = Date
            # Column H (index 7) = Notes
            # Column I (index 8) = Amount
            date_val = cells[2] if len(cells) > 2 else None
            notes = cells[7] if len(cells) > 7 else None
            amount = cells[8] if len(cells) > 8 else None
            
            if date_val and amount and notes:
                # Skip header row
                if str(notes).strip().lower() == "notes":
                    continue
                
                # Format date
                date_str = _format_date(date_val)
                
                # Other Items have no transaction id
                trans_id = ""
                description = str(notes).strip()
                
                try:
                    amt = float(amount)  # Keep sign as-is for Other Items
                except (ValueError, TypeError):
                    continue
                    
                transactions.append({
                    "property": property_name,
                    "date": date_str,
                    "transaction_id": trans_id,
                    "description": description,
                    "amount": amt,
                    "source_type": "other",
                })
    
    wb.close()
    
    # Build analysis metadata
    confidence = "high"
    if not property_name:
        confidence = "low"
    elif not year or not month:
        confidence = "medium"
    
    suggested_name = None
    if property_name and year and month:
        month_str = f"{month:02d}"
        suggested_name = f"{year}-{month_str}_Bank_Rec_{property_name}.xlsx"
    
    analysis = {
        "property": property_name,
        "year": year,
        "month": month,
        "suggested_name": suggested_name,
        "confidence": confidence,
        "original_name": excel_path.name,
        "original_path": str(excel_path),
    }
    
    # Auto-rename if requested
    if auto_rename and suggested_name and excel_path.name != suggested_name:
        success, message, new_path = rename_yardi_report(excel_path)
        analysis["rename_result"] = {
            "success": success,
            "message": message,
            "new_path": str(new_path) if new_path else None,
        }
    
    return transactions, analysis


def _excel_cell_to_text(cell) -> str:
    """Convert Excel cell to text, preserving formatting like leading zeros."""
    if cell is None:
        return ""

    value = getattr(cell, "value", None)
    if value is None:
        return ""

    # If it's already text in Excel, preserve it.
    if isinstance(value, str):
        return value.strip()

    # Datetimes/dates should not be used for IDs; fall back to string.
    if hasattr(value, "strftime"):
        return value.strftime("%m/%d/%Y")

    number_format = (getattr(cell, "number_format", None) or "").strip()

    # Common case: Excel stored numeric, openpyxl returns float.
    if isinstance(value, float) and value.is_integer():
        value = int(value)

    if isinstance(value, int):
        text = str(value)
        # Handle simple formats like '00000' (leading zeros).
        if number_format and set(number_format) == {"0"}:
            width = len(number_format)
            text = text.zfill(width)
        return text

    return str(value).strip()


def _format_date(date_val) -> str:
    """Format date value to YYYY-MM-DD string."""
    if isinstance(date_val, (datetime, date)):
        return date_val.strftime("%Y-%m-%d")
    else:
        # Try to parse string date
        try:
            parsed = datetime.strptime(str(date_val), "%m/%d/%Y")
            return parsed.strftime("%Y-%m-%d")
        except ValueError:
            return str(date_val)


def _extract_property_name(ws, excel_path: Path) -> str:
    """Extract property name from worksheet header or filename."""
    # Search header rows for property name pattern
    header_search_cells: List[str] = []
    for r in ws.iter_rows(min_row=1, max_row=15, values_only=True):
        for v in r:
            if isinstance(v, str) and v.strip():
                header_search_cells.append(v.strip())
    header_blob = " \n".join(header_search_cells)

    # Prefer patterns that end before "Notre" when present.
    m = re.search(r"\bSS\s+of\s+(.+?)\s+Notre\b", header_blob, flags=re.IGNORECASE)
    if not m:
        m = re.search(r"\bSS\s+of\s+(.+?)\s+Bank\b", header_blob, flags=re.IGNORECASE)
    if not m:
        m = re.search(r"\bSS\s+of\s+([A-Za-z][A-Za-z ]+)\b", header_blob, flags=re.IGNORECASE)

    if m:
        return _title_case_property(m.group(1))

    # Fallback: Extract property name from filename if present (e.g., Bank_Rec_Madison.xlsx)
    if "_" in excel_path.stem:
        parts = excel_path.stem.split("_")
        if len(parts) >= 3 and parts[2].strip():
            return _title_case_property(parts[2])
    
    return ""


def _title_case_property(name: str) -> str:
    """Convert property name to title case."""
    name = " ".join(name.split()).strip()
    if not name:
        return ""
    return name.title()


def _get_header_blob(ws) -> str:
    """Extract all text from header rows for analysis."""
    header_search_cells: List[str] = []
    for r in ws.iter_rows(min_row=1, max_row=20, values_only=True):
        for v in r:
            if isinstance(v, str) and v.strip():
                header_search_cells.append(v.strip())
            elif isinstance(v, (datetime, date)):
                header_search_cells.append(v.strftime("%m/%d/%Y"))
    return " \n".join(header_search_cells)


def _extract_period(ws, header_blob: str, excel_path: Path) -> Tuple[Optional[int], Optional[int]]:
    """
    Extract year and month from the report.
    
    Looks for patterns like:
    - "Bank Reconciliation as of 12/31/2025"
    - "as of December 31, 2025"
    - "Period Ending 12/31/2025"
    - Date cells in header area
    - Folder name like "12. Dec"
    - Filename like "2025-12_Bank_Rec_Madison.xlsx"
    
    Returns:
        Tuple of (year, month) - either can be None if not detected
    """
    year = None
    month = None
    
    # Pattern 1: "as of MM/DD/YYYY" or "as of M/D/YYYY"
    m = re.search(r"as\s+of\s+(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})", header_blob, re.IGNORECASE)
    if m:
        month = int(m.group(1))
        year = int(m.group(3))
        return year, month
    
    # Pattern 2: "as of Month DD, YYYY"
    m = re.search(r"as\s+of\s+([A-Za-z]+)\s+\d{1,2},?\s+(\d{4})", header_blob, re.IGNORECASE)
    if m:
        month_name = m.group(1).lower()
        if month_name in MONTH_NAME_TO_NUM:
            month = MONTH_NAME_TO_NUM[month_name]
            year = int(m.group(2))
            return year, month
    
    # Pattern 3: "Period Ending MM/DD/YYYY"
    m = re.search(r"(?:period|ending|through)\s+(?:ending\s+)?(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})", 
                  header_blob, re.IGNORECASE)
    if m:
        month = int(m.group(1))
        year = int(m.group(3))
        return year, month
    
    # Pattern 4: Look for any date in header that looks like end-of-month
    for r in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        for v in r:
            if isinstance(v, (datetime, date)):
                # Assume this date represents the period
                year = v.year
                month = v.month
                return year, month
    
    # Pattern 5: Check parent folder name (e.g., "12. Dec")
    parent_folder = excel_path.parent.name
    m = re.match(r"(\d{1,2})\.\s*([A-Za-z]+)", parent_folder)
    if m:
        folder_month = int(m.group(1))
        if 1 <= folder_month <= 12:
            month = folder_month
            # Try to get year from grandparent or current year
            grandparent = excel_path.parent.parent.name
            year_match = re.search(r"20\d{2}", grandparent)
            if year_match:
                year = int(year_match.group())
            else:
                # Use current year as fallback
                year = datetime.now().year
            return year, month
    
    # Pattern 6: Check filename (e.g., "2025-12_Bank_Rec_Madison.xlsx")
    m = re.match(r"(\d{4})-(\d{2})_", excel_path.stem)
    if m:
        year = int(m.group(1))
        month = int(m.group(2))
        return year, month
    
    # Pattern 7: Look for any MM/DD/YYYY or YYYY in the header
    m = re.search(r"(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})", header_blob)
    if m:
        month = int(m.group(1))
        year = int(m.group(3))
        return year, month
    
    return year, month
