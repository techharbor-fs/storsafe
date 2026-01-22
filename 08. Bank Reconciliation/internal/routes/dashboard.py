"""
Dashboard Routes

Main dashboard showing reconciliation periods and stats.
"""

from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify

from ..database import get_db

bp = Blueprint("dashboard", __name__, url_prefix="/dashboard")


@bp.route("/")
def index():
    """Dashboard home - list of reconciliation periods."""
    db = get_db()
    
    # Get all periods with property names and stats
    periods = db.execute("""
        SELECT 
            rp.id,
            p.name as property_name,
            rp.year,
            rp.month,
            rp.status,
            rp.created_at,
            rp.completed_at,
            (SELECT COUNT(*) FROM bank_transactions WHERE period_id = rp.id) as bank_count,
            (SELECT COUNT(*) FROM yardi_transactions WHERE period_id = rp.id) as yardi_count,
            (SELECT COUNT(*) FROM matches WHERE period_id = rp.id) as match_count
        FROM reconciliation_periods rp
        JOIN properties p ON rp.property_id = p.id
        ORDER BY rp.year DESC, rp.month DESC, p.name
    """).fetchall()
    
    # Get summary stats
    total_periods = len(periods)
    completed_periods = sum(1 for p in periods if p["status"] == "completed")
    in_progress_periods = total_periods - completed_periods
    
    return render_template(
        "dashboard.html",
        periods=periods,
        stats={
            "total": total_periods,
            "completed": completed_periods,
            "in_progress": in_progress_periods,
        }
    )


@bp.route("/period/<int:period_id>")
def period_detail(period_id: int):
    """Detail view for a specific reconciliation period."""
    db = get_db()
    
    # Get period info
    period = db.execute("""
        SELECT 
            rp.*,
            p.name as property_name
        FROM reconciliation_periods rp
        JOIN properties p ON rp.property_id = p.id
        WHERE rp.id = ?
    """, (period_id,)).fetchone()
    
    if not period:
        return "Period not found", 404
    
    # Get match summary by pass
    match_summary = db.execute("""
        SELECT 
            match_pass,
            COUNT(*) as count
        FROM matches
        WHERE period_id = ?
        GROUP BY match_pass
        ORDER BY match_pass
    """, (period_id,)).fetchall()
    
    # Get matched transactions with details
    matched_transactions = db.execute("""
        SELECT 
            m.id as match_id,
            m.match_pass,
            m.match_type,
            m.notes,
            -- Bank transaction (may be NULL for yardi-yardi matches)
            bt.id as bank_id,
            bt.date as bank_date,
            bt.transaction_id as bank_txid,
            bt.description as bank_desc,
            bt.amount as bank_amount,
            -- Yardi transaction (may be NULL for bank-bank matches)
            yt.id as yardi_id,
            yt.date as yardi_date,
            yt.transaction_id as yardi_txid,
            yt.description as yardi_desc,
            yt.amount as yardi_amount
        FROM matches m
        LEFT JOIN match_bank_transactions mbt ON m.id = mbt.match_id
        LEFT JOIN bank_transactions bt ON mbt.bank_transaction_id = bt.id
        LEFT JOIN match_yardi_transactions myt ON m.id = myt.match_id
        LEFT JOIN yardi_transactions yt ON myt.yardi_transaction_id = yt.id
        WHERE m.period_id = ? AND m.match_type != 'suggestion'
        ORDER BY m.match_pass, m.id
    """, (period_id,)).fetchall()
    
    # Group matched transactions by match_id (for multi-transaction matches like PASS 5)
    matches_grouped = {}
    for row in matched_transactions:
        match_id = row["match_id"]
        if match_id not in matches_grouped:
            matches_grouped[match_id] = {
                "match_id": match_id,
                "match_pass": row["match_pass"],
                "match_type": row["match_type"],
                "notes": row["notes"],
                "bank_txns": [],
                "yardi_txns": [],
            }
        
        # Add bank transaction if present and not already added
        if row["bank_id"]:
            bank_ids = [t["id"] for t in matches_grouped[match_id]["bank_txns"]]
            if row["bank_id"] not in bank_ids:
                matches_grouped[match_id]["bank_txns"].append({
                    "id": row["bank_id"],
                    "date": row["bank_date"],
                    "txid": row["bank_txid"],
                    "desc": row["bank_desc"],
                    "amount": row["bank_amount"],
                })
        
        # Add yardi transaction if present and not already added
        if row["yardi_id"]:
            yardi_ids = [t["id"] for t in matches_grouped[match_id]["yardi_txns"]]
            if row["yardi_id"] not in yardi_ids:
                matches_grouped[match_id]["yardi_txns"].append({
                    "id": row["yardi_id"],
                    "date": row["yardi_date"],
                    "txid": row["yardi_txid"],
                    "desc": row["yardi_desc"],
                    "amount": row["yardi_amount"],
                })
    
    # Calculate totals and check for associated adjustments
    for match_id, match in matches_grouped.items():
        # Calculate totals
        match["bank_total"] = sum(t["amount"] or 0 for t in match["bank_txns"])
        match["yardi_total"] = sum(t["amount"] or 0 for t in match["yardi_txns"])
        match["difference"] = round(match["bank_total"] - match["yardi_total"], 2)
        
        # Check if there's an associated differential adjustment
        adj = db.execute("""
            SELECT id, amount, category, confirmed 
            FROM adjustment_entries 
            WHERE source_match_id = ?
        """, (match_id,)).fetchone()
        
        if adj:
            match["adjustment"] = {
                "id": adj["id"],
                "amount": adj["amount"],
                "category": adj["category"],
                "confirmed": adj["confirmed"],
            }
        else:
            match["adjustment"] = None
    
    matches_list = list(matches_grouped.values())
    
    # Sort matches: MANUAL first, then by pass order (PASS 1, PASS 2, etc.)
    def get_match_sort_key(match):
        # Primary: match pass (MANUAL = 0 to appear first, then PASS 1, 2, 3...)
        pass_str = match.get("match_pass", "PASS 99") or "PASS 99"
        try:
            # Handle formats like "PASS 1", "PASS 2", "MANUAL", etc.
            if pass_str == "MANUAL":
                pass_num = 0  # Manual matches appear FIRST
            elif pass_str.startswith("PASS"):
                pass_num = int(pass_str.replace("PASS", "").strip())
            else:
                pass_num = 99
        except ValueError:
            pass_num = 99
        
        # Secondary: match_id (creation order within each pass)
        match_id = match.get("match_id", 0) or 0
        
        return (pass_num, match_id)
    
    matches_list.sort(key=get_match_sort_key)
    
    # Get unmatched bank transactions (exclude matched AND adjustment items)
    unmatched_bank_txns = db.execute("""
        SELECT bt.*
        FROM bank_transactions bt
        WHERE bt.period_id = ?
        AND NOT EXISTS (
            SELECT 1 FROM match_bank_transactions mbt
            WHERE mbt.bank_transaction_id = bt.id
        )
        AND NOT EXISTS (
            SELECT 1 FROM adjustment_entries ae
            WHERE ae.source_bank_txn_id = bt.id
        )
        ORDER BY bt.date ASC, bt.id ASC
    """, (period_id,)).fetchall()
    
    # Get unmatched yardi checks (Outstanding Checks - sorted ascending)
    unmatched_yardi_checks = db.execute("""
        SELECT yt.*
        FROM yardi_transactions yt
        WHERE yt.period_id = ?
        AND yt.source_type = 'check'
        AND NOT EXISTS (
            SELECT 1 FROM match_yardi_transactions myt
            WHERE myt.yardi_transaction_id = yt.id
        )
        ORDER BY yt.date ASC, yt.id ASC
    """, (period_id,)).fetchall()
    
    # Get unmatched yardi other items (Other Items - sorted ascending)
    unmatched_yardi_other = db.execute("""
        SELECT yt.*
        FROM yardi_transactions yt
        WHERE yt.period_id = ?
        AND (yt.source_type = 'other' OR yt.source_type IS NULL)
        AND NOT EXISTS (
            SELECT 1 FROM match_yardi_transactions myt
            WHERE myt.yardi_transaction_id = yt.id
        )
        ORDER BY yt.date ASC, yt.id ASC
    """, (period_id,)).fetchall()
    
    # Combine for backward compatibility (template may still use unmatched_yardi)
    unmatched_yardi_txns = list(unmatched_yardi_checks) + list(unmatched_yardi_other)
    
    # Get PASS 7 suggestions
    suggestions = db.execute("""
        SELECT 
            m.id as match_id,
            m.notes
        FROM matches m
        WHERE m.period_id = ? AND m.match_type = 'suggestion'
    """, (period_id,)).fetchall()
    
    # Get adjustments and adjustment suggestions
    from .adjustments import get_adjustments_for_period, get_suggested_adjustments, CATEGORY_LABELS
    adjustments = get_adjustments_for_period(period_id)
    adjustment_suggestions = get_suggested_adjustments(period_id)
    
    return render_template(
        "period_detail.html",
        period=period,
        match_summary=match_summary,
        matches=matches_list,
        unmatched_bank=unmatched_bank_txns,
        unmatched_yardi=unmatched_yardi_txns,
        unmatched_yardi_checks=unmatched_yardi_checks,
        adjustments=adjustments,
        adjustment_suggestions=adjustment_suggestions,
        category_labels=CATEGORY_LABELS,
        unmatched_yardi_other=unmatched_yardi_other,
        suggestions=suggestions,
    )


@bp.route("/unmatch/<int:match_id>", methods=["POST"])
def unmatch(match_id: int):
    """Remove a match, making transactions unmatched again."""
    db = get_db()
    
    # Get the match to find its period_id for redirect
    match = db.execute(
        "SELECT period_id FROM matches WHERE id = ?", (match_id,)
    ).fetchone()
    
    if not match:
        flash("Match not found.", "error")
        return redirect(url_for("dashboard.index"))
    
    period_id = match["period_id"]
    
    # Delete the match (cascades to link tables due to ON DELETE CASCADE)
    db.execute("DELETE FROM matches WHERE id = ?", (match_id,))
    db.commit()
    
    flash("Match removed. Transactions are now unmatched.", "success")
    return redirect(url_for("dashboard.period_detail", period_id=period_id))


@bp.route("/period/<int:period_id>/match", methods=["POST"])
def create_match(period_id: int):
    """Create a manual match from selected transactions."""
    db = get_db()
    
    # Get JSON data from request
    data = request.get_json()
    
    if not data:
        return jsonify({"success": False, "error": "No data provided"}), 400
    
    bank_ids = data.get("bank_ids", [])
    yardi_ids = data.get("yardi_ids", [])
    
    if not bank_ids or not yardi_ids:
        return jsonify({"success": False, "error": "Must select transactions from both sides"}), 400
    
    # Verify period exists
    period = db.execute(
        "SELECT id FROM reconciliation_periods WHERE id = ?", (period_id,)
    ).fetchone()
    
    if not period:
        return jsonify({"success": False, "error": "Period not found"}), 404
    
    # Verify all bank transactions exist and belong to this period
    for bank_id in bank_ids:
        txn = db.execute(
            "SELECT id FROM bank_transactions WHERE id = ? AND period_id = ?",
            (bank_id, period_id)
        ).fetchone()
        if not txn:
            return jsonify({"success": False, "error": f"Bank transaction {bank_id} not found"}), 404
        
        # Check if already matched
        existing = db.execute(
            "SELECT match_id FROM match_bank_transactions WHERE bank_transaction_id = ?",
            (bank_id,)
        ).fetchone()
        if existing:
            return jsonify({"success": False, "error": f"Bank transaction {bank_id} is already matched"}), 400
    
    # Verify all yardi transactions exist and belong to this period
    for yardi_id in yardi_ids:
        txn = db.execute(
            "SELECT id FROM yardi_transactions WHERE id = ? AND period_id = ?",
            (yardi_id, period_id)
        ).fetchone()
        if not txn:
            return jsonify({"success": False, "error": f"Yardi transaction {yardi_id} not found"}), 404
        
        # Check if already matched
        existing = db.execute(
            "SELECT match_id FROM match_yardi_transactions WHERE yardi_transaction_id = ?",
            (yardi_id,)
        ).fetchone()
        if existing:
            return jsonify({"success": False, "error": f"Yardi transaction {yardi_id} is already matched"}), 400
    
    # Calculate totals for differential check
    bank_total = 0
    for bank_id in bank_ids:
        txn = db.execute("SELECT amount FROM bank_transactions WHERE id = ?", (bank_id,)).fetchone()
        if txn:
            bank_total += txn["amount"]
    
    yardi_total = 0
    for yardi_id in yardi_ids:
        txn = db.execute("SELECT amount FROM yardi_transactions WHERE id = ?", (yardi_id,)).fetchone()
        if txn:
            yardi_total += txn["amount"]
    
    difference = round(bank_total - yardi_total, 2)
    
    # Create the match
    cursor = db.execute("""
        INSERT INTO matches (period_id, match_pass, match_type, notes)
        VALUES (?, 'MANUAL', 'bank_yardi', 'Created via manual matching')
    """, (period_id,))
    match_id = cursor.lastrowid
    
    # Link bank transactions
    for bank_id in bank_ids:
        db.execute(
            "INSERT INTO match_bank_transactions (match_id, bank_transaction_id) VALUES (?, ?)",
            (match_id, bank_id)
        )
    
    # Link yardi transactions
    for yardi_id in yardi_ids:
        db.execute(
            "INSERT INTO match_yardi_transactions (match_id, yardi_transaction_id) VALUES (?, ?)",
            (match_id, yardi_id)
        )
    
    db.commit()
    
    # AUTOMATICALLY create differential adjustment if amounts don't match
    adjustment_id = None
    if abs(difference) >= 0.01:
        cursor = db.execute("""
            INSERT INTO adjustment_entries 
            (period_id, source_type, source_match_id, date, description, amount, category, confirmed, notes)
            VALUES (?, 'differential', ?, DATE('now'), ?, ?, 'other', TRUE, ?)
        """, (
            period_id,
            match_id,
            f"Differential from manual match (Bank: ${bank_total:,.2f}, Yardi: ${yardi_total:,.2f})",
            difference,
            "Auto-created from unequal match",
        ))
        adjustment_id = cursor.lastrowid
        db.commit()
    
    return jsonify({
        "success": True,
        "match_id": match_id,
        "message": f"Matched {len(bank_ids)} bank + {len(yardi_ids)} yardi transactions",
        "bank_total": bank_total,
        "yardi_total": yardi_total,
        "difference": difference,
        "adjustment_id": adjustment_id,
    })


@bp.route("/period/<int:period_id>/rerun-matching", methods=["POST"])
def rerun_matching(period_id: int):
    """Clear existing matches and re-run auto-matching."""
    db = get_db()
    
    # Verify period exists
    period = db.execute(
        "SELECT id FROM reconciliation_periods WHERE id = ?", (period_id,)
    ).fetchone()
    
    if not period:
        flash("Period not found.", "error")
        return redirect(url_for("dashboard.index"))
    
    # Clear existing matches for this period
    db.execute("DELETE FROM matches WHERE period_id = ?", (period_id,))
    db.commit()
    
    # Re-run auto-matching
    try:
        from ..matching import run_auto_matching
        result = run_auto_matching(db, period_id)
        
        flash(
            f"Re-matched: {result.total_matched} matched, "
            f"{result.unmatched_bank} bank + {result.unmatched_yardi} yardi unmatched.",
            "success"
        )
    except Exception as e:
        flash(f"Error re-running matching: {str(e)}", "error")
    
    return redirect(url_for("dashboard.period_detail", period_id=period_id))


@bp.route("/period/<int:period_id>/toggle-status", methods=["POST"])
def toggle_status(period_id: int):
    """Toggle period status between in_progress and completed."""
    db = get_db()
    
    period = db.execute(
        "SELECT id, status FROM reconciliation_periods WHERE id = ?", (period_id,)
    ).fetchone()
    
    if not period:
        flash("Period not found.", "error")
        return redirect(url_for("dashboard.index"))
    
    new_status = "completed" if period["status"] == "in_progress" else "in_progress"
    completed_at = "CURRENT_TIMESTAMP" if new_status == "completed" else "NULL"
    
    if new_status == "completed":
        db.execute(
            "UPDATE reconciliation_periods SET status = ?, completed_at = CURRENT_TIMESTAMP WHERE id = ?",
            (new_status, period_id)
        )
    else:
        db.execute(
            "UPDATE reconciliation_periods SET status = ?, completed_at = NULL WHERE id = ?",
            (new_status, period_id)
        )
    db.commit()
    
    if new_status == "completed":
        flash("Reconciliation marked as completed.", "success")
    else:
        flash("Reconciliation reopened.", "info")
    
    return redirect(url_for("dashboard.period_detail", period_id=period_id))


@bp.route("/period/<int:period_id>/export")
def export_excel(period_id: int):
    """Export matched and unmatched transactions to Excel."""
    from flask import send_file
    from datetime import datetime
    import io
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash("openpyxl is required for Excel export.", "error")
        return redirect(url_for("dashboard.period_detail", period_id=period_id))
    
    db = get_db()
    
    # Get period info
    period = db.execute("""
        SELECT rp.*, p.name as property_name
        FROM reconciliation_periods rp
        JOIN properties p ON rp.property_id = p.id
        WHERE rp.id = ?
    """, (period_id,)).fetchone()
    
    if not period:
        flash("Period not found.", "error")
        return redirect(url_for("dashboard.index"))
    
    # Create workbook
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="052054", end_color="052054", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # =========== Sheet 1: Matched ===========
    ws_matched = wb.active
    ws_matched.title = "Matched"
    
    # Get matched transactions
    matched = db.execute("""
        SELECT 
            m.match_pass,
            bt.date as bank_date, bt.transaction_id as bank_txid, 
            bt.description as bank_desc, bt.amount as bank_amount,
            yt.date as yardi_date, yt.transaction_id as yardi_txid,
            yt.description as yardi_desc, yt.amount as yardi_amount
        FROM matches m
        LEFT JOIN match_bank_transactions mbt ON m.id = mbt.match_id
        LEFT JOIN bank_transactions bt ON mbt.bank_transaction_id = bt.id
        LEFT JOIN match_yardi_transactions myt ON m.id = myt.match_id
        LEFT JOIN yardi_transactions yt ON myt.yardi_transaction_id = yt.id
        WHERE m.period_id = ? AND m.match_type != 'suggestion'
        ORDER BY m.match_pass, m.id
    """, (period_id,)).fetchall()
    
    # Headers
    headers = ["Pass", "Bank Date", "Bank ID", "Bank Description", "Bank Amount",
               "Yardi Date", "Yardi ID", "Yardi Description", "Yardi Amount"]
    for col, header in enumerate(headers, 1):
        cell = ws_matched.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Data
    for row_num, row in enumerate(matched, 2):
        ws_matched.cell(row=row_num, column=1, value=row["match_pass"])
        ws_matched.cell(row=row_num, column=2, value=row["bank_date"])
        ws_matched.cell(row=row_num, column=3, value=row["bank_txid"])
        ws_matched.cell(row=row_num, column=4, value=row["bank_desc"])
        ws_matched.cell(row=row_num, column=5, value=row["bank_amount"])
        ws_matched.cell(row=row_num, column=6, value=row["yardi_date"])
        ws_matched.cell(row=row_num, column=7, value=row["yardi_txid"])
        ws_matched.cell(row=row_num, column=8, value=row["yardi_desc"])
        ws_matched.cell(row=row_num, column=9, value=row["yardi_amount"])
    
    # =========== Sheet 2: Unmatched Bank ===========
    ws_bank = wb.create_sheet("Unmatched Bank")
    
    unmatched_bank = db.execute("""
        SELECT bt.*
        FROM bank_transactions bt
        WHERE bt.period_id = ?
        AND NOT EXISTS (
            SELECT 1 FROM match_bank_transactions mbt WHERE mbt.bank_transaction_id = bt.id
        )
        ORDER BY bt.date
    """, (period_id,)).fetchall()
    
    headers = ["Date", "Transaction ID", "Description", "Amount", "Type"]
    for col, header in enumerate(headers, 1):
        cell = ws_bank.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    for row_num, row in enumerate(unmatched_bank, 2):
        ws_bank.cell(row=row_num, column=1, value=row["date"])
        ws_bank.cell(row=row_num, column=2, value=row["transaction_id"])
        ws_bank.cell(row=row_num, column=3, value=row["description"])
        ws_bank.cell(row=row_num, column=4, value=row["amount"])
        ws_bank.cell(row=row_num, column=5, value=row["transaction_type"])
    
    # =========== Sheet 3: Unmatched Yardi ===========
    ws_yardi = wb.create_sheet("Unmatched Yardi")
    
    unmatched_yardi = db.execute("""
        SELECT yt.*
        FROM yardi_transactions yt
        WHERE yt.period_id = ?
        AND NOT EXISTS (
            SELECT 1 FROM match_yardi_transactions myt WHERE myt.yardi_transaction_id = yt.id
        )
        ORDER BY yt.date
    """, (period_id,)).fetchall()
    
    headers = ["Date", "Transaction ID", "Description", "Amount", "Source Type"]
    for col, header in enumerate(headers, 1):
        cell = ws_yardi.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    for row_num, row in enumerate(unmatched_yardi, 2):
        ws_yardi.cell(row=row_num, column=1, value=row["date"])
        ws_yardi.cell(row=row_num, column=2, value=row["transaction_id"])
        ws_yardi.cell(row=row_num, column=3, value=row["description"])
        ws_yardi.cell(row=row_num, column=4, value=row["amount"])
        ws_yardi.cell(row=row_num, column=5, value=row["source_type"])
    
    # Adjust column widths
    for ws in [ws_matched, ws_bank, ws_yardi]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    month_names = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                   'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    filename = f"Bank_Rec_{period['property_name']}_{month_names[period['month']]}_{period['year']}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
