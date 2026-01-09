#!/usr/bin/env python3
"""
Bank Reconciliation Automation Script

Extracts transactions from bank statement PDFs and populates Google Sheets
for reconciliation between bank and Yardi data.

Usage:
    python bank_reconciliation.py
    
The script will prompt for folder selection containing bank statement PDFs.
"""

import logging
import json
import os
import sys
import tempfile
import tkinter as tk
import argparse
from datetime import date, datetime
from pathlib import Path
from tkinter import filedialog, messagebox
from typing import Dict, List, Optional, Tuple

import gspread

from bank_parsers import Transaction, get_parser_for_pdf
from bank_parsers.base_parser import BaseBankParser

# Configuration
DEBUG = False
DEFAULT_SHEET_ID = "1c-Jejkmm4LeirFEEJUalb49lKsPMQS8u8Z3OxwkToEA"
BANK_SHEET_NAME = "Bank"
YARDI_SHEET_NAME = "Yardi"
MATCHED_SHEET_NAME = "Matched"
UNMATCHED_SHEET_NAME = "Unmatched"

# Google Sheets API scopes
SCOPES = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive'
]

# Set up logging
logging.basicConfig(
    level=logging.DEBUG if DEBUG else logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def debug_print(msg: str):
    """Print debug message if DEBUG mode is enabled."""
    if DEBUG:
        print(f"[DEBUG] {msg}")


def connect_to_sheets(sheet_id: str) -> gspread.Spreadsheet:
    """Connect to Google Sheets using gspread service account."""
    # Prefer environment variable (best practice)
    temp_json_path: Optional[Path] = None
    service_account_json = os.environ.get("SERVICE_ACCOUNT_JSON")
    if service_account_json:
        try:
            temp_json_path = Path(tempfile.gettempdir()) / "service_account.json"
            temp_json_path.write_text(
                json.dumps(json.loads(service_account_json)),
                encoding="utf-8",
            )
        except Exception as e:
            raise RuntimeError("Invalid SERVICE_ACCOUNT_JSON; expected a JSON object string") from e

    env_path = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS") or os.environ.get("SERVICE_ACCOUNT_FILE")

    workspace_root = Path(__file__).resolve().parents[3]
    possible_paths = [
        temp_json_path,
        Path(env_path) if env_path else None,
        Path(__file__).parent / "service-account.json",
        Path(__file__).parent.parent / "service-account.json",
        Path.home() / ".config" / "gspread" / "service_account.json",
    ]
    
    for path in [p for p in possible_paths if p is not None]:
        if path.exists():
            logger.info(f"Using service account: {path}")
            client = gspread.service_account(filename=str(path))
            return client.open_by_key(sheet_id)
    
    # Try default gspread location
    try:
        client = gspread.service_account()
        return client.open_by_key(sheet_id)
    except Exception as e:
        raise FileNotFoundError(
            f"Could not find service account credentials. "
            "Set SERVICE_ACCOUNT_JSON or GOOGLE_APPLICATION_CREDENTIALS/SERVICE_ACCOUNT_FILE, "
            f"or place service-account.json in one of: {[str(p) for p in possible_paths]}"
        )


def select_folder() -> Optional[Path]:
    """Show folder selection dialog."""
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    folder_path = filedialog.askdirectory(
        title="Select folder containing bank statement PDFs",
        initialdir=Path(__file__).parent
    )
    
    root.destroy()
    
    if folder_path:
        return Path(folder_path)
    return None


def find_pdf_files(folder: Path) -> List[Path]:
    """Find all PDF files in the folder."""
    return list(folder.glob("*.pdf"))


def find_excel_files(folder: Path) -> List[Path]:
    """Find all Excel files in the folder (Bank_Rec*.xlsx pattern)."""
    files = list(folder.glob("Bank_Rec*.xlsx")) + list(folder.glob("*Bank_Rec*.xlsx"))
    # De-dupe while preserving order (patterns can overlap)
    seen = set()
    unique_files: List[Path] = []
    for f in files:
        if f not in seen:
            seen.add(f)
            unique_files.append(f)
    return unique_files


def extract_yardi_from_excel(excel_path: Path) -> List[List]:
    """
    Extract Outstanding Checks and Other Items from Bank Rec Excel file.
    
    Returns list of rows in format: [Property, Date, Transaction ID, Description, Amount]
    """
    from openpyxl import load_workbook

    def _excel_cell_to_text(cell) -> str:
        """Best-effort conversion to the same *visible* text Excel shows.

        Primarily used to keep IDs (like check numbers) stable:
        - Avoids converting integer-like floats to strings like '123.0'
        - Respects simple zero-padding formats like '00000'
        """
        if cell is None:
            return ""

        value = getattr(cell, "value", None)
        if value is None:
            return ""

        # If it's already text in Excel, preserve it.
        if isinstance(value, str):
            return value.strip()

        # Datetimes/dates should not be used for IDs; fall back to string.
        if hasattr(value, "strftime"):
            return value.strftime("%m/%d/%Y")

        number_format = (getattr(cell, "number_format", None) or "").strip()

        # Common case: Excel stored numeric, openpyxl returns float.
        if isinstance(value, float) and value.is_integer():
            value = int(value)

        if isinstance(value, int):
            text = str(value)
            # Handle simple formats like '00000' (leading zeros).
            if number_format and set(number_format) == {"0"}:
                width = len(number_format)
                text = text.zfill(width)
            return text

        return str(value).strip()

    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    assert ws is not None
    
    def _title_case_property(name: str) -> str:
        name = " ".join(name.split()).strip()
        if not name:
            return ""
        return name.title()

    transactions: List[List] = []

    # Property name should come from the Excel report content.
    # Example header: "SS of Madison Notre D operatin"
    property_name = ""
    header_search_cells: List[str] = []
    for r in ws.iter_rows(min_row=1, max_row=15, values_only=True):
        for v in r:
            if isinstance(v, str) and v.strip():
                header_search_cells.append(v.strip())
    header_blob = " \n".join(header_search_cells)

    # Prefer patterns that end before "Notre" when present.
    import re

    m = re.search(r"\bSS\s+of\s+(.+?)\s+Notre\b", header_blob, flags=re.IGNORECASE)
    if not m:
        m = re.search(r"\bSS\s+of\s+(.+?)\s+Bank\b", header_blob, flags=re.IGNORECASE)
    if not m:
        m = re.search(r"\bSS\s+of\s+([A-Za-z][A-Za-z ]+)\b", header_blob, flags=re.IGNORECASE)

    if m:
        property_name = _title_case_property(m.group(1))

    # Fallback: Extract property name from filename if present (e.g., Bank_Rec_Madison.xlsx)
    if not property_name and "_" in excel_path.stem:
        parts = excel_path.stem.split("_")
        if len(parts) >= 3 and parts[2].strip():
            property_name = _title_case_property(parts[2])
    
    current_section = None
    
    for row in ws.iter_rows():
        # Get cell values for section detection
        cells = [cell.value for cell in row]
        
        # Detect section headers
        if "Outstanding Checks" in str(cells):
            current_section = "checks"
            continue
        elif "Other Items" in str(cells):
            current_section = "other"
            continue
        elif "Less:" in str(cells) or "Plus/Minus:" in str(cells):
            # End of section marker
            continue
        
        if current_section == "checks":
            # Outstanding Checks format: [None, None, Date, None, None, Check#, None, Payee, Amount, ...]
            # Column C (index 2) = Date, Column F (index 5) = Check#, Column H (index 7) = Payee, Column I (index 8) = Amount
            date_val = cells[2] if len(cells) > 2 else None
            check_cell = row[5] if len(row) > 5 else None
            check_num = _excel_cell_to_text(check_cell)
            payee = cells[7] if len(cells) > 7 else None
            amount = cells[8] if len(cells) > 8 else None
            
            if date_val and amount and check_num:
                # Skip header row
                if str(check_num).strip().lower() == "check number":
                    continue
                    
                check_num_str = str(check_num).strip()
                # Requirement: Yardi Transaction ID must match Excel exactly (no prefix)
                trans_id = check_num_str
                # Requirement: Description must match Excel exactly
                description = str(payee).strip() if payee is not None else ""
                
                # Format date
                if isinstance(date_val, (datetime, date)):
                    date_str = date_val.strftime("%m/%d/%Y")
                else:
                    date_str = str(date_val)
                
                # Amount should be negative for checks (outflow)
                try:
                    amt = -abs(float(amount))
                except (ValueError, TypeError):
                    continue
                    
                transactions.append([property_name, date_str, trans_id, description, amt])
        
        elif current_section == "other":
            # Other Items format: [None, None, Date, None, None, '', None, Notes, Amount, ...]
            date_val = cells[2] if len(cells) > 2 else None
            notes = cells[7] if len(cells) > 7 else None
            amount = cells[8] if len(cells) > 8 else None
            
            if date_val and amount and notes:
                # Skip header row
                if str(notes).strip().lower() == "notes":
                    continue
                
                # Format date
                if isinstance(date_val, (datetime, date)):
                    date_str = date_val.strftime("%m/%d/%Y")
                else:
                    date_str = str(date_val)
                
                # Requirement: Other Items have no transaction id
                trans_id = ""
                # Requirement: Description must match Excel exactly
                description = str(notes).strip()
                
                try:
                    amt = float(amount)  # Keep sign as-is for Other Items
                except (ValueError, TypeError):
                    continue
                    
                transactions.append([property_name, date_str, trans_id, description, amt])
    
    wb.close()
    return transactions


def populate_yardi_sheet(spreadsheet: gspread.Spreadsheet, transactions: List[List]):
    """
    Populate the Yardi sheet with extracted transactions from Excel.
    
    Args:
        spreadsheet: Google Spreadsheet object
        transactions: List of transaction rows [Property, Date, Transaction ID, Description, Amount]
    """
    try:
        yardi_sheet = spreadsheet.worksheet(YARDI_SHEET_NAME)
    except gspread.WorksheetNotFound:
        logger.error(f"Sheet '{YARDI_SHEET_NAME}' not found. Please create it first.")
        return
    
    # Clear existing data (except header)
    yardi_sheet.batch_clear(["A2:E10000"])
    
    if not transactions:
        logger.warning("No Yardi transactions to populate")
        return
    
    # Sort by date
    def parse_date(row):
        try:
            return datetime.strptime(row[1], "%m/%d/%Y")
        except:
            return datetime.min
    
    transactions.sort(key=parse_date)
    
    def _force_text(value: str) -> str:
        value = (value or "").strip()
        return f"'{value}" if value else ""

    # Update sheet
    # Use USER_ENTERED so dates become real dates and amounts become real numbers.
    # Force Transaction ID to text to preserve leading zeros.
    rows = []
    for r in transactions:
        prop = r[0] if len(r) > 0 else ""
        date_val = r[1] if len(r) > 1 else ""
        txid = r[2] if len(r) > 2 else ""
        desc = r[3] if len(r) > 3 else ""
        amt = r[4] if len(r) > 4 else ""
        rows.append([prop, date_val, _force_text(str(txid)), desc, amt])

    if rows:
        yardi_sheet.update(
            values=rows,
            range_name=f"A2:E{len(rows) + 1}",
            value_input_option="USER_ENTERED",
        )

        # Explicit formats to prevent date/amount being treated as text.
        # B = Date, C = Transaction ID (text), E = Amount
        yardi_sheet.format("B2:B", {"numberFormat": {"type": "DATE", "pattern": "MM/dd/yyyy"}})
        yardi_sheet.format("C2:C", {"numberFormat": {"type": "TEXT"}})
        yardi_sheet.format("E2:E", {"numberFormat": {"type": "NUMBER", "pattern": "0.00"}})

        logger.info(f"Populated {len(rows)} transactions to '{YARDI_SHEET_NAME}' sheet")


def extract_property_name(folder: Path) -> str:
    """
    Extract property name from folder path.
    
    Convention: Folder names like "Madison", "Chicago", etc.
    Or parent folder if in monthly subfolder (e.g., "11. Nov")
    """
    folder_name = folder.name
    
    # If in a monthly subfolder, use parent folder name
    if folder_name.lower().startswith(("01.", "02.", "03.", "04.", "05.", "06.",
                                        "07.", "08.", "09.", "10.", "11.", "12.")):
        return folder.parent.name
    
    return folder_name


def extract_transactions_from_pdfs(pdf_files: List[Path], property_name: str) -> List[Transaction]:
    """
    Extract transactions from all PDF files.
    
    Args:
        pdf_files: List of PDF file paths
        property_name: Property name to assign to transactions
        
    Returns:
        List of all extracted transactions
    """
    all_transactions = []
    
    for pdf_path in pdf_files:
        logger.info(f"Processing: {pdf_path.name}")
        
        parser = get_parser_for_pdf(pdf_path)
        if parser is None:
            logger.warning(f"No parser found for: {pdf_path.name}")
            continue
        
        logger.info(f"Using parser: {parser.BANK_NAME}")
        
        try:
            transactions = parser.parse(pdf_path)
            logger.info(f"Extracted {len(transactions)} transactions from {pdf_path.name}")
            all_transactions.extend(transactions)
        except Exception as e:
            logger.error(f"Error parsing {pdf_path.name}: {e}")
    
    return all_transactions


def _generate_colors(count: int) -> List[Dict[str, float]]:
    """Generate visually distinct pastel colors using HSV color space.
    
    Args:
        count: Number of colors to generate
        
    Returns:
        List of color dicts with red, green, blue values (0-1 range)
    """
    import colorsys
    
    colors = []
    for i in range(count):
        # Distribute hues evenly across color spectrum
        hue = i / count
        
        # Use moderate saturation and high value for pastels (good with black text)
        saturation = 0.45 + (i % 3) * 0.15  # 0.45, 0.60, 0.75
        value = 0.85 + (i % 2) * 0.1  # 0.85, 0.95
        
        # Convert HSV to RGB
        r, g, b = colorsys.hsv_to_rgb(hue, saturation, value)
        
        # Google Sheets API color format (values 0-1)
        colors.append({
            "red": r,
            "green": g,
            "blue": b
        })
    
    return colors


def _apply_pass7_colors(worksheet, suggested_data: List[Tuple]) -> None:
    """Apply color highlighting to PASS 7 suggested matches in Unmatched sheet.
    
    Args:
        worksheet: gspread worksheet object
        suggested_data: List of (date, side, idx, row_data) tuples for suggested matches
    """
    if not suggested_data:
        return
    
    # Group suggested matches by (property, amount) to assign same color
    # property is in row_data[0] or row_data[5] depending on side
    # amount is in row_data[4] or row_data[9] depending on side
    groups: Dict[Tuple[str, float], List[int]] = {}  # (property, amount) -> [row_indices]
    
    for i, (date, side, idx, row_data) in enumerate(suggested_data):
        if side == "yardi":
            prop = row_data[0]  # Column A
            amt = row_data[4]   # Column E
        else:  # bank
            prop = row_data[5]  # Column F
            amt = row_data[9]   # Column J
        
        key = (prop, amt)
        row_number = 2 + i  # 1-indexed + header row
        if key not in groups:
            groups[key] = []
        groups[key].append(row_number)
    
    # Generate colors for each group
    colors = _generate_colors(len(groups))
    
    # Build batch requests for color formatting
    batch_requests = []
    for color_idx, (key, row_numbers) in enumerate(groups.items()):
        color = colors[color_idx]
        
        for row_number in row_numbers:
            # Color the entire row (columns A-J)
            batch_requests.append({
                "repeatCell": {
                    "range": {
                        "sheetId": worksheet.id,
                        "startRowIndex": row_number - 1,  # 0-indexed
                        "endRowIndex": row_number,
                        "startColumnIndex": 0,  # Column A
                        "endColumnIndex": 10  # Column J (exclusive)
                    },
                    "cell": {
                        "userEnteredFormat": {
                            "backgroundColor": color
                        }
                    },
                    "fields": "userEnteredFormat.backgroundColor"
                }
            })
    
    # Execute batch update if there are requests
    if batch_requests:
        worksheet.spreadsheet.batch_update({"requests": batch_requests})


def extract_property_name_from_pdf(pdf_path: Path) -> Optional[str]:
    """Extract property name from a bank statement PDF (best-effort).

    For Notre Dame FCU statements, the header often contains a line like:
    "STORSAFE OF MADISON LLC" -> "Madison".
    """
    try:
        text = BaseBankParser.extract_pdf_text(pdf_path).replace("\xa0", " ")
    except Exception:
        return None

    # Parser-specific extraction (currently Notre Dame FCU)
    try:
        import re

        m = re.search(r"\bSTORSAFE\s+OF\s+([A-Z][A-Z ]+?)\s+LLC\b", text, flags=re.IGNORECASE)
        if m:
            return " ".join(m.group(1).split()).title()
    except Exception:
        return None

    return None


def populate_bank_sheet(spreadsheet: gspread.Spreadsheet, transactions: List[Transaction], property_name: str):
    """
    Populate the Bank sheet with extracted transactions.
    
    Args:
        spreadsheet: Google Spreadsheet object
        transactions: List of transactions to populate
        property_name: Property name for the Property column
    """
    try:
        bank_sheet = spreadsheet.worksheet(BANK_SHEET_NAME)
    except gspread.WorksheetNotFound:
        logger.error(f"Sheet '{BANK_SHEET_NAME}' not found. Please create it first.")
        return
    
    # Clear existing data (except header)
    bank_sheet.batch_clear(["A2:E10000"])
    
    if not transactions:
        logger.warning("No transactions to populate")
        return
    
    def _force_text(value: str) -> str:
        value = (value or "").strip()
        return f"'{value}" if value else ""

    # Convert transactions to rows
    # Use USER_ENTERED so dates become real dates and amounts become real numbers.
    # Force Transaction ID to text to preserve leading zeros.
    rows = []
    for t in transactions:
        date_str = t.date.strftime("%m/%d/%Y")
        rows.append([
            property_name,
            date_str,
            _force_text(t.transaction_id),
            t.description,
            t.amount,
        ])
    
    # Sort by date
    rows.sort(key=lambda r: datetime.strptime(r[1], "%m/%d/%Y"))
    
    # Update sheet
    if rows:
        bank_sheet.update(
            values=rows,
            range_name=f"A2:E{len(rows) + 1}",
            value_input_option="USER_ENTERED",
        )

        # Explicit formats to prevent date/amount being treated as text.
        # B = Date, C = Transaction ID (text), E = Amount
        bank_sheet.format("B2:B", {"numberFormat": {"type": "DATE", "pattern": "MM/dd/yyyy"}})
        bank_sheet.format("C2:C", {"numberFormat": {"type": "TEXT"}})
        bank_sheet.format("E2:E", {"numberFormat": {"type": "NUMBER", "pattern": "0.00"}})

        logger.info(f"Populated {len(rows)} transactions to '{BANK_SHEET_NAME}' sheet")


def run_matching(spreadsheet: gspread.Spreadsheet) -> Tuple[int, int, int]:
    """
    Run transaction matching between Bank and Yardi sheets.
    
    Returns:
        Tuple of (matched_count, unmatched_bank_count, unmatched_yardi_count)
    """
    def _get_or_create_ws(title: str, rows: int = 2000, cols: int = 15) -> gspread.Worksheet:
        try:
            return spreadsheet.worksheet(title)
        except gspread.WorksheetNotFound:
            logger.info(f"Creating sheet '{title}'")
            return spreadsheet.add_worksheet(title=title, rows=rows, cols=cols)

    try:
        bank_sheet = spreadsheet.worksheet(BANK_SHEET_NAME)
        yardi_sheet = spreadsheet.worksheet(YARDI_SHEET_NAME)
    except gspread.WorksheetNotFound as e:
        logger.error(f"Sheet not found: {e}")
        return 0, 0
    
    # Get data from both sheets
    bank_data = bank_sheet.get_all_values()
    yardi_data = yardi_sheet.get_all_values()
    
    # Skip headers
    bank_rows = bank_data[1:] if len(bank_data) > 1 else []
    yardi_rows = yardi_data[1:] if len(yardi_data) > 1 else []
    
    if not bank_rows or not yardi_rows:
        logger.info("One or both sheets are empty. Skipping matching.")
        return 0, len(bank_rows)
    
    def _parse_date(val: str) -> Optional[datetime]:
        try:
            return datetime.strptime(val.strip(), "%m/%d/%Y")
        except Exception:
            return None

    def _parse_amount(val: str) -> Optional[float]:
        try:
            cleaned = val.replace(",", "").replace("$", "").strip()
            if cleaned == "":
                return None
            return float(cleaned)
        except Exception:
            return None

    def _cents(amount: float) -> int:
        # Avoid float drift
        return int(round(amount * 100))

    # Normalize sheet data
    # Columns: [Property, Date, Transaction ID, Description, Amount]
    bank_norm = []
    for idx, r in enumerate(bank_rows):
        prop = (r[0] if len(r) > 0 else "").strip()
        date_dt = _parse_date(r[1]) if len(r) > 1 else None
        txid = (r[2] if len(r) > 2 else "").strip()
        desc = (r[3] if len(r) > 3 else "").strip()
        amt = _parse_amount(r[4]) if len(r) > 4 else None
        bank_norm.append({
            "idx": idx,
            "prop": prop,
            "date": date_dt,
            "txid": txid,
            "desc": desc,
            "amt": amt,
        })

    yardi_norm = []
    for idx, r in enumerate(yardi_rows):
        prop = (r[0] if len(r) > 0 else "").strip()
        date_dt = _parse_date(r[1]) if len(r) > 1 else None
        txid = (r[2] if len(r) > 2 else "").strip()
        desc = (r[3] if len(r) > 3 else "").strip()
        amt = _parse_amount(r[4]) if len(r) > 4 else None
        yardi_norm.append({
            "idx": idx,
            "prop": prop,
            "date": date_dt,
            "txid": txid,
            "desc": desc,
            "amt": amt,
        })

    # Build yardi indices
    yardi_unmatched = set(range(len(yardi_norm)))

    def _prop_key(p: str) -> str:
        return " ".join(p.split()).strip().lower()

    yardi_by_txid: Dict[Tuple[str, str], List[int]] = {}
    for y in yardi_norm:
        if y["txid"]:
            yardi_by_txid.setdefault((_prop_key(y["prop"]), y["txid"]), []).append(y["idx"])

    # For pass-2 matching: (property, cents, date) buckets
    yardi_by_amount_date: Dict[Tuple[str, int, datetime], List[int]] = {}
    for y in yardi_norm:
        if y["amt"] is None or y["date"] is None:
            continue
        key = (
            _prop_key(y["prop"]),
            _cents(y["amt"]),
            y["date"].replace(hour=0, minute=0, second=0, microsecond=0),
        )
        yardi_by_amount_date.setdefault(key, []).append(y["idx"])

    # For pass-3 matching: (property, cents) buckets
    yardi_by_amount: Dict[Tuple[str, int], List[int]] = {}
    for y in yardi_norm:
        if y["amt"] is None:
            continue
        yardi_by_amount.setdefault((_prop_key(y["prop"]), _cents(y["amt"])), []).append(y["idx"])

    # Match results: bank_idx -> yardi_idx, plus pass label
    matches: Dict[int, int] = {}
    match_pass: Dict[int, str] = {}

    # Pass 1: Property + Transaction ID + Amount (strict 1:1)
    for b in bank_norm:
        if not b["txid"] or b["amt"] is None:
            continue

        candidates = yardi_by_txid.get((_prop_key(b["prop"]), b["txid"]), [])
        filtered: List[int] = []
        for y_idx in candidates:
            if y_idx not in yardi_unmatched:
                continue
            y = yardi_norm[y_idx]
            if y["amt"] is None:
                continue
            if _cents(y["amt"]) != _cents(b["amt"]):
                continue
            filtered.append(y_idx)

        if not filtered:
            continue

        if len(filtered) == 1:
            chosen = filtered[0]
            matches[b["idx"]] = chosen
            match_pass[b["idx"]] = "PASS 1"
            yardi_unmatched.remove(chosen)

    # Pass 2: Property + Date + Amount (strict 1:1)
    for b in bank_norm:
        if b["idx"] in matches:
            continue
        if b["amt"] is None or b["date"] is None:
            continue

        b_key = (
            _prop_key(b["prop"]),
            _cents(b["amt"]),
            b["date"].replace(hour=0, minute=0, second=0, microsecond=0),
        )
        candidate_ids = yardi_by_amount_date.get(b_key, [])
        filtered = [y_idx for y_idx in candidate_ids if y_idx in yardi_unmatched]

        if len(filtered) == 1:
            chosen = filtered[0]
            matches[b["idx"]] = chosen
            match_pass[b["idx"]] = "PASS 2"
            yardi_unmatched.remove(chosen)

    # Pass 3: Property + Amount + Date within +/- 3 days (strict 1:1)
    for b in bank_norm:
        if b["idx"] in matches:
            continue
        if b["amt"] is None or b["date"] is None:
            continue

        candidates = yardi_by_amount.get((_prop_key(b["prop"]), _cents(b["amt"])), [])
        filtered: List[int] = []
        for y_idx in candidates:
            if y_idx not in yardi_unmatched:
                continue
            y = yardi_norm[y_idx]
            if y["date"] is None:
                continue
            if abs((b["date"] - y["date"]).days) <= 3:
                filtered.append(y_idx)

        if len(filtered) == 1:
            chosen = filtered[0]
            matches[b["idx"]] = chosen
            match_pass[b["idx"]] = "PASS 3"
            yardi_unmatched.remove(chosen)

    # Pass 4: Property + Amount + Date within +/- 7 days (strict 1:1)
    for b in bank_norm:
        if b["idx"] in matches:
            continue
        if b["amt"] is None or b["date"] is None:
            continue

        candidates = yardi_by_amount.get((_prop_key(b["prop"]), _cents(b["amt"])), [])
        filtered: List[int] = []
        for y_idx in candidates:
            if y_idx not in yardi_unmatched:
                continue
            y = yardi_norm[y_idx]
            if y["date"] is None:
                continue
            if abs((b["date"] - y["date"]).days) <= 7:
                filtered.append(y_idx)

        if len(filtered) == 1:
            chosen = filtered[0]
            matches[b["idx"]] = chosen
            match_pass[b["idx"]] = "PASS 4"
            yardi_unmatched.remove(chosen)

    # Pass 5: Property + Side (same side: Bank↔Bank or Yardi↔Yardi) + Date + Amount (same absolute value, opposite signs)
    # This catches reversals/corrections on the same side (either Bank or Yardi)
    # Track these separately since they're not Bank↔Yardi matches
    
    pass5_bank_matches: List[Tuple[int, int]] = []  # List of (bank_idx1, bank_idx2) pairs
    pass5_yardi_matches: List[Tuple[int, int]] = []  # List of (yardi_idx1, yardi_idx2) pairs
    
    # Build index for Bank-to-Bank matching (opposite sign, same date)
    bank_by_prop_date_abs_amt: Dict[Tuple[str, datetime, int], List[int]] = {}
    for b in bank_norm:
        if b["idx"] in matches or b["amt"] is None or b["date"] is None:
            continue
        key = (
            _prop_key(b["prop"]),
            b["date"].replace(hour=0, minute=0, second=0, microsecond=0),
            abs(_cents(b["amt"]))
        )
        bank_by_prop_date_abs_amt.setdefault(key, []).append(b["idx"])

    # Match Bank entries with opposite signs (1:1)
    matched_bank_pass5 = set()
    for key, bank_idxs in bank_by_prop_date_abs_amt.items():
        if len(bank_idxs) < 2:
            continue
        
        # Group by sign
        positive = [idx for idx in bank_idxs if bank_norm[idx]["amt"] > 0 and idx not in matched_bank_pass5]
        negative = [idx for idx in bank_idxs if bank_norm[idx]["amt"] < 0 and idx not in matched_bank_pass5]
        
        # Match 1:1 between positive and negative
        for pos_idx, neg_idx in zip(positive, negative):
            pass5_bank_matches.append((neg_idx, pos_idx))  # Store as (negative, positive) tuple
            matched_bank_pass5.add(pos_idx)
            matched_bank_pass5.add(neg_idx)

    # Build index for Yardi-to-Yardi matching (opposite sign, same date)
    yardi_by_prop_date_abs_amt: Dict[Tuple[str, datetime, int], List[int]] = {}
    for y in yardi_norm:
        if y["idx"] not in yardi_unmatched or y["amt"] is None or y["date"] is None:
            continue
        key = (
            _prop_key(y["prop"]),
            y["date"].replace(hour=0, minute=0, second=0, microsecond=0),
            abs(_cents(y["amt"]))
        )
        yardi_by_prop_date_abs_amt.setdefault(key, []).append(y["idx"])

    # Match Yardi entries with opposite signs (1:1)
    matched_yardi_pass5 = set()
    for key, yardi_idxs in yardi_by_prop_date_abs_amt.items():
        if len(yardi_idxs) < 2:
            continue
        
        # Group by sign
        positive = [idx for idx in yardi_idxs if yardi_norm[idx]["amt"] > 0 and idx not in matched_yardi_pass5]
        negative = [idx for idx in yardi_idxs if yardi_norm[idx]["amt"] < 0 and idx not in matched_yardi_pass5]
        
        # Match 1:1 between positive and negative
        for pos_idx, neg_idx in zip(positive, negative):
            pass5_yardi_matches.append((neg_idx, pos_idx))  # Store as (negative, positive) tuple
            matched_yardi_pass5.add(pos_idx)
            matched_yardi_pass5.add(neg_idx)
            yardi_unmatched.discard(pos_idx)
            yardi_unmatched.discard(neg_idx)

    # Pass 6: Property + Amount (exact) - but only if ALL have perfect pairs
    # Group remaining unmatched by (property, amount)
    bank_by_prop_amt: Dict[Tuple[str, int], List[int]] = {}
    for b in bank_norm:
        if b["idx"] in matches or b["amt"] is None:
            continue
        key = (_prop_key(b["prop"]), _cents(b["amt"]))
        bank_by_prop_amt.setdefault(key, []).append(b["idx"])

    yardi_by_prop_amt: Dict[Tuple[str, int], List[int]] = {}
    for y in yardi_norm:
        if y["idx"] not in yardi_unmatched or y["amt"] is None:
            continue
        key = (_prop_key(y["prop"]), _cents(y["amt"]))
        yardi_by_prop_amt.setdefault(key, []).append(y["idx"])

    # Find keys that exist in both sides
    common_keys = set(bank_by_prop_amt.keys()) & set(yardi_by_prop_amt.keys())

    pass6_matches: Dict[int, int] = {}
    for key in common_keys:
        bank_idxs = bank_by_prop_amt[key]
        yardi_idxs = yardi_by_prop_amt[key]

        # Only match if counts are equal (everyone has a partner)
        if len(bank_idxs) == len(yardi_idxs):
            # Pair them up (1:1 in order)
            for b_idx, y_idx in zip(bank_idxs, yardi_idxs):
                pass6_matches[b_idx] = y_idx

    # Apply PASS 6 matches
    for b_idx, y_idx in pass6_matches.items():
        matches[b_idx] = y_idx
        match_pass[b_idx] = "PASS 6"
        yardi_unmatched.remove(y_idx)

    # Pass 7: Property + Amount (exact) but incomplete pairing → suggested matches
    # These stay in Unmatched tab but get a "Suggested Match" flag
    suggested_matches: Dict[int, List[int]] = {}  # bank_idx -> list of yardi_idx candidates
    suggested_yardi: Dict[int, List[int]] = {}    # yardi_idx -> list of bank_idx candidates

    for key in common_keys:
        bank_idxs = [b for b in bank_by_prop_amt[key] if b not in matches]
        yardi_idxs = [y for y in yardi_by_prop_amt[key] if y in yardi_unmatched]

        # Only suggest if there's at least one unpaired (PASS 6 didn't match them)
        if bank_idxs and yardi_idxs and (len(bank_idxs) != len(yardi_idxs)):
            for b_idx in bank_idxs:
                suggested_matches[b_idx] = yardi_idxs
            for y_idx in yardi_idxs:
                suggested_yardi[y_idx] = bank_idxs

    matched = len(matches) + len(pass5_bank_matches) + len(pass5_yardi_matches)
    unmatched_bank = len(bank_norm) - len(matches) - len(matched_bank_pass5)
    unmatched_yardi = len(yardi_unmatched)

    logger.info(f"Bank transactions: {len(bank_norm)}")
    logger.info(f"Yardi transactions: {len(yardi_norm)}")
    logger.info(f"Matched (Bank↔Yardi): {len(matches)}; PASS 5 (Bank↔Bank): {len(pass5_bank_matches)}; PASS 5 (Yardi↔Yardi): {len(pass5_yardi_matches)}")
    logger.info(f"Unmatched bank: {unmatched_bank}; Unmatched yardi: {unmatched_yardi}")
    logger.info(f"Suggested matches (PASS 7): {len(suggested_matches)} bank + {len(suggested_yardi)} yardi")

    # Populate Matched / Unmatched tabs (primary output for matching workflow)
    matched_ws = _get_or_create_ws(MATCHED_SHEET_NAME, rows=max(2000, matched + 10), cols=15)
    unmatched_ws = _get_or_create_ws(UNMATCHED_SHEET_NAME, rows=max(2000, (len(bank_norm) + len(yardi_norm)) + 10), cols=10)

    matched_ws.clear()
    unmatched_ws.clear()

    def _force_text(value: str) -> str:
        value = (value or "").strip()
        return f"'{value}" if value else ""

    matched_header = [
        "Yardi Property",
        "Yardi Date",
        "Yardi Transaction ID",
        "Yardi Description",
        "Yardi Amount",
        "Bank Property",
        "Bank Date",
        "Bank Transaction ID",
        "Bank Description",
        "Bank Amount",
        "Match Pass",
    ]
    unmatched_header = [
        "Yardi Property",
        "Yardi Date",
        "Yardi Transaction ID",
        "Yardi Description",
        "Yardi Amount",
        "Bank Property",
        "Bank Date",
        "Bank Transaction ID",
        "Bank Description",
        "Bank Amount",
    ]

    matched_rows: List[List] = []

    def _append_matched_row(b: Dict):
        y = yardi_norm[matches[b["idx"]]]
        matched_rows.append([
            y["prop"],
            y["date"].strftime("%m/%d/%Y") if y["date"] else "",
            _force_text(y["txid"]),
            y["desc"],
            y["amt"] if y["amt"] is not None else "",
            b["prop"],
            b["date"].strftime("%m/%d/%Y") if b["date"] else "",
            _force_text(b["txid"]),
            b["desc"],
            b["amt"] if b["amt"] is not None else "",
            match_pass.get(b["idx"], ""),
        ])

    # Order matched rows by pass group
    # PASS 5 is special: same-side matches (Bank↔Bank or Yardi↔Yardi)
    # Other passes: Bank↔Yardi matches
    for pass_label in ("PASS 1", "PASS 2", "PASS 3", "PASS 4"):
        for b in bank_norm:
            if b["idx"] not in matches:
                continue
            if match_pass.get(b["idx"], "") != pass_label:
                continue
            _append_matched_row(b)
    
    # Add PASS 5 Bank↔Bank matches
    for bank_idx1, bank_idx2 in pass5_bank_matches:
        b1 = bank_norm[bank_idx1]
        b2 = bank_norm[bank_idx2]
        matched_rows.append([
            "",  # No Yardi property
            "",  # No Yardi date
            "",  # No Yardi txid
            "",  # No Yardi desc
            "",  # No Yardi amount
            b1["prop"],
            b1["date"].strftime("%m/%d/%Y") if b1["date"] else "",
            _force_text(b1["txid"]),
            b1["desc"],
            b1["amt"] if b1["amt"] is not None else "",
            "PASS 5",
        ])
        matched_rows.append([
            "",
            "",
            "",
            "",
            "",
            b2["prop"],
            b2["date"].strftime("%m/%d/%Y") if b2["date"] else "",
            _force_text(b2["txid"]),
            b2["desc"],
            b2["amt"] if b2["amt"] is not None else "",
            "PASS 5",
        ])
    
    # Add PASS 5 Yardi↔Yardi matches
    for yardi_idx1, yardi_idx2 in pass5_yardi_matches:
        y1 = yardi_norm[yardi_idx1]
        y2 = yardi_norm[yardi_idx2]
        matched_rows.append([
            y1["prop"],
            y1["date"].strftime("%m/%d/%Y") if y1["date"] else "",
            _force_text(y1["txid"]),
            y1["desc"],
            y1["amt"] if y1["amt"] is not None else "",
            "",  # No Bank property
            "",  # No Bank date
            "",  # No Bank txid
            "",  # No Bank desc
            "",  # No Bank amount
            "PASS 5",
        ])
        matched_rows.append([
            y2["prop"],
            y2["date"].strftime("%m/%d/%Y") if y2["date"] else "",
            _force_text(y2["txid"]),
            y2["desc"],
            y2["amt"] if y2["amt"] is not None else "",
            "",
            "",
            "",
            "",
            "",
            "PASS 5",
        ])
    
    # Add PASS 6 matches
    for pass_label in ("PASS 6",):
        for b in bank_norm:
            if b["idx"] not in matches:
                continue
            if match_pass.get(b["idx"], "") != pass_label:
                continue
            _append_matched_row(b)

    # Collect all unmatched entries with dates for sorting
    # Separate suggested (PASS_7) from non-suggested
    suggested_unmatched = []
    non_suggested_unmatched = []
    
    for y in yardi_norm:
        if y["idx"] not in yardi_unmatched:
            continue
        row_data = [
            y["prop"],
            y["date"].strftime("%m/%d/%Y") if y["date"] else "",
            _force_text(y["txid"]),
            y["desc"],
            y["amt"] if y["amt"] is not None else "",
            "",
            "",
            "",
            "",
            "",
        ]
        if y["idx"] in suggested_yardi:
            suggested_unmatched.append((y["date"] or datetime.min, "yardi", y["idx"], row_data))
        else:
            non_suggested_unmatched.append((y["date"] or datetime.min, "yardi", y["idx"], row_data))
    
    for b in bank_norm:
        if b["idx"] in matches or b["idx"] in matched_bank_pass5:
            continue
        row_data = [
            "",
            "",
            "",
            "",
            "",
            b["prop"],
            b["date"].strftime("%m/%d/%Y") if b["date"] else "",
            _force_text(b["txid"]),
            b["desc"],
            b["amt"] if b["amt"] is not None else "",
        ]
        if b["idx"] in suggested_matches:
            suggested_unmatched.append((b["date"] or datetime.min, "bank", b["idx"], row_data))
        else:
            non_suggested_unmatched.append((b["date"] or datetime.min, "bank", b["idx"], row_data))
    
    # Sort each group by date
    suggested_unmatched.sort(key=lambda x: x[0])
    non_suggested_unmatched.sort(key=lambda x: x[0])
    
    # Combine: suggested first, then non-suggested
    all_unmatched = suggested_unmatched + non_suggested_unmatched
    unmatched_rows = [row for _, _, _, row in all_unmatched]
    
    # Track row indices (2-based) for color highlighting
    suggested_row_indices = list(range(2, 2 + len(suggested_unmatched)))

    matched_ws.update(range_name="A1:K1", values=[matched_header], value_input_option="RAW")
    if matched_rows:
        matched_ws.update(
            range_name=f"A2:K{len(matched_rows) + 1}",
            values=matched_rows,
            value_input_option="USER_ENTERED",
        )

    unmatched_ws.update(range_name="A1:J1", values=[unmatched_header], value_input_option="RAW")
    if unmatched_rows:
        unmatched_ws.update(
            range_name=f"A2:J{len(unmatched_rows) + 1}",
            values=unmatched_rows,
            value_input_option="USER_ENTERED",
        )

    # Explicit formats to prevent date/amount being treated as text.
    # Matched: B/G dates, C/H text IDs, E/J amounts (Yardi A-E, Bank F-J)
    matched_ws.format("B2:B", {"numberFormat": {"type": "DATE", "pattern": "MM/dd/yyyy"}})
    matched_ws.format("G2:G", {"numberFormat": {"type": "DATE", "pattern": "MM/dd/yyyy"}})
    matched_ws.format("C2:C", {"numberFormat": {"type": "TEXT"}})
    matched_ws.format("H2:H", {"numberFormat": {"type": "TEXT"}})
    matched_ws.format("E2:E", {"numberFormat": {"type": "NUMBER", "pattern": "0.00"}})
    matched_ws.format("J2:J", {"numberFormat": {"type": "NUMBER", "pattern": "0.00"}})

    # Unmatched: B/G dates, C/H text IDs, E/J amounts (Yardi A-E, Bank F-J)
    unmatched_ws.format("B2:B", {"numberFormat": {"type": "DATE", "pattern": "MM/dd/yyyy"}})
    unmatched_ws.format("G2:G", {"numberFormat": {"type": "DATE", "pattern": "MM/dd/yyyy"}})
    unmatched_ws.format("C2:C", {"numberFormat": {"type": "TEXT"}})
    unmatched_ws.format("H2:H", {"numberFormat": {"type": "TEXT"}})
    unmatched_ws.format("E2:E", {"numberFormat": {"type": "NUMBER", "pattern": "0.00"}})
    unmatched_ws.format("J2:J", {"numberFormat": {"type": "NUMBER", "pattern": "0.00"}})
    
    # Apply color highlighting to PASS_7 suggested matches
    if suggested_unmatched:
        _apply_pass7_colors(unmatched_ws, suggested_unmatched)

    return matched, unmatched_bank, unmatched_yardi


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(add_help=True)
    parser.add_argument(
        "--mode",
        choices=["parse", "match", "all"],
        default="all",
        help="Run only parsing (parse), only matching (match), or both (all).",
    )
    parser.add_argument(
        "--sheet-id",
        default=DEFAULT_SHEET_ID,
        help="Google Sheet ID to use.",
    )
    parser.add_argument(
        "--folder",
        default=None,
        help="Folder containing bank statement PDFs / Bank_Rec Excel. If omitted, a folder picker is shown.",
    )
    args = parser.parse_args()

    print("=" * 60)
    print("Bank Reconciliation Automation")
    print("=" * 60)
    print()

    # Matching-only workflow: no local folder selection.
    if args.mode == "match":
        print("Running matching only (no parsing)...")
        print("Connecting to Google Sheets...")
        try:
            spreadsheet = connect_to_sheets(args.sheet_id)
            print(f"Connected to: {spreadsheet.title}")
        except Exception as e:
            print(f"Error connecting to Google Sheets: {e}")
            return

        print("Running transaction matching...")
        matched, unmatched_bank, unmatched_yardi = run_matching(spreadsheet)
        print(f"Matched: {matched}, Unmatched: {unmatched_bank + unmatched_yardi}")
        print(f"Populated tabs: {MATCHED_SHEET_NAME}, {UNMATCHED_SHEET_NAME}")
        return
    
    # Select folder (or use provided one)
    if args.folder:
        folder = Path(args.folder)
    else:
        folder = select_folder()
    if not folder:
        print("No folder selected. Exiting.")
        return

    if not folder.exists() or not folder.is_dir():
        print(f"Folder does not exist or is not a directory: {folder}")
        return
    
    print(f"Selected folder: {folder}")
    
    # Find PDF files
    pdf_files = find_pdf_files(folder)
    excel_files = find_excel_files(folder)
    
    if not pdf_files and not excel_files:
        print(f"No PDF or Excel files found in {folder}")
        return
    
    if pdf_files:
        print(f"Found {len(pdf_files)} PDF file(s)")
        for pdf in pdf_files:
            print(f"  - {pdf.name}")
    
    if excel_files:
        print(f"Found {len(excel_files)} Excel file(s)")
        for excel in excel_files:
            print(f"  - {excel.name}")
    print()
    
    # Extract property name
    # Bank sheet: derive from the bank statement itself when possible.
    property_name = extract_property_name(folder)
    if pdf_files:
        parsed_property = extract_property_name_from_pdf(pdf_files[0])
        if parsed_property:
            property_name = parsed_property
    print(f"Property: {property_name}")
    print()
    
    # Connect to Google Sheets
    print("Connecting to Google Sheets...")
    try:
        spreadsheet = connect_to_sheets(args.sheet_id)
        print(f"Connected to: {spreadsheet.title}")
    except Exception as e:
        print(f"Error connecting to Google Sheets: {e}")
        return
    print()
    
    # Extract and populate Bank transactions from PDFs
    if pdf_files:
        print("Extracting transactions from PDFs...")
        transactions = extract_transactions_from_pdfs(pdf_files, property_name)
        print(f"Total bank transactions extracted: {len(transactions)}")
        
        if transactions:
            print("\nBank Transaction Summary:")
            deposits = sum(1 for t in transactions if t.amount > 0)
            withdrawals = sum(1 for t in transactions if t.amount < 0)
            total_deposits = sum(t.amount for t in transactions if t.amount > 0)
            total_withdrawals = sum(t.amount for t in transactions if t.amount < 0)
            print(f"  Deposits: {deposits} (${total_deposits:,.2f})")
            print(f"  Withdrawals: {withdrawals} (${abs(total_withdrawals):,.2f})")
            print(f"  Net: ${sum(t.amount for t in transactions):,.2f}")
        print()
        
        print("Populating Bank sheet...")
        populate_bank_sheet(spreadsheet, transactions, property_name)
        print()
    
    # Extract and populate Yardi transactions from Excel
    if excel_files:
        print("Extracting transactions from Excel (Yardi data)...")
        all_yardi_transactions = []
        for excel_path in excel_files:
            yardi_txns = extract_yardi_from_excel(excel_path)
            print(f"  Extracted {len(yardi_txns)} transactions from {excel_path.name}")
            all_yardi_transactions.extend(yardi_txns)
        
        print(f"Total Yardi transactions extracted: {len(all_yardi_transactions)}")
        
        if all_yardi_transactions:
            print("\nYardi Transaction Summary:")
            checks = sum(1 for t in all_yardi_transactions if str(t[2]).strip() != "")
            other = len(all_yardi_transactions) - checks
            total_checks = sum(t[4] for t in all_yardi_transactions if str(t[2]).strip() != "")
            total_other = sum(t[4] for t in all_yardi_transactions if str(t[2]).strip() == "")
            print(f"  Outstanding Checks: {checks} (${abs(total_checks):,.2f})")
            print(f"  Other Items: {other} (${total_other:,.2f})")
        print()
        
        print("Populating Yardi sheet...")
        populate_yardi_sheet(spreadsheet, all_yardi_transactions)
        print()
    
    # Run matching (separate workflow)
    if args.mode == "all":
        print("Running transaction matching...")
        matched, unmatched_bank, unmatched_yardi = run_matching(spreadsheet)
        print(f"Matched: {matched}, Unmatched: {unmatched_bank + unmatched_yardi}")
        print(f"Populated tabs: {MATCHED_SHEET_NAME}, {UNMATCHED_SHEET_NAME}")
        print()
    
    print("=" * 60)
    print("Bank reconciliation complete!")
    print(f"View results at: https://docs.google.com/spreadsheets/d/{DEFAULT_SHEET_ID}")
    print("=" * 60)


if __name__ == "__main__":
    main()
